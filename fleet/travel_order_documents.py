import os
import re

from django.conf import settings
from django.http import JsonResponse
from openpyxl import load_workbook


def sanitize_filename(filename):
    """
    Uklanja nedozvoljene znakove iz naziva fajla.
    Dozvoljeni znakovi: slova, brojevi, crtice i donje crte.
    """
    return re.sub(r"[^a-zA-Z0-9_\-]", "_", filename)


def populate_putni_nalog_template(putni_nalog):
    """
    Popunjava Excel šablon sa podacima putnog naloga i vraća generisani fajl.
    """
    template_path = os.path.join(settings.BASE_DIR, "dokumenta", "iz077.xlsx")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Šablon ne postoji: {template_path}")

    workbook = load_workbook(template_path)

    if "zadnja strana" in workbook.sheetnames:
        sheet1 = workbook["zadnja strana"]
        sheet1["P1"] = str(putni_nalog.job_code.name)
        sheet1["N2"] = str(putni_nalog.order_number)
        sheet1["M3"] = str(putni_nalog.order_date.strftime("%d.%m.%Y"))
        sheet1["O6"] = str(putni_nalog.employee) if putni_nalog.employee else str(putni_nalog.other_employee_name or "")
        sheet1["M8"] = str(putni_nalog.employee.position) if putni_nalog.employee else ""
        sheet1["R9"] = putni_nalog.travel_date.strftime("%d.%m.%Y")
        sheet1["N10"] = putni_nalog.travel_location
        sheet1["M12"] = putni_nalog.task
        sheet1["M12"] = putni_nalog.contract_offer
        sheet1["M16"] = str(putni_nalog.vehicle)
        sheet1["S17"] = putni_nalog.daily_allowance
        sheet1["R18"] = putni_nalog.number_of_days
        sheet1["R22"] = float(putni_nalog.advance_payment)
        sheet1["R23"] = putni_nalog.job_code.code
    else:
        raise ValueError("Nema radnog lista 'zadnja strana' u šablonu.")

    if "prednja strana" in workbook.sheetnames:
        sheet2 = workbook["prednja strana"]
        sheet2["P1"] = str(putni_nalog.job_code.name)
        sheet2["N2"] = str(putni_nalog.order_number)
        sheet2["M3"] = str(putni_nalog.order_date.strftime("%d.%m.%Y"))
        sheet2["O6"] = str(putni_nalog.employee) if putni_nalog.employee else str(putni_nalog.other_employee_name or "")
        sheet2["M8"] = str(putni_nalog.employee.position) if putni_nalog.employee else ""
        sheet2["R9"] = putni_nalog.travel_date.strftime("%d.%m.%Y")
        sheet2["N10"] = putni_nalog.travel_location
        sheet2["M12"] = putni_nalog.task
        sheet1["M12"] = putni_nalog.contract_offer
        sheet2["M16"] = str(putni_nalog.vehicle)
        sheet2["S17"] = putni_nalog.daily_allowance
        sheet2["R18"] = putni_nalog.number_of_days
        sheet2["R22"] = float(putni_nalog.advance_payment)
        sheet2["R23"] = putni_nalog.job_code.code
    else:
        raise ValueError("Nema radnog lista 'prednja strana' u šablonu.")

    output_dir = os.path.join(settings.MEDIA_ROOT, "travel_orders")
    os.makedirs(output_dir, exist_ok=True)

    sanitized_order_number = sanitize_filename(putni_nalog.order_number)
    file_name = f"PutniNalog_{sanitized_order_number}.xlsx"
    file_path = os.path.join(output_dir, file_name)

    try:
        workbook.save(file_path)
    except Exception as exc:
        return JsonResponse({"error": f"Greška pri čuvanju fajla: {str(exc)}"}, status=500)

    if not os.path.exists(file_path):
        return JsonResponse({"error": f"Fajl nije pronađen nakon čuvanja: {file_path}"}, status=500)

    file_url = os.path.join(settings.MEDIA_URL, "travel_orders", file_name)
    return JsonResponse({"file_url": file_url})

from django.shortcuts import get_object_or_404, render, redirect
from django.db import connections, IntegrityError
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DetailView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from .models import *
from .forms import *
from .filters import *
from django.db.models import OuterRef, Subquery
from django_filters.views import FilterView
from django.contrib import messages
import logging
from django.http import HttpResponseForbidden
from django.db.models import Sum, Exists,Count,Avg
import datetime 
from datetime import date, timedelta
from django.db.models import F
from django.utils import timezone
from datetime import timedelta
from .utils import calculate_average_fuel_consumption, calculate_average_fuel_consumption_ever, update_vehicle_values, delete_complete_drafts, sanitize_filename
from django.db.models.functions import TruncMonth, TruncYear
from django.db.models import Q
from .utils import fetch_requisition_data, fetch_service_data, fetch_policy_data, populate_putni_nalog_template
from .models import DraftServiceTransaction
import threading
from .utils import migrate_draft_to_service_transaction, get_fuel_consumption_queryset
from django.shortcuts import render
from django.db import connection
from django.http import FileResponse
import os
from django.urls import reverse
from django.conf import settings
from django.http import HttpResponseRedirect
from openpyxl import Workbook
import pandas as pd
from .forms import PutnickaFilterForm

# <!-- ======================================================================= -->
#                           <!-- DASHBOARD I ANALITIKA -->
# <!-- ======================================================================= -->
def dashboard(request):    
    # Count the number of objects where vehicle is None
    services_without_vehicle = DraftServiceTransaction.objects.count()
    policies_without_vehicle = DraftPolicy.objects.count()
    requisitions_without_vehicle = DraftRequisition.objects.count()
    
    today = date.today()
    thirty_days_from_now = today + timedelta(days=30)

    # Subquery za pronalaženje najnovijeg datuma završetka za isti automobil i tip osiguranja
    newest_policy = Policy.objects.filter(
        vehicle=OuterRef('vehicle'),
        insurance_type=OuterRef('insurance_type'),
        is_renewable=True  # Dodato da se uzimaju u obzir samo polise koje se obnavljaju
    ).order_by('-end_date').values('end_date')[:1]

    # Filtriranje polisa koje uskoro ističu i koje se obnavljaju
    expiring_policies = Policy.objects.annotate(
        latest_end_date=Subquery(newest_policy)
    ).filter(
        end_date__gte=today,
        end_date__lte=thirty_days_from_now,
        end_date=F('latest_end_date'),
        is_renewable=True  # Dodato da filtrira samo polise koje se obnavljaju
    )

    expiring_policies_count = expiring_policies.count()

    # Subquery da proveri da li postoji nova polisa za isto vozilo i tip osiguranja
    newer_policy_exists = Policy.objects.filter(
        vehicle=OuterRef('vehicle'),
        insurance_type=OuterRef('insurance_type'),
        start_date__gt=OuterRef('start_date'),
        is_renewable=True  # Dodato da proveri samo polise koje se obnavljaju
    )

    # Filtriranje polisa koje su istekle i nisu obnovljene
    expired_unrenewed_policies = Policy.objects.annotate(
        has_newer_policy=Exists(newer_policy_exists)
    ).filter(
        end_date__lt=today,           # Polise koje su već istekle
        has_newer_policy=False,       # Proveri da li nema novije polise
        is_renewable=True             # Dodato da proveri samo polise koje se obnavljaju
    )

    expired_unrenewed_policies_count = expired_unrenewed_policies.count()

    # Current year and last day of the previous month
    current_year = datetime.datetime.now().year
    first_day_of_current_month = date.today().replace(day=1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)
    start_of_year = date.today().replace(month=1, day=1)

    # Number of vehicles
    total_vehicles = Vehicle.objects.filter(otpis=False).count()
    passenger_vehicles = Vehicle.objects.filter(category='PUTNICKO VOZILO').filter(otpis=False).count()
    transport_vehicles = Vehicle.objects.filter(category='TERETNO VOZILO').filter(otpis=False).count()



    # Average vehicle age
    average_age = Vehicle.objects.aggregate(avg_age=(current_year - Avg('year_of_manufacture')))

    # Book value as of the last day of the previous month
    book_value = Vehicle.objects.filter(purchase_date__lte=last_day_of_previous_month).aggregate(total_value=Sum('value'))

    # Yearly costs
    yearly_fuel_costs = FuelConsumption.objects.filter(date__gte=start_of_year).aggregate(total_fuel_cost=Sum('cost_bruto'))
    yearly_service_costs = ServiceTransaction.objects.filter(datum__gte=start_of_year).aggregate(total_service_cost=Sum('potrazuje'))

    # Vehicles in red zone
    red_zone_vehicles = Vehicle.objects.filter(otpis=True)  # or any other criteria
   
    # Podupit za poslednju dodelu jedinice po vozilu
    latest_jobcode = JobCode.objects.filter(
        vehicle=OuterRef('pk')
    ).order_by('-assigned_date')

    # Dohvati sva vozila sa podacima, bez grupisanja
    vehicles = Vehicle.objects.annotate(
        center_code=Subquery(latest_jobcode.values('organizational_unit__center')[:1]),
        manufacture_year=F('year_of_manufacture'),
        current_value=F('value'),  # promenjeno ime!
        fuel_amount=Subquery(
            FuelConsumption.objects.filter(vehicle=OuterRef('pk'))
            .values('vehicle')
            .annotate(total=Sum('amount'))
            .values('total')[:1]
        ),
        fuel_cost=Subquery(
            FuelConsumption.objects.filter(vehicle=OuterRef('pk'))
            .values('vehicle')
            .annotate(total=Sum('cost_bruto'))
            .values('total')[:1]
        )
    )
    # Grupisanje u Pythonu
    from collections import defaultdict
    grouped = defaultdict(list)

    for v in vehicles:
        grouped[v.center_code].append(v)

    # Napravi center_data ručno
    center_data = []
    for center, vehicle_list in grouped.items():
        count = len(vehicle_list)
        total_value = sum([v.current_value or 0 for v in vehicle_list])
        avg_value = total_value / count if count else 0
        total_fuel_quantity = sum([v.fuel_amount or 0 for v in vehicle_list])
        total_fuel_price = sum([v.fuel_cost or 0 for v in vehicle_list])
        avg_year = sum([v.year_of_manufacture or 0 for v in vehicle_list]) / count if count else 0
        avg_age = current_year - avg_year if avg_year else None

        center_data.append({
            'center_code': center,
            'vehicle_count': count,
            'avg_age': avg_age,
            'total_value': total_value,
            'avg_value': avg_value,
            'total_fuel_quantity': total_fuel_quantity,
            'total_fuel_price': total_fuel_price,
        })



    context = {
        'services_without_vehicle': services_without_vehicle,
        'policies_without_vehicle': policies_without_vehicle,
        'requisitions_without_vehicle': requisitions_without_vehicle,
        'expiring_policies': expiring_policies,
        'expiring_policies_count': expiring_policies_count,
        'expired_unrenewed_policies': expired_unrenewed_policies,
        'expired_unrenewed_policies_count': expired_unrenewed_policies_count,
        'total_vehicles': total_vehicles,
        'passenger_vehicles': passenger_vehicles,
        'transport_vehicles': transport_vehicles,
        # 'vehicles_by_center': vehicles_by_center,
        'average_age': average_age['avg_age'],
        'book_value': book_value['total_value'],
        'yearly_fuel_costs': yearly_fuel_costs['total_fuel_cost'],
        'yearly_service_costs': yearly_service_costs['total_service_cost'],
        'red_zone_vehicles': red_zone_vehicles,
        'centers': center_data
    }

    return render(request, 'fleet/dashboard.html', context)

def center_statistics(request, center_code):
    # Check if the user has access to this center
    if not request.user.allowed_centers.filter(center=center_code).exists():
        return HttpResponseForbidden("Nemate pristup ovim podacima.")
    
    # Podupit za poslednju OU (tj. centar) za vozilo
    latest_center_code = JobCode.objects.filter(
        vehicle=OuterRef('vehicle')
    ).order_by('-assigned_date').values('organizational_unit__center')[:1]

    # Filtriranje FuelConsumption po centru
    fuel_data = FuelConsumption.objects.annotate(
        vehicle_center_code=Subquery(latest_center_code)
    ).filter(
        vehicle_center_code=center_code
    ).annotate(
        year=TruncYear('date'),
        month=TruncMonth('date')
    ).values('year', 'month').annotate(
        total_fuel_quantity=Sum('amount'),
        total_fuel_cost=Sum('cost_bruto')
    ).order_by('year', 'month')

    # Service costs statistics (grouped by service type, month, and year, filtered by center)
    service_data = ServiceTransaction.objects.annotate(
        vehicle_center_code=Subquery(latest_center_code)
    ).filter(
        vehicle_center_code=center_code
    ).annotate(
        year=TruncYear('datum'),
        month=TruncMonth('datum')
    ).values('year', 'month').annotate(
    total_cost_gume=Sum('potrazuje', filter=Q(popravka_kategorija__name__icontains='gume')),
    total_cost_redovan_servis=Sum('potrazuje', filter=Q(popravka_kategorija__name__icontains='redovan servis')),
    total_cost_tehnicki_pregled=Sum('potrazuje', filter=Q(popravka_kategorija__name__icontains='tehnicki pregled')),
    total_cost_registracija=Sum('potrazuje', filter=Q(popravka_kategorija__name__icontains='registracija'))
    ).order_by('year', 'month')

    # Ispravljen upit za registracije
    insurance_data = Policy.objects.annotate(
        vehicle_center_code=Subquery(latest_center_code)
    ).filter(
        vehicle_center_code=center_code
    ).annotate(
        year=TruncYear('issue_date'),
        month=TruncMonth('issue_date')
    ).values('year', 'month').annotate(
        total_registration_cost=Sum('premium_amount')
    ).order_by('year', 'month')

    # Combine all data based on year and month
    consolidated_data = {}
    
    # Consolidating fuel data
    for fuel in fuel_data:
        year = fuel['year'].year
        month = fuel['month'].month
        consolidated_data[(year, month)] = {
            'total_fuel_quantity': fuel['total_fuel_quantity'],
            'total_fuel_cost': fuel['total_fuel_cost'],
            'total_cost_gume': 0,
            'total_cost_redovan_servis': 0,
            'total_cost_tehnicki_pregled': 0,
            'total_cost_registracija': 0,
            'total_registration_cost': 0
        }

    # Consolidating service data
    for service in service_data:
        year = service['year'].year
        month = service['month'].month
        if (year, month) not in consolidated_data:
            consolidated_data[(year, month)] = {
                'total_fuel_quantity': 0,
                'total_fuel_cost': 0,
                'total_cost_gume': service['total_cost_gume'],
                'total_cost_redovan_servis': service['total_cost_redovan_servis'],
                'total_cost_tehnicki_pregled': service['total_cost_tehnicki_pregled'],
                'total_cost_registracija': service['total_cost_registracija'],
                'total_registration_cost': 0
            }
        else:
            consolidated_data[(year, month)].update({
                'total_cost_gume': service['total_cost_gume'],
                'total_cost_redovan_servis': service['total_cost_redovan_servis'],
                'total_cost_tehnicki_pregled': service['total_cost_tehnicki_pregled'],
                'total_cost_registracija': service['total_cost_registracija']
            })

    # Consolidating insurance data
    for insurance in insurance_data:
        year = insurance['year'].year
        month = insurance['month'].month
        if (year, month) not in consolidated_data:
            consolidated_data[(year, month)] = {
                'total_fuel_quantity': 0,
                'total_fuel_cost': 0,
                'total_cost_gume': 0,
                'total_cost_redovan_servis': 0,
                'total_cost_tehnicki_pregled': 0,
                'total_cost_registracija': 0,
                'total_registration_cost': insurance['total_registration_cost']
            }
        else:
            consolidated_data[(year, month)].update({
                'total_registration_cost': insurance['total_registration_cost']
            })

    context = {
        'consolidated_data': consolidated_data,
        'center_code': center_code,
    }

    return render(request, 'fleet/dashboard_center.html', context)

# <!-- ======================================================================= -->
#                           <!-- ORGANIZATIONAL UNITS -->
# <!-- ======================================================================= -->

class OrganizationalUnitListView(ListView):
    model = OrganizationalUnit
    template_name = 'fleet/organizational_units_list.html'
    context_object_name = 'units'


class OrganizationalUnitCreateView(CreateView):
    model = OrganizationalUnit
    form_class = OrganizationalUnitForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('organizational_unit_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Kreiraj novu organizacionu jedinicu'
        context['submit_button_label'] = 'Sačuvaj'
        return context


class OrganizationalUnitUpdateView(UpdateView):
    model = OrganizationalUnit
    form_class = OrganizationalUnitForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('organizational_unit_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni organizacionu jedinicu'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context
    
# class OrganizationalUnitDeleteView(DeleteView):
#     model = OrganizationalUnit
#     success_url = reverse_lazy('organizational_unit_list')

# <!-- ======================================================================= -->
#                           <!-- VEHICLE -->
# <!-- ======================================================================= -->
class VehicleListView(LoginRequiredMixin, ListView):
    model = Vehicle
    template_name = 'fleet/vehicle_list.html'
    context_object_name = 'vehicles'
    form_class = VehicleFilterForm  # Dodajemo formu za filtriranje 

    def get_queryset(self):
        queryset = super().get_queryset()

        # ✔️ Ako korisnik NIJE tražio otpisana, filtriraj ih
        show_archived = self.request.GET.get('show_archived')
        if show_archived != 'yes':
            queryset = queryset.filter(otpis=False)
        
        # Dohvati vrednosti iz GET parametara
        org_unit = self.request.GET.get('org_unit')
        fuel_in_last_6_months = self.request.GET.get('fuel_in_last_6_months')
        center_code = self.request.GET.get('center_code')

        # Subquery to get the latest org_unit for each Vehicle
        latest_org_unit_subquery = JobCode.objects.filter(
            vehicle_id=OuterRef('pk')
        ).order_by('-assigned_date').values('organizational_unit__code')[:1]

        # Subquery to get the latest TrafficCard for each Vehicle
        latest_traffic_card_subquery = TrafficCard.objects.filter(
            vehicle_id=OuterRef('pk')
        ).order_by('-issue_date').values('registration_number')[:1]
        
        last_mileage_subquery = FuelConsumption.objects.filter(
            vehicle=OuterRef('pk')
        ).order_by('-mileage').values('mileage')[:1]

        latest_job_code_id_subquery = JobCode.objects.filter(
            vehicle=OuterRef('pk')
        ).order_by('-assigned_date').values('id')[:1]

        queryset = queryset.annotate(
            latest_job_code_id=Subquery(latest_job_code_id_subquery),
            latest_org_unit=Subquery(latest_org_unit_subquery),
            registration_number=Subquery(latest_traffic_card_subquery),
            total_repairs=Sum('service_transactions__potrazuje'),
            mileage=Subquery(last_mileage_subquery),
        )

        if org_unit:
            # Create a subquery that identifies the IDs of JobCode objects
            # which are the LATEST for their vehicle AND match the org_unit_id
            matching_latest_job_codes = JobCode.objects.filter(
                id=OuterRef('latest_job_code_id'), # Match the ID of the latest JobCode from the outer Vehicle query
                organizational_unit_id=org_unit # Filter by the selected org unit ID
            ).values('vehicle_id') # Get the vehicle ID for these matching JobCodes

            # Filter the main queryset to include only vehicles whose PK is in the list
            # of vehicle_ids derived from the matching latest job codes.
            queryset = queryset.filter(pk__in=Subquery(matching_latest_job_codes))


        # Filter for Center Code (center_code)
        if center_code:
            # Create a subquery that identifies the IDs of JobCode objects
            # which are the LATEST for their vehicle AND match the center_code
            matching_latest_job_codes_by_center = JobCode.objects.filter(
                id=OuterRef('latest_job_code_id'), # Match the ID of the latest JobCode from the outer Vehicle query
                organizational_unit__center=center_code # Filter by the selected center code string
            ).values('vehicle_id')

            # Filter the main queryset to include only vehicles whose PK is in the list
            # of vehicle_ids derived from the matching latest job codes by center.
            queryset = queryset.filter(pk__in=Subquery(matching_latest_job_codes_by_center))

        # Filter za gorivo u poslednjih 6 meseci
        if fuel_in_last_6_months == 'yes':
            six_months_ago = timezone.now() - timedelta(days=180)
            queryset = queryset.filter(
                fuel_consumptions__date__gte=six_months_ago
            ).distinct()  # Da ne vraća duplikate ako postoji više sipanja
        elif fuel_in_last_6_months == 'no':
            six_months_ago = timezone.now() - timedelta(days=180)
            queryset = queryset.exclude(
                fuel_consumptions__date__gte=six_months_ago
            ).distinct()  # Da ne vraća duplikate ako postoji više sipanja
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vehicles = context['vehicles']
        vehicle_consumption_data = {}

        for vehicle in vehicles:
            vehicle_consumption_data[vehicle.id] = calculate_average_fuel_consumption(vehicle)

        form = VehicleFilterForm(self.request.GET or None)
        context['vehicle_consumption_data'] = vehicle_consumption_data
        context['form'] = form
        context['title'] = 'Lista vozila'

        return context
    
# DETALJI VOZILA
class VehicleDetailView(LoginRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'fleet/vehicle_detail.html'
    context_object_name = 'vehicle'

    def get(self, request, *args, **kwargs):
        print("VehicleDetailView get() method called")
        vehicle = self.get_object()
        print(vehicle)

        # Subquery to get the latest org_unit for each Vehicle
        # Ovo vam vraća kod centra iz JobCode-a, bazirano na 'organizational_unit__center'
        latest_org_unit_subquery = JobCode.objects.filter(
            vehicle_id=OuterRef('pk')
        ).order_by('-assigned_date').values('organizational_unit__center')[:1]

        # Annotate the vehicle queryset with the latest org_unit code
        vehicle_with_latest_org_unit = Vehicle.objects.annotate(
            latest_org_unit=Subquery(latest_org_unit_subquery)
        ).get(pk=vehicle.pk)

        # --- Modifikovana logika za proveru dozvola ---
        user_allowed_centers_manager = request.user.allowed_centers # Dobijamo ManyRelatedManager

        # Proveravamo da li korisnik uopšte ima definisane dozvoljene centre
        # Prazan manager (kada korisnik nema dodeljene centre) se u if uslovu ponaša kao False
        if user_allowed_centers_manager.exists(): # Proverava da li manager sadrži ijedan srodni objekat
             # Iz managers dobijamo listu kodova dozvoljenih centara.
             # *** VAŽNO: Zamenite 'center' u values_list('center', flat=True)
             # *** sa STVARNIM IMENOM POLJA na modelu na koji ukazuje
             # *** request.user.allowed_centers, a koje sadrži kod centra.
             # *** Na osnovu vašeg subquery-a 'organizational_unit__center',
             # *** verovatno se to polje zove 'center' ili 'code'.
             allowed_centers_codes = user_allowed_centers_manager.values_list('center', flat=True) # Vraća QuerySet sa listom vrednosti, flat=True daje Python listu


             # Sada proveravamo da li kod poslednjeg organizacione jedinice vozila
             # NIJE u listi dozvoljenih kodova za korisnika.
             # Dodata provera da li latest_org_unit nije None (ako vozilo nema JobCode)
             if vehicle_with_latest_org_unit.latest_org_unit is not None and \
                vehicle_with_latest_org_unit.latest_org_unit not in allowed_centers_codes:
                 return HttpResponseForbidden("Nemate dozvolu za pristup ovom vozilu.")
        # else: Ako user_allowed_centers_manager.exists() vrati False, to znači da korisnik
        # nema eksplicitno definisane centre kojima može da pristupi. Ako je željeno ponašanje
        # da takav korisnik ima pristup svim vozilima, onda ova if struktura to omogućava
        # jer se provera dozvole preskače.
        # --- Kraj modifikovane logike ---


        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        vehicle = self.get_object()

        # 1. Aktivne polise osiguranja
        active_policies = Policy.objects.filter(vehicle=vehicle, end_date__gte=datetime.date.today())

        # 2. Job Code
        current_job_code = JobCode.objects.filter(vehicle=vehicle).order_by('-assigned_date').first()
        job_codes = JobCode.objects.filter(vehicle=vehicle).order_by('-assigned_date')

        
        # 3. Gorivo, kilometraza i kartice
        nis_card = vehicle.nis_transactions.filter().first()
        omv_card = vehicle.omv_transactions.filter().first()

        mileage = vehicle.fuel_consumptions.order_by('-mileage').values_list('mileage', flat=True).first()

        consumptions = vehicle.fuel_consumptions.all() # Za tabelu potrosnje
        average_consumption = calculate_average_fuel_consumption(vehicle) # poslednjih 10
        average_consumption_ever = calculate_average_fuel_consumption_ever(vehicle) # Sva tocenja
        
        # Grupisanje po mesecima i godinama uz dobavljača, agregiranje količine i bruto cene
        fuel_data = FuelConsumption.objects.filter(vehicle=vehicle).annotate(
            month=TruncMonth('date'),
            year=TruncYear('date')
        ).values('month', 'year', 'supplier').annotate(
            total_liters=Sum('amount'),
            total_cost_bruto=Sum('cost_bruto')
        ).order_by('year', 'month', 'supplier')

        # Razdvajanje podataka za OMV i NIS
        omv_data = fuel_data.filter(supplier='OMV')
        nis_data = fuel_data.filter(supplier='NIS')

        # 4. Leasing data
        lease_info = Lease.objects.filter(vehicle=vehicle).order_by('-start_date').first()
        lease_intrests = LeaseInterest.objects.filter(lease=lease_info).order_by('-year')

        # 5. General status (red/green light based on repair costs)
        repair_costs = vehicle.service_transactions.aggregate(total_repairs=Sum('potrazuje'))['total_repairs'] or 0
        requisition_costs = vehicle.requisitions.aggregate(total_requisitions=Sum('vrednost_nab'))['total_requisitions'] or 0

        service_list = vehicle.service_transactions.order_by('-datum')
        requisition_list = vehicle.requisitions.order_by('-datum_trebovanja')
        
        # 6. Saobracajna dozvol i istorija
        trafic_cards = TrafficCard.objects.filter(vehicle=vehicle).order_by('-issue_date')
        trafic_card = trafic_cards.first()
        status_light = 'green' if repair_costs < vehicle.purchase_value else 'red'

        
        context.update({
            'lease_info': lease_info,
            'lease_intrests':lease_intrests,
            'nis_card': nis_card,
            'omv_card': omv_card,
            'mileage': mileage,
            'active_policies': active_policies,
            'average_consumption': average_consumption,
            'average_consumption_ever': average_consumption_ever,
            'omv_data':omv_data,
            'nis_data':nis_data,
            'current_job_code': current_job_code,
            'job_codes': job_codes,
            'status_light': status_light,
            'repair_costs': repair_costs,
            'requisition_costs':requisition_costs,
            'service_list':service_list,
            'requisition_list':requisition_list,
            'consumptions': consumptions,
            'trafic_cards':trafic_cards,
            'trafic_card':trafic_card,
            'title':f"Detalji vozila {self.object.brand} {self.object.model}"
        })
        return context



class VehicleCreateView(LoginRequiredMixin, CreateView):
    model = Vehicle
    form_class = VehicleForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('vehicle_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        return redirect('trafficcard_create', vehicle_id=self.object.id)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Kreiraj novo vozilo'
        context['submit_button_label'] = 'Dodaj vozilo'
        return context

class VehicleUpdateView(LoginRequiredMixin, UpdateView):
    model = Vehicle
    form_class = VehicleForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('vehicle_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni podatke vozila'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class VehicleTogleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        vehicle = get_object_or_404(Vehicle, pk=pk)
        vehicle.otpis = not vehicle.otpis
        vehicle.save()
        status = "aktivano" if vehicle.otpis else "otpisano"
        messages.success(request, f"Vozilo je uspešno {status}.")
        return redirect('vehicle_detail', pk=pk)


class VehicleDeleteView(LoginRequiredMixin, DeleteView):
    model = Vehicle
    success_url = reverse_lazy('vehicle_list')
    template_name = 'fleet/vehicle_confirm_delete.html'
    context_object_name = 'vehicle'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši vozilo'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)



# <!-- ======================================================================= -->
#                           <!-- TRAFIC CARD -->
# <!-- ======================================================================= -->
class TrafficCardListView(LoginRequiredMixin, ListView):
    model = TrafficCard
    template_name = 'fleet/trafficcard_list.html'
    context_object_name = 'traffic_cards'
    form_class = TrafficCardFilterForm

    def get_queryset(self):
        queryset = super().get_queryset().select_related('vehicle')
        self.filter_form = self.form_class(self.request.GET or None)

        # Subquery: poslednji JobCode po datumu
        latest_org_unit_subquery = JobCode.objects.filter(
            vehicle_id=OuterRef('vehicle_id')
        ).order_by('-assigned_date').values('organizational_unit__code')[:1]

        latest_center_subquery = JobCode.objects.filter(
            vehicle_id=OuterRef('vehicle_id')
        ).order_by('-assigned_date').values('organizational_unit__center')[:1]

        queryset = queryset.annotate(
            latest_org_unit=Subquery(latest_org_unit_subquery),
            latest_center=Subquery(latest_center_subquery),
        )

        if self.filter_form.is_valid():
            org_unit = self.filter_form.cleaned_data.get('organizational_unit')
            center = self.filter_form.cleaned_data.get('center')

            if org_unit:
                queryset = queryset.filter(latest_org_unit=org_unit.code)

            if center:
                queryset = queryset.filter(latest_center=center)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = self.filter_form
        return context
class TrafficCardCreateView(LoginRequiredMixin, CreateView):
    model = TrafficCard
    form_class = TrafficCardForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('trafficcard_list')
   
    def get_initial(self):
        vehicle_id = self.kwargs.get('vehicle_id')
        return {'vehicle': vehicle_id}

    def form_valid(self, form):
        response = super().form_valid(form)
        return redirect('jobcode_create', vehicle_id = self.object.vehicle.id)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Kreiraj novu saobraćajnu dozvolu'
        context['submit_button_label'] = 'Dodaj saobraćajnu dozvolu'
        return context

class TrafficCardUpdateView(LoginRequiredMixin, UpdateView):
    model = TrafficCard
    form_class = TrafficCardForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('trafficcard_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni podatke saobraćajne dozvole'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class TrafficCardDetailView(LoginRequiredMixin, DetailView):
    model = TrafficCard
    template_name = 'fleet/trafficcard_detail.html'
    context_object_name = 'traffic_card'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji saobraćajne dozvole {self.object.registration_number}"
        return context

class TrafficCardDeleteView(LoginRequiredMixin, DeleteView):
    model = TrafficCard
    success_url = reverse_lazy('trafficcard_list')
    template_name = 'fleet/trafficcard_confirm_delete.html'
    context_object_name = 'traffic_card'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši saobraćajnu dozvolu'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)



# <!-- ======================================================================= -->
#                           <!-- JOB CODE -->
# <!-- ======================================================================= -->
class JobCodeListView(LoginRequiredMixin, ListView):
    model = JobCode
    template_name = 'fleet/jobcode_list.html'
    context_object_name = 'job_codes'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista šifara poslova'
        return context

class JobCodeCreateView(LoginRequiredMixin, CreateView):
    model = JobCode
    form_class = JobCodeForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('jobcode_list')

    def get_initial(self):
        return {
            'vehicle': self.kwargs.get('vehicle_id')
        }
    
    def form_valid(self, form):
        response = super().form_valid(form)
        return redirect('vehicle_detail', pk = self.object.vehicle.id)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Kreiraj novu šifru posla'
        context['submit_button_label'] = 'Dodaj šifru posla'
        return context
    
class JobCodeUpdateView(LoginRequiredMixin, UpdateView):
    model = JobCode
    form_class = JobCodeForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('jobcode_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni šifru posla'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class JobCodeDetailView(LoginRequiredMixin, DetailView):
    model = JobCode
    template_name = 'fleet/jobcode_detail.html'
    context_object_name = 'job_code'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji šifre posla {self.object.job_code}"
        return context

class JobCodeDeleteView(LoginRequiredMixin, DeleteView):
    model = JobCode
    success_url = reverse_lazy('jobcode_list')
    template_name = 'fleet/jobcode_confirm_delete.html'
    context_object_name = 'job_code'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši šifru posla'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)



# <!-- ======================================================================= -->
#                           <!-- LEASE -->
# <!-- ======================================================================= -->
class LeaseListView(LoginRequiredMixin, ListView):
    model = Lease
    template_name = 'fleet/lease_list.html'
    context_object_name = 'leases'
    # Dodajte permission_required atribut
    permission_required = 'fleet.view_lease'
    raise_exception = True
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista zakupa vozila'
        return context

class LeaseCreateView(LoginRequiredMixin, CreateView):
    model = Lease
    form_class = LeaseForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('lease_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Kreiraj novi zakup'
        context['submit_button_label'] = 'Dodaj zakup'
        return context
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Lease

def export_leases_to_excel(request):
    # Kreiranje novog Excel fajla
    wb = Workbook()
    ws = wb.active
    ws.title = "Lizing ugovori"

    # Naslovi kolona
    headers = [
        "Vozilo (šasija)",
        "Šifra partnera",
        "Naziv partnera",
        "Šifra posla",
        "Broj ugovora",
        "Trenutna vrednost otplate",
        "Vrsta lizinga",
        "Datum početka",
        "Datum završetka",
        "Napomena"
    ]
    ws.append(headers)

    # Podaci iz baze
    leases = Lease.objects.select_related('vehicle').all()
    for lease in leases:
        ws.append([
            lease.vehicle.chassis_number if lease.vehicle else '',
            lease.partner_code,
            lease.partner_name,
            lease.job_code,
            lease.contract_number,
            float(lease.current_payment_amount),
            lease.get_lease_type_display(),
            lease.start_date.strftime('%d.%m.%Y'),
            lease.end_date.strftime('%d.%m.%Y'),
            lease.note or ''
        ])

    # Odgovor kao Excel fajl
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="lizing_ugovori.xlsx"'
    wb.save(response)
    return response

class LeaseUpdateView(LoginRequiredMixin, UpdateView):
    model = Lease
    form_class = LeaseForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('lease_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni zakup'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class LeaseDetailView(LoginRequiredMixin, DetailView):
    model = Lease
    template_name = 'fleet/lease_detail.html'
    context_object_name = 'lease'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji zakupa {self.object.partner_name}"
        return context

class LeaseDeleteView(LoginRequiredMixin, DeleteView):
    model = Lease
    success_url = reverse_lazy('lease_list')
    template_name = 'fleet/lease_confirm_delete.html'
    context_object_name = 'lease'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši zakup'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)
    


# <!-- ======================================================================= -->
#                           <!-- POLICY -->
# <!-- ======================================================================= -->
class PolicyListView(LoginRequiredMixin, ListView):
    model = Policy
    template_name = 'fleet/policy_list.html'
    context_object_name = 'policies'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista polisa osiguranja'
        return context

class PolicyFixingListView(LoginRequiredMixin, ListView):
    model = Policy
    template_name = 'fleet/draft_policy_list.html'
    context_object_name = 'policies'

    def get_queryset(self):
        # Filter policies where the vehicle is None
        return DraftPolicy.objects.all

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista polisa osiguranja koje morate dopuniti'
        return context

class ExpiringAndNotRenewedPolicyView(LoginRequiredMixin, ListView):
    template_name = 'fleet/policy_expiring.html'
    model = Policy

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        thirty_days_from_now = today + timedelta(days=30)
        
        # Pronađi najnoviju polisu po vozilu i tipu osiguranja
        newest_policy = Policy.objects.filter(
            vehicle=OuterRef('vehicle'),
            insurance_type=OuterRef('insurance_type')
        ).order_by('-end_date').values('end_date', 'is_renewable')[:1]

        # Dodaj anotaciju sa datumom i informacijom da li je polisa obnovljiva
        expiring_policies = Policy.objects.annotate(
            latest_end_date=Subquery(newest_policy.values('end_date')[:1]),
            latest_is_renewable=Subquery(newest_policy.values('is_renewable')[:1])
        ).filter(
            end_date__gte=today,
            end_date__lte=thirty_days_from_now,
            end_date=F('latest_end_date'),
            latest_is_renewable=True  # Prikazuj samo ako je obnovljiva
        )

        # Proveri da li postoji novija polisa
        newer_policy_exists = Policy.objects.filter(
            vehicle=OuterRef('vehicle'),
            insurance_type=OuterRef('insurance_type'),
            start_date__gt=OuterRef('start_date')
        )

        # Pronađi istekle polise koje nisu obnovljene
        expired_unrenewed_policies = Policy.objects.annotate(
            has_newer_policy=Subquery(newer_policy_exists.values('id')[:1]),
            latest_is_renewable=Subquery(newest_policy.values('is_renewable')[:1])
        ).filter(
            end_date__lt=today,
            has_newer_policy__isnull=True,   # Nema novije polise
            latest_is_renewable=True         # Prikazuj samo ako je obnovljiva
        )


        context['expiring_policies'] = expiring_policies
        context['expired_unrenewed_policies'] = expired_unrenewed_policies
        context['title'] = 'Liste polisa koje ističu i koje su istekle i nisu obnovljene'
        return context

class PolicyCreateView(LoginRequiredMixin, CreateView):
    model = Policy
    form_class = PolicyForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('policy_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Kreiraj novu polisu osiguranja'
        context['submit_button_label'] = 'Dodaj polisu'
        return context

class PolicyUpdateView(LoginRequiredMixin, UpdateView):
    model = Policy
    form_class = PolicyForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('policy_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni polisu osiguranja'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context
    
    def form_valid(self, form):
        # Prvo sačuvaj izmene
        response = super().form_valid(form)
        next_url = self.request.GET.get('next')

        # Zatim preusmeri korisnika nazad ako postoji 'next' parametar
        next_url = self.request.GET.get('next')
        if next_url:
            return HttpResponseRedirect(next_url)
        
        return response


class PolicyDetailView(LoginRequiredMixin, DetailView):
    model = Policy
    template_name = 'fleet/policy_detail.html'
    context_object_name = 'policy'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji polise {self.object.policy_number}"
        return context

class PolicyDeleteView(LoginRequiredMixin, DeleteView):
    model = Policy
    success_url = reverse_lazy('policy_list')
    template_name = 'fleet/policy_confirm_delete.html'
    context_object_name = 'policy'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši polisu osiguranja'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)

class DraftPolicyUpdateView(UpdateView):
    model = DraftPolicy
    form_class = PolicyForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('policy_list')  # Preusmeravanje nakon uspešne izmene

    def form_valid(self, form):
        # Sačuvaj izmene u draft tabeli
        draft = form.save(commit=False)
        print("Izmene sačuvane u draft tabeli.")

        # Provera da li su svi potrebni podaci sada prisutni osim opcionalnih polja
        required_fields = [
            'partner_pib',
            'partner_name',
            'invoice_id',
            'invoice_number',
            'issue_date',
            'insurance_type',
            'policy_number',
            'premium_amount',
            'start_date',
            'end_date',
            'first_installment_amount',
            'other_installments_amount',
            'number_of_installments'
        ]
        is_complete = all(
            getattr(draft, field) is not None and getattr(draft, field) != ''
            for field in required_fields
        )

        if is_complete:
            # Ako su podaci kompletni, prebacujemo ih u glavni model
            policy = Policy(
                vehicle=draft.vehicle,
                partner_pib=draft.partner_pib,
                partner_name=draft.partner_name,
                invoice_id=draft.invoice_id,
                invoice_number=draft.invoice_number,
                issue_date=draft.issue_date,
                insurance_type=draft.insurance_type,
                policy_number=draft.policy_number,
                premium_amount=draft.premium_amount,
                start_date=draft.start_date,
                end_date=draft.end_date,
                first_installment_amount=draft.first_installment_amount,
                other_installments_amount=draft.other_installments_amount,
                number_of_installments=draft.number_of_installments
            )
            policy.save()
            print("Podaci migrirani u glavnu tabelu Policy.")
            draft.delete()  # Obrisan unos iz draft tabele
            return redirect(self.success_url)

        # Ako podaci nisu kompletni, sačuvaj ih samo u draft tabeli
        else:
            draft.save()  # Sačuvaj izmene u draft tabeli
            print("Podaci nisu kompletni, ostaju u draft tabeli.")
            return super().form_valid(form)


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Dopuni polisu'
        context['submit_button_label'] = 'Sačuvaj'
        return context

# <!-- ======================================================================================== -->
#                           <!-- FUEL CONSUMPTION -->
# <!-- ======================================================================================== -->
from django_filters.views import FilterView

class FuelConsumptionListView(LoginRequiredMixin, FilterView):
    model = FuelConsumption
    filterset_class = FuelFilterForm
    template_name = 'fleet/fuelconsumption_list.html'
    context_object_name = 'fuel_consumptions'

    def get_queryset(self):
        # Subquery to get the latest TrafficCard for each Vehicle
        latest_traffic_card_subquery = TrafficCard.objects.filter(
            vehicle=OuterRef('vehicle')
        ).order_by('-issue_date').values('registration_number')[:1]

        # Base queryset with annotation
        queryset = super().get_queryset().annotate(
            registration_number=Subquery(latest_traffic_card_subquery)
        )

        # Default filtering logic
        if not self.request.GET:  # If there are no GET parameters
            today = timezone.now().date()
            forty_days_ago = today - timedelta(days=40)
            return queryset.filter(date__gte=forty_days_ago, date__lte=today)

        # Apply filter if GET parameters are present
        form = self.filterset_class(self.request.GET, queryset=queryset)
        if form.is_valid():
            return form.qs
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = self.filterset_class(self.request.GET, queryset=self.get_queryset())
        context.update({
            'filter': form,
            'title': 'Lista potrošnje goriva',
        })
        return context

class FuelTransactionsListView(LoginRequiredMixin, ListView):
    template_name = 'fleet/fuel_transactions_list.html'
    context_object_name = 'fuel_transactions'

    def get_queryset(self):
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        # Postavi podrazumevane vrednosti ako GET parametri nisu prisutni
        if not start_date:
            start_date = date.today() - timedelta(days=40)
        if not end_date:
            end_date = date.today()

        return get_fuel_consumption_queryset(start_date=start_date, end_date=end_date)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Inicijalizacija forme sa trenutnim GET vrednostima ili podrazumevanim datumima
        context['filter_form'] = FuelTransactionFilterForm(self.request.GET or {
            'start_date': (date.today() - timedelta(days=40)).strftime('%Y-%m-%d'),
            'end_date': date.today().strftime('%Y-%m-%d')
        })

        return context



class FuelConsumptionCreateView(LoginRequiredMixin, CreateView):
    model = FuelConsumption
    form_class = FuelConsumptionForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('fuelconsumption_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Dodaj potrošnju goriva'
        context['submit_button_label'] = 'Dodaj'
        return context

class FuelConsumptionUpdateView(LoginRequiredMixin, UpdateView):
    model = FuelConsumption
    form_class = FuelConsumptionForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('fuelconsumption_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni potrošnju goriva'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class FuelConsumptionDetailView(LoginRequiredMixin, DetailView):
    model = FuelConsumption
    template_name = 'fleet/fuelconsumption_detail.html'
    context_object_name = 'fuel_consumption'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji potrošnje goriva {self.object.date}"
        return context

class FuelConsumptionDeleteView(LoginRequiredMixin, DeleteView):
    model = FuelConsumption
    success_url = reverse_lazy('fuelconsumption_list')
    template_name = 'fleet/fuelconsumption_confirm_delete.html'
    context_object_name = 'fuel_consumption'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši potrošnju goriva'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)





# <!-- ======================================================================================== -->
#                           <!-- EMPLOYEES -->
# <!-- ======================================================================================== -->
class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'fleet/employee_list.html'
    context_object_name = 'employees'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista zaposlenih'
        return context

class EmployeeCreateView(LoginRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('employee_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Kreiraj novog zaposlenog'
        context['submit_button_label'] = 'Dodaj zaposlenog'
        return context

class EmployeeUpdateView(LoginRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('employee_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni podatke zaposlenog'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class EmployeeDetailView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = 'fleet/employee_detail.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji zaposlenog {self.object.name}"
        return context

class EmployeeDeleteView(LoginRequiredMixin, DeleteView):
    model = Employee
    success_url = reverse_lazy('employee_list')
    template_name = 'fleet/employee_confirm_delete.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši zaposlenog'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)




# <!-- ======================================================================================== -->
#                           <!-- FUEL CONSUMPTION -->
# <!-- ======================================================================================== -->
class IncidentListView(LoginRequiredMixin, ListView):
    model = Incident
    template_name = 'fleet/incident_list.html'
    context_object_name = 'incidents'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista incidenata'
        return context

class IncidentCreateView(LoginRequiredMixin, CreateView):
    model = Incident
    form_class = IncidentForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('incident_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Dodaj incident'
        context['submit_button_label'] = 'Dodaj'
        return context

class IncidentUpdateView(LoginRequiredMixin, UpdateView):
    model = Incident
    form_class = IncidentForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('incident_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni incident'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class IncidentDetailView(LoginRequiredMixin, DetailView):
    model = Incident
    template_name = 'fleet/incident_detail.html'
    context_object_name = 'incident'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji incidenta {self.object.date}"
        return context

class IncidentDeleteView(LoginRequiredMixin, DeleteView):
    model = Incident
    success_url = reverse_lazy('incident_list')
    template_name = 'fleet/incident_confirm_delete.html'
    context_object_name = 'incident'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši incident'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)




# <!-- ======================================================================================== -->
#                           <!-- PUTNI NALOG -->
# <!-- ======================================================================================== -->
class PutniNalogListView(LoginRequiredMixin, ListView):
    model = PutniNalog
    template_name = 'fleet/putninalog_list.html'
    context_object_name = 'putni_nalozi'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista putnih naloga'
        return context

from django.http import FileResponse, HttpResponse
import os
from django.conf import settings

def download_travel_order_excel(request, pk):
    """
    Preuzmi generisani Excel fajl za dati PutniNalog.
    """
    try:
        # Dobavi putni nalog
        putni_nalog = PutniNalog.objects.get(pk=pk)

        # Generiši tačno ime fajla kao u `populate_putni_nalog_template()`
        sanitized_order_number = sanitize_filename(putni_nalog.order_number)
        file_name = f"PutniNalog_{sanitized_order_number}.xlsx"
        file_path = os.path.join(settings.MEDIA_ROOT, "travel_orders", file_name)

        # ✅ Proveri da li fajl postoji pre preuzimanja
        if not os.path.exists(file_path):
            return HttpResponse(f"Greška: Fajl nije pronađen {file_path}", status=404)

        # ✅ Vraćanje fajla kao odgovor
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)

    except PutniNalog.DoesNotExist:
        return HttpResponse("Greška: Putni nalog ne postoji.", status=404)
    except Exception as e:
        return HttpResponse(f"Greška: {str(e)}", status=500)


class PutniNalogCreateView(LoginRequiredMixin, CreateView):
    model = PutniNalog
    form_class = PutniNalogForm
    template_name = 'fleet/putni_nalog_form.html'
    success_url = reverse_lazy('putninalog_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Dodaj putni nalog'
        context['submit_button_label'] = 'Dodaj'
        return context
    
    def form_valid(self, form):
        # Sačuvaj objekat
        self.object = form.save()

        # Generiši Excel fajl
        file_path = populate_putni_nalog_template(self.object)

        # Vrati URL za preuzimanje i preusmeravanje
        return JsonResponse({
            'file_url': reverse('download_travel_order_excel', args=[self.object.id]),
            'redirect_url': reverse('putninalog_list')
        })


class PutniNalogUpdateView(LoginRequiredMixin, UpdateView):
    model = PutniNalog
    form_class = PutniNalogForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('putninalog_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni putni nalog'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context
    
    def form_valid(self, form):
        # Save the object
        self.object = form.save()

        # Generate the Excel file
        file_path = populate_putni_nalog_template(self.object)

        # Serve the file as a response
        response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=f"PutniNalog_{self.object.id}.xlsx")
        return response
class PutniNalogDetailView(LoginRequiredMixin, DetailView):
    model = PutniNalog
    template_name = 'fleet/putninalog_detail.html'
    context_object_name = 'putni_nalog'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji putnog naloga {self.object.travel_date}"
        return context

class PutniNalogDeleteView(LoginRequiredMixin, DeleteView):
    model = PutniNalog
    success_url = reverse_lazy('putninalog_list')
    template_name = 'fleet/putninalog_confirm_delete.html'
    context_object_name = 'putni_nalog'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši putni nalog'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)




# <!-- ======================================================================================== -->
#                           <!-- SERVICE TYPES -->
# <!-- ======================================================================================== -->
class ServiceTypeListView(LoginRequiredMixin, ListView):
    model = ServiceType
    template_name = 'fleet/servicetype_list.html'
    context_object_name = 'service_types'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista tipova servisa'
        return context

class ServiceTypeCreateView(LoginRequiredMixin, CreateView):
    model = ServiceType
    form_class = ServiceTypeForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('servicetype_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Dodaj tip servisa'
        context['submit_button_label'] = 'Dodaj'
        return context

class ServiceTypeUpdateView(LoginRequiredMixin, UpdateView):
    model = ServiceType
    form_class = ServiceTypeForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('servicetype_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni tip servisa'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class ServiceTypeDetailView(LoginRequiredMixin, DetailView):
    model = ServiceType
    template_name = 'fleet/servicetype_detail.html'
    context_object_name = 'service_type'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji tipa servisa {self.object.name}"
        return context

class ServiceTypeDeleteView(LoginRequiredMixin, DeleteView):
    model = ServiceType
    success_url = reverse_lazy('servicetype_list')
    template_name = 'fleet/servicetype_confirm_delete.html'
    context_object_name = 'service_type'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši tip servisa'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)



# <!-- ======================================================================================== -->
#                           <!-- SERVICES -->
# <!-- ======================================================================================== -->



class ServiceListView(LoginRequiredMixin, ListView):
    model = Service
    template_name = 'fleet/service_list.html'
    context_object_name = 'services'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista servisa'
        return context

class ServiceFixingListView(LoginRequiredMixin, ListView):
    model = DraftServiceTransaction
    template_name = 'fleet/draft_service_transactions_list.html'
    context_object_name = 'service_transactions'

    def get_queryset(self):
        queryset = DraftServiceTransaction.objects.select_related('vehicle').all()
        self.form = ServiceFixingFilterForm(self.request.GET)

        if self.form.is_valid():
            datum_od = self.form.cleaned_data.get('datum_od')
            datum_do = self.form.cleaned_data.get('datum_do')
            vozilo = self.form.cleaned_data.get('vozilo')
            partner = self.form.cleaned_data.get('partner')
            nije_garaza = self.form.cleaned_data.get('nije_garaza')

            if datum_od:
                queryset = queryset.filter(datum__gte=datum_od)
            if datum_do:
                queryset = queryset.filter(datum__lte=datum_do)
            if partner:
                queryset = queryset.filter(naz_par_pl__icontains=partner)
            if nije_garaza:
                queryset = queryset.filter(nije_garaza=True)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista servisa koje morate dopuniti'
        context['form'] = self.form
        return context


class ServiceCreateView(LoginRequiredMixin, CreateView):
    model = Service
    form_class = ServiceForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('service_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Dodaj servis'
        context['submit_button_label'] = 'Dodaj'
        return context

class ServiceUpdateView(LoginRequiredMixin, UpdateView):
    model = Service
    form_class = ServiceForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('service_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmeni servis'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class ServiceDetailView(LoginRequiredMixin, DetailView):
    model = Service
    template_name = 'fleet/service_detail.html'
    context_object_name = 'service'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f"Detalji servisa {self.object.service_date}"
        return context

class ServiceDeleteView(LoginRequiredMixin, DeleteView):
    model = Service
    success_url = reverse_lazy('service_list')
    template_name = 'fleet/service_confirm_delete.html'
    context_object_name = 'service'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obriši servis'
        return context

    def get_object(self, queryset=None):
        return super().get_object(queryset)
    


# <!-- ======================================================================================== -->
#                           <!-- SERVICE TRANSACTIONS -->
# <!-- ======================================================================================== -->
class ServiceTransactionListView(ListView):
    model = ServiceTransaction
    template_name = 'fleet/service_transactions_list.html'
    context_object_name = 'service_transactions'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista servisa - popravke van IMS-a'
        return context

class ServiceTransactionFixingListView(LoginRequiredMixin, ListView):
    model = DraftServiceTransaction
    template_name = 'fleet/draft_service_transactions_list.html'
    context_object_name = 'service_transactions'

    def get_queryset(self):
            queryset = DraftServiceTransaction.objects.select_related('vehicle').all()
            self.form = ServiceFixingFilterForm(self.request.GET)

            if self.form.is_valid():
                datum_od = self.form.cleaned_data.get('datum_od')
                datum_do = self.form.cleaned_data.get('datum_do')
                vozilo = self.form.cleaned_data.get('vozilo')
                partner = self.form.cleaned_data.get('partner')
                nije_garaza = self.form.cleaned_data.get('nije_garaza')

                if datum_od:
                    queryset = queryset.filter(datum__gte=datum_od)
                if datum_do:
                    queryset = queryset.filter(datum__lte=datum_do)
                if vozilo:
                    queryset = queryset.filter(vehicle__registration_number__icontains=vozilo)
                if partner:
                    queryset = queryset.filter(naz_par_pl__icontains=partner)
                if nije_garaza:
                    queryset = queryset.filter(nije_garaza=True)

            return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Lista servisa koje morate dopuniti'
        context['form'] = self.form
        return context

class ServiceTransactionCreateView(CreateView):
    model = ServiceTransaction
    form_class = ServiceTransactionForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('service_transaction_list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Kreiranje servisa'
        context['submit_button_label'] = 'Sačuvaj insformacije o servisu'
        return context
    
class ServiceTransactionUpdateView(UpdateView):
    model = ServiceTransaction
    form_class = ServiceTransactionForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('service_transaction_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmena servisa'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

class ServiceTransactionDeleteView(DeleteView):
    model = ServiceTransaction
    template_name = 'service_transaction_confirm_delete.html'
    success_url = reverse_lazy('service_transaction_list')

class DraftServiceTransactionUpdateView(UpdateView):
    model = DraftServiceTransaction
    form_class = DraftServiceTransactionForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('service_fixing_list')  # Preusmeri nakon uspešne migracije

    def form_valid(self, form):
        # Sačuvaj izmene u draft tabeli
        draft = form.save(commit=False)
        print("Izmene sačuvane u draft tabeli.")

        # Proveri da li su svi potrebni podaci sada prisutni osim `kom` i `napomena`
        is_complete = all([
            draft.vehicle_id is not None,
            draft.god is not None,
            draft.sif_par_pl not in [None, ''],
            draft.naz_par_pl not in [None, ''],
            draft.datum is not None,
            draft.sif_vrs not in [None, ''],
            draft.br_naloga not in [None, ''],
            draft.vez_dok not in [None, ''],
            draft.knt_pl not in [None, ''],
            draft.potrazuje is not None,
            draft.sif_par_npl not in [None, ''],
            draft.knt_npl not in [None, ''],
            draft.duguje is not None,
            draft.konto_vozila not in [None, ''],
            draft.popravka_kategorija not in [None, '']
        ])


        # Ako su svi potrebni podaci prisutni, pokreni migraciju u glavnu tabelu
        if is_complete:
            draft.save()
            migrate_draft_to_service_transaction(draft.id)
            print("Podaci migrirani u glavnu tabelu.")
            messages.success(self.request, "✅ Podaci su uspešno migrirani u glavnu tabelu.")
            return redirect(self.success_url)
        
        # Ako podaci nisu kompletni, sačuvaj samo u draft tabeli
        else:
            print("Podaci nisu kompletni, ostaju u draft tabeli.")
            messages.warning(self.request, "⚠️ Podaci nisu kompletni, ostaju u draft tabeli.")
            draft.save()  # Sačuvaj bez migracije
            return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Dopunite informacije o servisu'
        context['submit_button_label'] = 'Sačuvaj insformacije o servisu'
        return context

# <!-- ======================================================================================== -->
#                           <!-- REQUISTION - TREBOVANJA -->
# <!-- ======================================================================================== -->

class RequisitionListView(ListView):
    model = Requisition
    template_name = 'fleet/requisition_list.html'
    context_object_name = 'requisitions'
   
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Trebovanja'
        return context

class RequisitionDetailView(ListView):
    model = Requisition
    template_name = 'fleet/requisition_detail.html'
    context_object_name = 'stavke'

    def get_queryset(self):
        return Requisition.objects.filter(
            br_dok=self.kwargs['br_dok'],
            god=self.kwargs['god']
        ).order_by('stavka')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['br_dok'] = self.kwargs['br_dok']
        context['god'] = self.kwargs['god']
        return context

# Draft    
class RequisitionFixingListView(ListView): 
    model = DraftRequisition
    template_name = 'fleet/draft_requisition_list.html'
    context_object_name = 'requisitions'
    
    def get_queryset(self):
        return DraftRequisition.objects.filter(nije_garaza=False)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Trebovanja koja je potrebno dopuniti'
        return context
    
class RequisitionCreateView(CreateView):
    model = Requisition
    form_class = RequisitionForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('requisition_list')
    success_message = "Requisition successfully created."

class RequisitionUpdateView(UpdateView):
    model = Requisition
    form_class = RequisitionForm
    template_name = 'fleet/generic_form.html'
    success_url = reverse_lazy('requisition_list')
    success_message = "Trebovanje uspešno izmenjeno!"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Izmena trebovanja'
        context['submit_button_label'] = 'Sačuvaj izmene'
        return context

from django.urls import reverse

class DraftRequisitionUpdateView(UpdateView):
    model = DraftRequisition
    form_class = DraftRequisitionForm
    template_name = 'fleet/generic_form_draft.html'
    success_message = "Trebovanje uspešno izmenjeno!"

    def form_valid(self, form):
        current_instance = form.save()

        # Ažuriraj ostale redove
        DraftRequisition.objects.filter(
            br_dok=current_instance.br_dok,
            god=current_instance.god
        ).exclude(id=current_instance.id).update(
            vehicle=current_instance.vehicle,
            datum_trebovanja=current_instance.datum_trebovanja,
            mesec_unosa=current_instance.mesec_unosa,
            popravka_kategorija=current_instance.popravka_kategorija,
            kilometraza=current_instance.kilometraza,
            nije_garaza=current_instance.nije_garaza,
            napomena=current_instance.napomena
        )

        # Obradi zapise i premesti one koji su kompletni
        draft_requisitions = DraftRequisition.objects.filter(
            br_dok=current_instance.br_dok,
            god=current_instance.god
        )

        for draft in draft_requisitions:
            print(f"Obrada: {draft}, kompletan: {draft.is_complete()}")
            if draft.is_complete():
                print("→ Premeštam u Requisition")
                Requisition.objects.create(
                    vehicle=draft.vehicle,
                    sif_pred=draft.sif_pred,
                    god=draft.god,
                    br_dok=draft.br_dok,
                    sif_vrsart=draft.sif_vrsart,
                    stavka=draft.stavka,
                    sif_art=draft.sif_art,
                    naz_art=draft.naz_art,
                    kol=draft.kol,
                    cena=draft.cena,
                    vrednost_nab=draft.vrednost_nab,
                    popravka_kategorija=draft.popravka_kategorija,
                    mesec_unosa=draft.mesec_unosa,
                    kilometraza=draft.kilometraza,
                    nije_garaza=draft.nije_garaza,
                    datum_trebovanja=draft.datum_trebovanja,
                    napomena=draft.napomena
                )
                draft.delete()

        # Proveri da li još uvek ima nedovršenih zapisa
        ostali_draftovi = DraftRequisition.objects.filter(
            br_dok=current_instance.br_dok,
            god=current_instance.god
        ).exists()

        # Pozovi funkciju za brisanje kompletnih zapisa
        delete_complete_drafts()

        if ostali_draftovi:
            return redirect('requisition_fixing_list')  # zameni sa stvarnim imenom URL-a
        else:
            return redirect('requisition_detail', god=current_instance.god, br_dok=current_instance.br_dok)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Izmena trebovanja {self.object.br_dok}'
        context['submit_button_label'] = 'Sačuvaj izmene'
        context['manual'] = 'Kilometražu je poželjno uneti uvek, ali nije obavezno. ' \
        'Kada se radi o redovnim servisima, kilometraža je obavezna.'
        return context

class RequisitionDeleteView(DeleteView):
    model = Requisition
    template_name = 'requisition/requisition_confirm_delete.html'
    success_url = reverse_lazy('requisition_list')
    success_message = "Requisition successfully deleted."


# <!-- ======================================================================================== -->
#                                     <!-- USERS -->
# <!-- ======================================================================================== -->

class UserListView(ListView):
    model = CustomUser
    template_name = 'fleet/user_list.html'  # Specify your template
    context_object_name = 'users'     # The name of the variable to use in the template

    # Optionally, you can override get_queryset to filter users if needed
    def get_queryset(self):
        # You can apply any filters if needed, otherwise return all users
        return CustomUser.objects.all()
    

    
# <!-- ======================================================================================== -->
#                           <!-- FETCHING FUNCTIONS -->
# <!-- ======================================================================================== -->

from django.shortcuts import render, redirect

from django.core.management import call_command
from django.http import JsonResponse
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def fetch_data_view(request):
    """
    View za prikaz HTML stranice koja sadrži sve fetching forme.
    """
    if request.method == 'POST':
        command = request.POST.get('command')
        try:
            # Provera i pokretanje komande
            if command == 'nis_command':
                call_command('nis_command')
            elif command == 'omv_command_putnicka':
                call_command('omv_command_putnicka')
            elif command == 'omv_command_teretna':
                call_command('omv_command_teretna')
            else:
                return JsonResponse({'status': 'error', 'message': 'Nepoznata komanda.'}, status=400)
            
            return JsonResponse({'status': 'success', 'message': f'Komanda {command} uspešno izvršena.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    # Prikaz stranice sa svim fetching formama
    return render(request, 'fleet/fetch_data.html')

# @staff_member_required
# def run_nis_command_view(request):
#     if request.method == "POST":
#         def run_command_in_background():
#             execute_nis_command()
        
#         # Pokreće komandu u pozadini
#         thread = threading.Thread(target=run_command_in_background)
#         thread.start()
        
#         return JsonResponse({"status": "success", "message": "Komanda pokrenuta u pozadini."})
#     return JsonResponse({"status": "error", "message": "Nevalidan zahtev."})

# POVLACENJE PODATAKA IZ DRUGE BAZE
logger = logging.getLogger(__name__)  
def fetch_vehicle_value_view(request):
    if request.method == 'POST':
        # Povlačenje podataka iz druge baze
        with connections['test_db'].cursor() as cursor:
            cursor.execute("""
                SELECT sif_osn, vrednost FROM dbo.vrednost_vozila
            """)
            rows = cursor.fetchall()

        # Broj ažuriranih vozila
        updated_vehicles_count = 0

        for row in rows:
            sif_osn = row[0].strip()  # Polje 'sif_osn' iz druge baze (odgovara 'inventory_number' u modelu Vehicle)
            vrednost = row[1]  # Polje 'vrednost' iz druge baze (odgovara polju 'value' u modelu Vehicle)

            try:
                # Pronađi vozilo po inventory_number (sif_osn)
                print(sif_osn)
                vehicle = Vehicle.objects.get(inventory_number=sif_osn)
                #print(vehicle)
                # Ažuriraj polje 'value' sa novom vrednošću
                vehicle.value = vrednost
                vehicle.save()
                updated_vehicles_count += 1

            except Vehicle.DoesNotExist:
                # Ako vozilo sa datim 'sif_osn' ne postoji, preskoči i zapiši grešku
                logger.warning(f"Vozilo sa inventory_number (sif_osn) {sif_osn} nije pronađeno.")
                continue

            except Exception as e:
                # U slučaju bilo koje druge greške
                logger.error(f"Greška prilikom ažuriranja vozila sa inventory_number {sif_osn}: {e}")
                messages.error(request, "Došlo je do greške prilikom ažuriranja podataka o vozilu.")
                return redirect('fetch_policies')

        # Poruka o uspešnom ažuriranju
        messages.success(request, f"Uspešno ažurirano {updated_vehicles_count} vozila.")

        # Preusmeravanje nakon uspešnog povlačenja podataka
        return redirect('fetch_policies')  # Postavi URL na koji želiš da preusmeriš korisnika

    return render(request, 'fleet/fetch_data.html')


# LEASE KAMATE FETCH
def fetch_lease_interest_data(request):
    if request.method == 'POST':
        # Povlačenje podataka iz view-a u bazi
        with connections['test_db'].cursor() as cursor:
            cursor.execute("""
                SELECT god, ugovor, iznos FROM dbo.lizing_kamate
            """)
            rows = cursor.fetchall()

        for row in rows:
            year = row[0]
            contract_number = row[1].strip()
            interest_amount = row[2]

            try:
                # Pronađi ugovor lizinga po broju ugovora
                lease = Lease.objects.get(contract_number=contract_number)

                # Proveri da li već postoji zapis za tu godinu i lizing ugovor
                lease_interest, created = LeaseInterest.objects.get_or_create(
                    lease=lease,
                    year=year,
                    defaults={'interest_amount': interest_amount}
                )

                if not created:
                    # Ako zapis već postoji, možeš ga ažurirati ako je potrebno
                    lease_interest.interest_amount = interest_amount
                    lease_interest.save()

            except Lease.DoesNotExist:
                logger.warning(f"Lizing ugovor sa brojem {contract_number} nije pronađen.")
                continue

        # Nakon uspešne obrade, preusmeravanje ili prikaz poruke
        return redirect('fetch_policies')  # Preusmeri na odgovarajući URL za prikaz lizing kamata

    return render(request, 'fleet/fetch_data.html')

def fetch_policy_data_view(request):
    if request.method == 'POST':
        days = request.POST.get('days')
        result = None
        
        try:
            if days:
                days = int(days)
                if days < 0:
                    raise ValueError
                result = fetch_policy_data(last_24_hours=False, days=days)
            else:
                result = fetch_policy_data()
            
            if result.startswith('Critical error'):
                messages.error(request, result)
            else:
                messages.success(request, result)
                
        except ValueError:
            messages.error(request, "Invalid number of days")
        
        return redirect('policy_list')

    return render(request, 'fleet/fetch_policy_data.html')


# POVLACENJE PODATAKA IZ DRUGE BAZE
def fetch_service_data_view(request):
    if request.method == 'POST':
        # Preuzmi broj dana iz POST zahteva (opciono)
        days = request.POST.get('days', None)
        
        # Ako korisnik unese broj dana, pretvori ga u integer
        if days:
            try:
                days = int(days)
            except ValueError:
                messages.error(request, "Uneta vrednost za broj dana nije validna.")
                return redirect('fetch_service_data')  # Vrati korisnika na stranicu sa greškom

        # Pozovi funkciju za povlačenje podataka sa odgovarajućim parametrima
        result = fetch_service_data(last_24_hours=False, days=days)
        messages.success(request, result)
        return redirect('service_transaction_list')

    # Prikaz forme za unos broja dana
    return render(request, 'fleet/fetch_service_data.html')


def fetch_requisition_data_view(request):
    if request.method == 'POST':
        days = request.POST.get('days', None)
        if days:
            try:
                days = int(days)
            except ValueError:
                messages.error(request, "Uneta vrednost za broj dana nije validna.")
                return redirect('fetch_policies')

        result = fetch_requisition_data(last_24_hours=False, days=days)
        messages.success(request, result)
        return redirect('requisition_list')

    return render(request, 'fleet/fetch_data.html')


# <!-- ======================================================================================== -->
#                           <!-- IZVESTAJI -->
# <!-- ======================================================================================== -->

def reports_index(request):
    """Početna stranica za izveštaje sa linkovima."""
    sections = {
        "Finansije": [
            {"name": "Spisak vozila po šiframa posla", "url": "vehicle_list"},
            {"name": "Pregled potrošnje goriva po šiframa posla - OMV putnicka", "url": "omv_putnicka"},
            {"name": "Pregled potrošnje goriva po šiframa posla - OMV teretna", "url": "omv_teretna"},
            {"name": "Pregled potrošnje goriva po šiframa posla - NIS putnicka", "url": "nis_putnicka"},
            {"name": "Pregled potrošnje goriva po šiframa posla - NIS teretna", "url": "nis_teretna"},
        ],
        "Centri": [
            # {"name": "Pregled putnih naloga po godinama", "url": "kasko_rate"},
            {"name": "Zatvoreni putni nalozi", "url": "zatvoreni_putni"},
        ],
        "Garaža": [
            {"name": "Trenutno stanje u magacinu", "url": "magacin"},
            {"name": "Spisak otpisanih vozila", "url": "otpis"},
            # {"name": "Trenutno stanje u magacinu", "url": "tro_gorivo_mesec"},
        ],
        "Uprava": [
            {"name": "Promet goriva po mesecima", "url": "tro_gorivo_mesec"},
            {"name": "Pregled ukupnih troskova, pa po kontima, pa po centrima, po mesecima ", "url": "troskovi_svi"},
            {"name": "Troškovi praćenja vozila", "url": "tro_pracenja_vozila"},
            {"name": "Troškovi tahografa ", "url": "troskovi_tahograf"},
            {"name": "Troškovi parkinga", "url": "tro_parking"},
            {"name": "Pregled Potraživanja od osiguranja", "url": "potrazivanje_ddor"},
            {"name": "Pregled Najvećih Dobavljača Usluga", "url": "po_dobavljacima"},
            
        ],
    }

    return render(request, 'fleet/reports_index.html', {"sections": sections})

def omv_putnicka_view(request):
    form = OMVPutnickaFilterForm(request.GET or None)

    query = """
        SELECT sifpos, godina, mesec, tipvozila, polovina, bruto, neto
        FROM OMV_putnicka_sp
        WHERE 1=1
    """

    filters = []
    params = []

    if form.is_valid():
        godina = form.cleaned_data.get('godina')
        mesec = form.cleaned_data.get('mesec')
        polovina = form.cleaned_data.get('polovina')

        if godina:
            filters.append("AND godina = %s")
            params.append(godina)

        if mesec:
            filters.append("AND mesec = %s")
            params.append(mesec)

        if polovina:
            filters.append("AND polovina = %s")
            params.append(polovina)

    query += " " + " ".join(filters)

    data = get_data_from_secondary_db(query, 'test_db', params=params)

    # Excel export
    if 'export' in request.GET:
        import pandas as pd
        from django.http import HttpResponse

        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=omv_putnicka.xlsx'
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='OMV Putnicka')
        return response

    return render(request, 'fleet/reports/omv_putnicka.html', {
        'data': data,
        'form': form,
        'title': 'OMV Putnička vozila'
    })


def export_omv_putnicka_excel(request):
    form = PutnickaFilterForm(request.GET or None)

    query = """
        SELECT sifpos, godina, mesec, tipvozila, polovina, bruto, neto
        FROM OMV_putnicka_sp
        WHERE 1=1
    """

    filters = []
    params = []

    if form.is_valid():
        godina = form.cleaned_data.get('godina')
        mesec = form.cleaned_data.get('mesec')
        polovina = form.cleaned_data.get('polovina')

        if godina:
            filters.append("AND godina = %s")
            params.append(int(godina))

        if mesec:
            filters.append("AND mesec = %s")
            params.append(int(mesec))

        if polovina:
            filters.append("AND polovina = %s")
            params.append(int(polovina))

    if filters:
        query += " " + " ".join(filters)

    data = get_data_from_secondary_db(query, 'test_db', params=params)

    wb = Workbook()
    ws = wb.active
    ws.title = "OMV Putnička"

    # Header
    headers = ["Šifra pos", "Godina", "Mesec", "Tip vozila", "Polovina", "Bruto", "Neto"]
    ws.append(headers)

    # Rows
    for row in data:
        ws.append([
            row['sifpos'],
            row['godina'],
            row['mesec'],
            row['tipvozila'],
            row['polovina'],
            row['bruto'],
            row['neto']
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=omv_putnicka.xlsx'
    wb.save(response)
    return response

def nis_putnicka_view(request):
    form = PutnickaFilterForm(request.GET or None)

    query = """
        SELECT tipvozila, sifpos, godina, mesec, polovina, bruto, neto
        FROM dbo.NIS_putnicka_sp
        WHERE 1=1
    """

    filters = []
    params = []

    if form.is_valid():
        godina = form.cleaned_data.get('godina')
        mesec = form.cleaned_data.get('mesec')
        polovina = form.cleaned_data.get('polovina')

        if godina:
            filters.append("AND godina = %s")
            params.append(godina)

        if mesec:
            filters.append("AND mesec = %s")
            params.append(mesec)

        if polovina:
            filters.append("AND polovina = %s")
            params.append(polovina)

    if filters:
        query += " " + " ".join(filters)

    data = get_data_from_secondary_db(query, 'test_db', params=params)

    # Excel export
    if 'export' in request.GET:
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=nis_putnicka.xlsx'
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='NIS Putnicka')
        return response

    return render(request, 'fleet/reports/nis_putnicka.html', {
        'data': data,
        'form': form,
        'title': 'NIS Putnička vozila'
    })

def export_nis_putnicka_excel(request):
    form = PutnickaFilterForm(request.GET or None)

    query = """
        SELECT tipvozila, sifpos, godina, mesec, polovina, bruto, neto
        FROM dbo.NIS_putnicka_sp
        WHERE 1=1
    """

    filters = []
    params = []

    if form.is_valid():
        godina = form.cleaned_data.get('godina')
        mesec = form.cleaned_data.get('mesec')
        polovina = form.cleaned_data.get('polovina')

        if godina:
            filters.append("AND godina = %s")
            params.append(int(godina))

        if mesec:
            filters.append("AND mesec = %s")
            params.append(int(mesec))

        if polovina:
            filters.append("AND polovina = %s")
            params.append(int(polovina))

    if filters:
        query += " " + " ".join(filters)

    data = get_data_from_secondary_db(query, 'test_db', params=params)

    wb = Workbook()
    ws = wb.active
    ws.title = "NIS Putnička"

    headers = ["Tip vozila", "Šifra pos", "Godina", "Mesec", "Polovina", "Bruto", "Neto"]
    ws.append(headers)

    for row in data:
        ws.append([
            row['tipvozila'],
            row['sifpos'],
            row['godina'],
            row['mesec'],
            row['polovina'],
            row['bruto'],
            row['neto']
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=nis_putnicka.xlsx'
    wb.save(response)
    return response

def nis_teretna_view(request):
    form = PutnickaFilterForm(request.GET or None)

    query = """
        SELECT tipvozila, sifpos, regozn, kartica, datum, proizvod, kolicina, cena, bruto, neto
        FROM dbo.nis_teretna
        WHERE 1=1
    """

    filters = []
    params = []

    if form.is_valid():
        godina = form.cleaned_data.get('godina')
        mesec = form.cleaned_data.get('mesec')
        polovina = form.cleaned_data.get('polovina')

        if godina:
            filters.append("AND YEAR(datum) = %s")
            params.append(godina)

        if mesec:
            filters.append("AND MONTH(datum) = %s")
            params.append(mesec)

        if polovina:
            if polovina == 1:
                filters.append("AND DAY(datum) <= 15")
            elif polovina == 2:
                filters.append("AND DAY(datum) > 15")

    if filters:
        query += " " + " ".join(filters)

    data = get_data_from_secondary_db(query, 'test_db', params=params)

    # Excel export
    if 'export' in request.GET:
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=nis_teretna.xlsx'
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='NIS Teretna')
        return response

    return render(request, 'fleet/reports/nis_teretna.html', {
        'data': data,
        'form': form,
        'title': 'NIS Teretna vozila'
    })



def export_nis_teretna_excel(request):
    form = PutnickaFilterForm(request.GET or None)

    query = """
        SELECT tipvozila, sifpos, regozn, kartica, datum, proizvod, kolicina, cena, bruto, neto
        FROM dbo.nis_teretna
        WHERE 1=1
    """

    filters = []
    params = []

    if form.is_valid():
        godina = form.cleaned_data.get('godina')
        mesec = form.cleaned_data.get('mesec')
        polovina = form.cleaned_data.get('polovina')

        if godina:
            filters.append("AND YEAR(datum) = %s")
            params.append(int(godina))

        if mesec:
            filters.append("AND MONTH(datum) = %s")
            params.append(int(mesec))

        if polovina:
            if int(polovina) == 1:
                filters.append("AND DAY(datum) <= 15")
            elif int(polovina) == 2:
                filters.append("AND DAY(datum) > 15")

    if filters:
        query += " " + " ".join(filters)

    data = get_data_from_secondary_db(query, 'test_db', params=params)

    wb = Workbook()
    ws = wb.active
    ws.title = "NIS Teretna"

    headers = ["Tip vozila", "Šifra posla", "Reg oznaka", "Kartica", "Datum", "Proizvod", "Količina", "Cena", "Bruto", "Neto"]
    ws.append(headers)

    for row in data:
        ws.append([
            row['tipvozila'],
            row['sifpos'],
            row['regozn'],
            row['kartica'],
            row['datum'].strftime("%d.%m.%Y %H:%M") if row['datum'] else '',
            row['proizvod'],
            row['kolicina'],
            row['cena'],
            row['bruto'],
            row['neto']
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=nis_teretna.xlsx'
    wb.save(response)
    return response


def omv_teretna_view(request):
    form = PutnickaFilterForm(request.GET or None)

    query = """
        SELECT tipvozila, sifpos, godina, mesec, polovina, bruto, neto
        FROM dbo.OMV_teretna_sp
        WHERE 1=1
    """

    filters = []
    params = []

    if form.is_valid():
        godina = form.cleaned_data.get('godina')
        mesec = form.cleaned_data.get('mesec')
        polovina = form.cleaned_data.get('polovina')

        if godina:
            filters.append("AND godina = %s")
            params.append(godina)

        if mesec:
            filters.append("AND mesec = %s")
            params.append(mesec)

        if polovina:
            filters.append("AND polovina = %s")
            params.append(polovina)

    if filters:
        query += " " + " ".join(filters)

    data = get_data_from_secondary_db(query, 'test_db', params=params)

    # Excel export
    if 'export' in request.GET:
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=omv_teretna.xlsx'
        with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='OMV Teretna')
        return response

    return render(request, 'fleet/reports/omv_teretna.html', {
        'data': data,
        'form': form,
        'title': 'OMV Teretna vozila'
    })

def export_omv_teretna_excel(request):
    form = PutnickaFilterForm(request.GET or None)

    query = """
        SELECT tipvozila, sifpos, godina, mesec, polovina, bruto, neto
        FROM dbo.OMV_teretna_sp
        WHERE 1=1
    """

    filters = []
    params = []

    if form.is_valid():
        godina = form.cleaned_data.get('godina')
        mesec = form.cleaned_data.get('mesec')
        polovina = form.cleaned_data.get('polovina')

        if godina:
            filters.append("AND godina = %s")
            params.append(int(godina))

        if mesec:
            filters.append("AND mesec = %s")
            params.append(int(mesec))

        if polovina:
            filters.append("AND polovina = %s")
            params.append(int(polovina))

    if filters:
        query += " " + " ".join(filters)

    data = get_data_from_secondary_db(query, 'test_db', params=params)

    wb = Workbook()
    ws = wb.active
    ws.title = "OMV Teretna"

    headers = ["Tip vozila", "Šifra pos", "Godina", "Mesec", "Polovina", "Bruto", "Neto"]
    ws.append(headers)

    for row in data:
        ws.append([
            row['tipvozila'],
            row['sifpos'],
            row['godina'],
            row['mesec'],
            row['polovina'],
            row['bruto'],
            row['neto']
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=omv_teretna.xlsx'
    wb.save(response)
    return response

def get_data_from_secondary_db(query, db_alias, params=None):
    """
    Izvršava SQL upit na drugoj bazi i vraća rezultat kao listu rečnika.
    """
    with connections[db_alias].cursor() as cursor:
        cursor.execute(query, params or [])
        columns = [col[0] for col in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]

def kasko_rate_view(request):
    """
    View za prikaz podataka iz dbo.kasko_rate.
    """
    query = "SELECT * FROM dbo.kasko_rate"
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/kasko_rate.html', {'data': data})

	
def zatvoren_putni_view(request):
    """
    View za prikaz podataka iz dbo.zatvoren_putni.
    """
    query = "SELECT * FROM dbo.zatvoren_putni"
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/zatvoreni_putni.html', {'data': data})


def magacin_view(request):
    """
    View za prikaz podataka iz dbo.magacin_garaza.
    """
    query = """
        SELECT sif_pred, god, oj, sif_mag, sif_art, kolul, koliz, popkol, vrulnab, vriznab,
               vrulvp, vrizvp, revalzal, razliz, mag_cena, kolpon, cenapon, naz_art,
               sif_vrsart, naz_vrsart
        FROM dbo.magacin_garaza
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/magacin.html', {'data': data})


def otpis_view(request):
    """
    View za prikaz podataka iz dbo.otpis.
    """
    query = """
        SELECT sif_pred, god, sif_osn, rb, naz_osn, inv_br, kol, jed_mere, sif_par, knt, oj, sif_lok,
               sif_amort, sif_reval, stopa_dogam, dat_stavlj, dat_prest, iznos_val, skr_naz, poreklo,
               nab_vred, osnovica, otpis, status, br_fakture, zemljiste_ar, zemljiste_m, u_gramima,
               sif_amortP, sif_revalP, otpisP, otudjena_vrednost, ind_trosak, opis, osnovicaP,
               ind_manjak, ind_amort, knt_ispravka, sif_kor, stopa_amort
        FROM dbo.otpis
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/otpis.html', {'data': data})

def tro_gorivo_mesec_view(request):
    """
    View za prikaz podataka iz dbo.TroGorivoMesec.
    """
    query = """
        SELECT god, mesec, kategorija, iznos
        FROM dbo.TroGorivoMesec
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/tro_gorivo_mesec.html', {'data': data})

def troskovi_svi_view(request):
    """
    View za prikaz podataka iz dbo.Troskovi_svi.
    """
    query = """
        SELECT god, sif_vrs, datum, br_naloga, stavka, oj, knt, naz_knt, duguje, sif_pos
        FROM dbo.Troskovi_svi
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/troskovi_svi.html', {'data': data})

def tro_pracenja_vozila_view(request):
    """
    View za prikaz podataka iz dbo.TroPracenjaVozila.
    """
    query = """
        SELECT PartnerPIB, PartnerIme, ID, BrojFakture, issuedate, ZaPlacanje, Konto_tro
        FROM dbo.TroPracenjaVozila
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/tro_pracenja_vozila.html', {'data': data})

def tahograf_partneri_view(request):
    """
    View za prikaz podataka iz dbo.TroTahografa.
    """
    query = """
        SELECT *
        FROM dbo.TroTahografa
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/tro_tahografa.html', {'data': data})

def tro_zarade_view(request):
    """
    View za prikaz podataka iz dbo.tro_zarade.
    """
    query = """
        SELECT oj, god, mesec, rasif, ranaz, neto, bruto, bruto2
        FROM dbo.tro_zarade
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/tro_zarade.html', {'data': data})

def tro_parking_view(request):
    """
    View za prikaz podataka iz dbo.tro_parking.
    """
    query = """
        SELECT PartnerPIB, PartnerIme, ID, BrojFakture, issuedate, note, naziv, ZaPlacanje
        FROM dbo.tro_parking
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/tro_parking.html', {'data': data})

def po_dobavljacima_view(request):
    """
    View za prikaz podataka iz dbo.po_dobavljacima.
    """
    query = """
        SELECT naz_par, sif_pred, god, sif_vrs, br_naloga, stavka, oj, knt, grupa, sif_par, datum, vez_dok,
               duguje, potrazuje, skr_naz, deviza, kom, stavka_k, dpo, promena, sif_pos, dat_naloga, d_p, placeno
        FROM dbo.po_dobavljacima
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/po_dobavljacima.html', {'data': data})

def potrazivanje_ddor_view(request):
    """
    View za prikaz podataka iz dbo.potrazivanje_ddor.
    """
    query = """
        SELECT god, sif_vrs, br_naloga, stavka, oj, knt, datum, vez_dok, potrazuje
        FROM dbo.potrazivanje_ddor
    """
    data = get_data_from_secondary_db(query, 'test_db')  # test_db je alias za sekundarnu bazu
    return render(request, 'fleet/reports/potrazivanje_ddor.html', {'data': data})


from io import BytesIO
from types import SimpleNamespace
from unittest.mock import Mock

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.exceptions import PermissionDenied
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .exporting import (
    csv_attachment_response,
    dataframe_xlsx_response,
    rows_to_xlsx_response,
    xlsx_attachment_response,
)
from .mixins import RolePermissionRequiredMixin, role_permission_required, user_has_role_permission
from .models import ActivityLog, CustomUser, OrganizationalUnit, PermissionCode, Role, RolePermission, TaskHistory


class ExportingTests(SimpleTestCase):
    def test_csv_attachment_response_sets_download_headers(self):
        response = csv_attachment_response("vozila.csv")

        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="vozila.csv"')

    def test_xlsx_attachment_response_sets_download_headers(self):
        response = xlsx_attachment_response("report.xlsx")

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(response["Content-Disposition"], "attachment; filename=report.xlsx")

    def test_rows_to_xlsx_response_creates_valid_workbook(self):
        response = rows_to_xlsx_response(
            "report.xlsx",
            "Report",
            ["Name", "Value"],
            [["Alpha", 12]],
            bold_header=True,
            auto_width=True,
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active

        self.assertEqual(worksheet.title, "Report")
        self.assertEqual(worksheet["A1"].value, "Name")
        self.assertTrue(worksheet["A1"].font.bold)
        self.assertEqual(worksheet["A2"].value, "Alpha")
        self.assertEqual(worksheet["B2"].value, 12)

    def test_dataframe_xlsx_response_creates_valid_workbook(self):
        response = dataframe_xlsx_response(
            [{"Name": "Alpha", "Value": 12}],
            "data.xlsx",
            "Data",
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active

        self.assertEqual(worksheet.title, "Data")
        self.assertEqual(worksheet["A1"].value, "Name")
        self.assertEqual(worksheet["A2"].value, "Alpha")
        self.assertEqual(worksheet["B2"].value, 12)


class RolePermissionMixinTests(SimpleTestCase):
    def test_user_has_role_permission_returns_false_without_code(self):
        user = SimpleNamespace(is_authenticated=True, is_superuser=False, roles=Mock())

        self.assertFalse(user_has_role_permission(user, None))
        user.roles.filter.assert_not_called()

    def test_user_has_role_permission_uses_active_role_filter(self):
        filter_mock = Mock()
        filter_mock.exists.return_value = True
        roles = Mock()
        roles.filter.return_value = filter_mock
        user = SimpleNamespace(is_authenticated=True, is_superuser=False, roles=roles)

        self.assertTrue(user_has_role_permission(user, "fleet:vehicle_list"))
        roles.filter.assert_called_once_with(
            permissions__code="fleet:vehicle_list",
            is_active=True,
        )

    def test_role_permission_mixin_uses_resolver_match_view_name(self):
        filter_mock = Mock()
        filter_mock.exists.return_value = True
        roles = Mock()
        roles.filter.return_value = filter_mock
        user = SimpleNamespace(is_authenticated=True, is_superuser=False, roles=roles)
        request = SimpleNamespace(
            user=user,
            resolver_match=SimpleNamespace(view_name="naplata:lista_dugovanja"),
        )

        class TestMixin(RolePermissionRequiredMixin):
            def __init__(self, request):
                self.request = request

        self.assertTrue(TestMixin(request).test_func())
        roles.filter.assert_called_once_with(
            permissions__code="naplata:lista_dugovanja",
            is_active=True,
        )

    def test_role_permission_required_decorator_raises_without_permission(self):
        filter_mock = Mock()
        filter_mock.exists.return_value = False
        roles = Mock()
        roles.filter.return_value = filter_mock
        user = SimpleNamespace(is_authenticated=True, is_superuser=False, roles=roles)
        request = SimpleNamespace(
            user=user,
            resolver_match=SimpleNamespace(view_name="fleet:vehicle_list"),
        )

        @role_permission_required()
        def protected_view(request):
            return "ok"

        with self.assertRaises(PermissionDenied):
            protected_view(request)


class PermissionCodeSyncTests(TestCase):
    def test_sync_permission_codes_grants_employee_sync_to_sekretarijat(self):
        from .permissions import sync_permission_codes

        sync_permission_codes()

        role = Role.objects.get(slug="sekretarijat")
        codes = set(role.permissions.values_list("code", flat=True))
        self.assertIn("employee_list", codes)
        self.assertIn("employee_sync", codes)
        self.assertIn("putninalog_create", codes)
        self.assertIn("putninalog_detail", codes)
        self.assertIn("putninalog_foreign_print", codes)
        self.assertIn("putninalog_print", codes)
        self.assertIn("putninalog_print_list", codes)
        self.assertIn("putninalog_storniraj", codes)
        self.assertIn("putninalog_update", codes)
        self.assertNotIn("putninalog_set_opravdan", codes)

    def test_sync_permission_codes_grants_self_service_permissions_to_zaposleni(self):
        from .permissions import sync_permission_codes

        sync_permission_codes()

        role = Role.objects.get(slug="zaposleni")
        codes = set(role.permissions.values_list("code", flat=True))
        self.assertIn("vehicle_travel_order_create", codes)
        self.assertIn("vehicle_travel_order_detail", codes)
        self.assertIn("vehicle_travel_order_request", codes)
        self.assertIn("vehicle_travel_order_fuel_report", codes)
        self.assertNotIn("putninalog_create", codes)
        self.assertNotIn("vehicle_travel_order_update", codes)

    def test_sync_permission_codes_grants_isplate_permissions_to_blagajna(self):
        from .permissions import sync_permission_codes

        sync_permission_codes()

        role = Role.objects.get(slug="blagajna")
        codes = set(role.permissions.values_list("code", flat=True))
        self.assertIn("isplate:neoporezive_isplate", codes)
        self.assertIn("isplate:converter", codes)

    def test_sync_permission_codes_grants_readonly_naplata_role(self):
        from .permissions import sync_permission_codes

        sync_permission_codes()

        role = Role.objects.get(slug="pregled-naplate")
        codes = set(role.permissions.values_list("code", flat=True))
        self.assertIn("naplata:lista_dugovanja_po_bucketima", codes)
        self.assertIn("naplata:lista_avans_klijenti", codes)
        self.assertIn("naplata:detalji_partner", codes)
        self.assertIn("naplata:export_dugovanja_excel", codes)
        self.assertIn("naplata:pravna_detalj", codes)
        self.assertNotIn("naplata:toggle_avans_klijent", codes)
        self.assertNotIn("naplata:dodaj_kontakt", codes)
        self.assertNotIn("naplata:izmeni_kontakt", codes)
        self.assertNotIn("naplata:obrisi_kontakt", codes)
        self.assertNotIn("naplata:pravna_izmeni", codes)

    def test_sync_permission_codes_links_sekretarijat_group_users_to_role(self):
        from .permissions import sync_permission_codes

        group = Group.objects.create(name="Sekretarijat")
        user = get_user_model().objects.create_user("sekretar", password="test")
        user.groups.add(group)

        result = sync_permission_codes()

        self.assertEqual(result["sekretarijat_group_users_synced"], 1)
        self.assertTrue(user.roles.filter(slug="sekretarijat").exists())

    def test_sync_permission_codes_links_pregled_naplate_group_users_to_role(self):
        from .permissions import sync_permission_codes

        group = Group.objects.create(name="Pregled naplate")
        user = get_user_model().objects.create_user("pregled-naplate-user", password="test")
        user.groups.add(group)

        result = sync_permission_codes()

        self.assertEqual(result["pregled_naplate_group_users_synced"], 1)
        self.assertTrue(user.roles.filter(slug="pregled-naplate").exists())


class OrganizationalUnitLocationTests(SimpleTestCase):
    def test_organizational_unit_keeps_fleet_app_label(self):
        self.assertEqual(OrganizationalUnit._meta.app_label, "fleet")

    def test_organizational_unit_is_reexported_from_fleet_models(self):
        from fleet.models import OrganizationalUnit as FleetOrganizationalUnit

        self.assertIs(FleetOrganizationalUnit, OrganizationalUnit)


class FleetAuthModelLocationTests(SimpleTestCase):
    def test_fleet_auth_models_keep_fleet_app_label(self):
        self.assertEqual(ActivityLog._meta.app_label, "fleet")
        self.assertEqual(CustomUser._meta.app_label, "fleet")
        self.assertEqual(Role._meta.app_label, "fleet")
        self.assertEqual(PermissionCode._meta.app_label, "fleet")
        self.assertEqual(RolePermission._meta.app_label, "fleet")
        self.assertEqual(TaskHistory._meta.app_label, "fleet")

    def test_fleet_auth_models_are_reexported_from_fleet_models(self):
        from fleet.models import (
            ActivityLog as FleetActivityLog,
            CustomUser as FleetCustomUser,
            PermissionCode as FleetPermissionCode,
            Role as FleetRole,
            RolePermission as FleetRolePermission,
            TaskHistory as FleetTaskHistory,
        )

        self.assertIs(FleetActivityLog, ActivityLog)
        self.assertIs(FleetCustomUser, CustomUser)
        self.assertIs(FleetRole, Role)
        self.assertIs(FleetPermissionCode, PermissionCode)
        self.assertIs(FleetRolePermission, RolePermission)
        self.assertIs(FleetTaskHistory, TaskHistory)


class ActivityLogTests(TestCase):
    def test_log_activity_stores_actor_snapshot(self):
        from core.activity import log_activity

        user = get_user_model().objects.create_user(
            username="admin-test",
            first_name="Admin",
            last_name="Test",
            password="test",
        )

        log = log_activity(
            user=user,
            action=ActivityLog.ACTION_MANUAL,
            description="Test aktivnost",
            app_label="fleet",
            view_name="activity_log_list",
        )

        self.assertEqual(log.user, user)
        self.assertEqual(log.actor_username, "admin-test")
        self.assertEqual(log.actor_display_name, "Admin Test")
        self.assertEqual(log.description, "Test aktivnost")

    def test_activity_log_list_renders_for_superuser(self):
        user = get_user_model().objects.create_superuser(
            username="super-admin",
            email="admin@example.com",
            password="test-pass",
        )
        ActivityLog.objects.create(
            user=user,
            actor_username=user.username,
            action=ActivityLog.ACTION_MANUAL,
            description="Provera ekrana",
            app_label="fleet",
        )

        self.client.force_login(user)
        response = self.client.get(reverse("activity_log_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Provera ekrana")

    def test_activity_log_list_shows_user_summary(self):
        admin = get_user_model().objects.create_superuser(
            username="activity-admin",
            email="activity-admin@example.com",
            password="test-pass",
        )
        active_user = get_user_model().objects.create_user("aktivni", password="test-pass")
        quiet_user = get_user_model().objects.create_user("tihi", password="test-pass")
        ActivityLog.objects.create(
            user=active_user,
            actor_username=active_user.username,
            action=ActivityLog.ACTION_REQUEST,
            description="Prva akcija",
            status_code=200,
        )
        ActivityLog.objects.create(
            user=active_user,
            actor_username=active_user.username,
            action=ActivityLog.ACTION_REQUEST,
            description="Druga akcija",
            status_code=404,
        )
        ActivityLog.objects.create(
            user=quiet_user,
            actor_username=quiet_user.username,
            action=ActivityLog.ACTION_LOGIN,
            description="Prijava",
        )

        self.client.force_login(admin)
        response = self.client.get(reverse("activity_log_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rekapitulacija po korisniku")
        summary = response.context["user_activity_summary"]
        active_row = next(row for row in summary if row["actor_username"] == "aktivni")
        self.assertEqual(active_row["total"], 2)
        self.assertEqual(active_row["request_count"], 2)
        self.assertEqual(active_row["error_count"], 1)


class TaskHistoryTests(TestCase):
    def test_sync_summary_with_skipped_counts_is_success(self):
        from ims_erp.celery import _task_status_from_result

        status = _task_status_from_result(
            "SUCCESS",
            "Sync HR Employees: ukupno=837, kreirano=0, azurirano=332, "
            "azurirano_neaktivni=34, preskoceno_neaktivni=471",
        )

        self.assertEqual(status, TaskHistory.STATUS_SUCCESS)

    def test_putni_nalozi_sync_summary_with_skipped_counts_is_success(self):
        from ims_erp.celery import _task_status_from_result

        status = _task_status_from_result(
            "SUCCESS",
            "Sync isplaceno putni nalozi: view_redova=2795, pronadjeno_naloga=2397, "
            "azurirano=47, preskoceno_bez_promene=2350",
        )

        self.assertEqual(status, TaskHistory.STATUS_SUCCESS)

    def test_generic_sync_summary_with_preskoceno_is_success(self):
        from ims_erp.celery import _task_status_from_result

        summaries = [
            "Fetch Service Data: Servisi sync: povuceno=0, kreirano=0, preskoceno=0, problemi=0",
            "Fetch DDOR Insurance Data: DDOR sync: povuceno=30, kreirano=0, preskoceno=30, problemi=0",
            "Fetch Requisition Data: Trebovanja sync: povuceno=0, kreirano=0, azurirano=0, bez_vozila=0, preskoceno=0, problemi=0",
            "NIS sync zavrsen. Gorivo: redova 398, upisano 10, preskoceno 388. Transakcije: redova 398, upisano 10, preskoceno 388.",
        ]

        for summary in summaries:
            with self.subTest(summary=summary):
                self.assertEqual(
                    _task_status_from_result("SUCCESS", summary),
                    TaskHistory.STATUS_SUCCESS,
                )

    def test_explicit_skip_result_is_skipped(self):
        from ims_erp.celery import _task_status_from_result

        self.assertEqual(
            _task_status_from_result("SUCCESS", "SKIP: task 'sync_hr_employees_task' je vec aktivan."),
            TaskHistory.STATUS_SKIPPED,
        )

    def test_task_history_list_renders_for_superuser(self):
        user = get_user_model().objects.create_superuser(
            username="task-admin",
            email="task-admin@example.com",
            password="test-pass",
        )
        TaskHistory.objects.create(
            task_id="task-1",
            task_name="fleet.tasks.sync_hr_employees_task",
            display_name="Kadrovi - sinhronizacija zaposlenih",
            status=TaskHistory.STATUS_SUCCESS,
            short_message="Sync HR Employees: ukupno=1, kreirano=1, azurirano=0, preskoceno_neaktivni=0",
        )

        self.client.force_login(user)
        response = self.client.get(reverse("task_history_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Kadrovi - sinhronizacija zaposlenih")
        self.assertContains(response, "kreirano=1")

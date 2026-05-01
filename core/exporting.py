from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def attachment_response(filename, content_type, quoted=False):
    response = HttpResponse(content_type=content_type)
    disposition_filename = f'"{filename}"' if quoted else filename
    response["Content-Disposition"] = f"attachment; filename={disposition_filename}"
    return response


def csv_attachment_response(filename, charset="utf-8", quoted=True):
    content_type = "text/csv"
    if charset:
        content_type = f"{content_type}; charset={charset}"
    return attachment_response(filename, content_type, quoted=quoted)


def xlsx_attachment_response(filename, quoted=False):
    return attachment_response(filename, XLSX_CONTENT_TYPE, quoted=quoted)


def workbook_response(workbook, filename, quoted=False):
    response = xlsx_attachment_response(filename, quoted=quoted)
    workbook.save(response)
    return response


def create_xlsx_workbook(sheet_title=None):
    workbook = Workbook()
    worksheet = workbook.active
    if sheet_title:
        worksheet.title = sheet_title
    return workbook, worksheet


def style_header_row(worksheet, row_number=1):
    header_font = Font(bold=True)
    for cell in worksheet[row_number]:
        cell.font = header_font


def set_column_widths(worksheet, widths):
    for col_num, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = width


def set_uniform_column_width(worksheet, width):
    set_column_widths(worksheet, [width] * worksheet.max_column)


def autofit_columns(worksheet, min_width=10, max_width=70, padding=2):
    for col_num, column_cells in enumerate(worksheet.columns, start=1):
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in column_cells
        )
        worksheet.column_dimensions[get_column_letter(col_num)].width = min(
            max(max_len + padding, min_width),
            max_width,
        )


def rows_to_xlsx_response(
    filename,
    sheet_title,
    headers,
    rows,
    quoted=False,
    bold_header=False,
    auto_width=False,
    fixed_column_width=None,
):
    workbook, worksheet = create_xlsx_workbook(sheet_title)
    if headers:
        worksheet.append(headers)

    for row in rows:
        worksheet.append(row)

    if headers and bold_header:
        style_header_row(worksheet)

    if fixed_column_width is not None:
        set_uniform_column_width(worksheet, fixed_column_width)

    if auto_width:
        autofit_columns(worksheet)

    return workbook_response(workbook, filename, quoted=quoted)


def dataframe_xlsx_response(data, filename, sheet_name, quoted=False, engine="openpyxl"):
    import pandas as pd

    dataframe = data if hasattr(data, "to_excel") else pd.DataFrame(data)
    response = xlsx_attachment_response(filename, quoted=quoted)
    with pd.ExcelWriter(response, engine=engine) as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
    return response

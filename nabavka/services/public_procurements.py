from __future__ import annotations

import hashlib
import json
import re
import tempfile
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.db.models import Max

from ..models import PublicProcurementPlanItem, PublicProcurementPlanVersion


@dataclass(frozen=True)
class ParsedPlanItem:
    plan_type: str
    stable_key: str
    content_hash: str
    source_sheet: str
    source_row: int
    item_number: str
    subject_type: str
    title: str
    estimated_value: Decimal | None
    procurement_category: str
    procedure_type: str
    quarter: str
    cpv: str
    nuts: str
    technique: str
    conducted_by_other: str
    exemption_basis: str
    valuation_method: str
    note: str
    raw_data: dict


HEADER_ALIASES = {
    "item_number": [
        "rbr",
        "jnbr",
        "jn broj",
        "redni broj",
    ],
    "subject_type": [
        "vrsta predmeta",
        "vrstapredmeta",
    ],
    "title": [
        "predmet javne nabavke",
        "opis predmeta nabavke",
        "opis nabavke",
    ],
    "estimated_value": [
        "procenjena vrednost",
        "procenjena vrednost bez pdv",
    ],
    "procedure_type": [
        "vrsta postupka",
    ],
    "procurement_category": [
        "predmet nabavke",
    ],
    "quarter": [
        "okvirno vreme pokretanja",
        "kvartal pokretanja nabavke",
    ],
    "cpv": [
        "cpv",
    ],
    "nuts": [
        "nstj izvrsenja isporuke",
        "nstj",
    ],
    "technique": [
        "tehnika",
    ],
    "conducted_by_other": [
        "sprovodi drugi narucilac",
    ],
    "exemption_basis": [
        "osnov izuzeca",
    ],
    "valuation_method": [
        "nacin procene vrednosti",
    ],
    "note": [
        "napomena",
    ],
}


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def _latinize(value):
    text = _clean(value).casefold()
    replacements = {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "ђ": "dj",
        "е": "e",
        "ж": "z",
        "з": "z",
        "и": "i",
        "ј": "j",
        "к": "k",
        "л": "l",
        "љ": "lj",
        "м": "m",
        "н": "n",
        "њ": "nj",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "ћ": "c",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "c",
        "ч": "c",
        "џ": "dz",
        "ш": "s",
        "đ": "dj",
    }
    text = "".join(replacements.get(ch, ch) for ch in text)
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _compact(value):
    return re.sub(r"[^a-z0-9]+", "", _latinize(value))


def _to_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = _clean(value)
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _canonical_header(value):
    normalized = _latinize(value)
    compact = _compact(value)
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if normalized == alias or compact == _compact(alias):
                return field
    return ""


def _detect_header_row(worksheet):
    best_row = None
    best_mapping = {}
    best_score = 0
    for row_idx in range(1, min(worksheet.max_row, 40) + 1):
        mapping = {}
        for col_idx in range(1, worksheet.max_column + 1):
            field = _canonical_header(worksheet.cell(row_idx, col_idx).value)
            if field and field not in mapping:
                mapping[field] = col_idx
        score = len(mapping)
        if "item_number" in mapping and "title" in mapping and score > best_score:
            best_row = row_idx
            best_mapping = mapping
            best_score = score
    if best_row is None:
        return None, {}
    return best_row, best_mapping


def _detect_plan_type(mapping):
    if "exemption_basis" in mapping or "valuation_method" in mapping:
        return PublicProcurementPlanItem.PlanType.EXEMPT
    return PublicProcurementPlanItem.PlanType.PUBLIC


def _stable_key(
    plan_type,
    item_number,
    title,
    source_sheet,
    section="",
    subject_type="",
    procurement_category="",
):
    item = _clean(item_number).lstrip("0") or _clean(item_number)
    sheet_key = _compact(source_sheet)[:30]
    section_key = _compact(section or subject_type)[:40]
    category_key = _compact(procurement_category)[:40]
    if item:
        if section_key and category_key:
            return f"{plan_type}:{sheet_key}:{section_key}:{category_key}:{item}"
        if section_key:
            return f"{plan_type}:{sheet_key}:{section_key}:{item}"
        return f"{plan_type}:{sheet_key}:{item}"
    title_hash = hashlib.sha256(_latinize(title).encode("utf-8")).hexdigest()[:24]
    return f"{plan_type}:{sheet_key}:title:{title_hash}"


def _content_hash(data):
    comparable = {key: _clean(value) for key, value in data.items() if key not in {"source_row"}}
    payload = json.dumps(comparable, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _is_total_or_summary_title(value):
    normalized = _latinize(value)
    return normalized.startswith(("ukupno", "svega"))


def _read_cell(worksheet, row_idx, mapping, field):
    col_idx = mapping.get(field)
    if not col_idx:
        return ""
    return worksheet.cell(row_idx, col_idx).value


def parse_public_procurement_excel(path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    parsed = []
    skipped_sheets = []
    duplicate_keys = set()
    seen_keys = set()

    for worksheet in workbook.worksheets:
        header_row, mapping = _detect_header_row(worksheet)
        if not header_row:
            skipped_sheets.append(worksheet.title)
            continue
        plan_type = _detect_plan_type(mapping)
        current_section = ""
        headers = {
            field: _clean(worksheet.cell(header_row, col_idx).value)
            for field, col_idx in mapping.items()
        }
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            item_number = _clean(_read_cell(worksheet, row_idx, mapping, "item_number"))
            title = _clean(_read_cell(worksheet, row_idx, mapping, "title"))
            if _is_total_or_summary_title(title):
                continue
            if not item_number:
                if title:
                    current_section = title
                continue
            raw_data = {
                headers[field] or field: _clean(_read_cell(worksheet, row_idx, mapping, field))
                for field in mapping
            }
            row_data = {
                "plan_type": plan_type,
                "source_sheet": worksheet.title,
                "source_row": row_idx,
                "item_number": item_number,
                "subject_type": _clean(_read_cell(worksheet, row_idx, mapping, "subject_type")),
                "title": title,
                "estimated_value": _to_decimal(_read_cell(worksheet, row_idx, mapping, "estimated_value")),
                "procurement_category": _clean(_read_cell(worksheet, row_idx, mapping, "procurement_category")),
                "procedure_type": _clean(_read_cell(worksheet, row_idx, mapping, "procedure_type")),
                "quarter": _clean(_read_cell(worksheet, row_idx, mapping, "quarter")),
                "cpv": _clean(_read_cell(worksheet, row_idx, mapping, "cpv")),
                "nuts": _clean(_read_cell(worksheet, row_idx, mapping, "nuts")),
                "technique": _clean(_read_cell(worksheet, row_idx, mapping, "technique")),
                "conducted_by_other": _clean(_read_cell(worksheet, row_idx, mapping, "conducted_by_other")),
                "exemption_basis": _clean(_read_cell(worksheet, row_idx, mapping, "exemption_basis")),
                "valuation_method": _clean(_read_cell(worksheet, row_idx, mapping, "valuation_method")),
                "note": _clean(_read_cell(worksheet, row_idx, mapping, "note")),
                "raw_data": raw_data,
            }
            raw_data["Sekcija"] = current_section
            stable_key = _stable_key(
                plan_type,
                item_number,
                title,
                worksheet.title,
                current_section,
                row_data["subject_type"],
                row_data["procurement_category"],
            )
            if stable_key in seen_keys:
                duplicate_keys.add(stable_key)
                stable_key = f"{stable_key}:row:{row_idx}"
            seen_keys.add(stable_key)
            parsed.append(
                ParsedPlanItem(
                    stable_key=stable_key,
                    content_hash=_content_hash(row_data),
                    **row_data,
                )
            )
    return parsed, skipped_sheets, sorted(duplicate_keys)


def _copy_previous_removed_item(version, previous_item):
    return PublicProcurementPlanItem(
        version=version,
        previous_item=previous_item,
        plan_type=previous_item.plan_type,
        diff_status=PublicProcurementPlanItem.DiffStatus.REMOVED,
        stable_key=previous_item.stable_key,
        content_hash=previous_item.content_hash,
        source_sheet=previous_item.source_sheet,
        source_row=previous_item.source_row,
        item_number=previous_item.item_number,
        subject_type=previous_item.subject_type,
        title=previous_item.title,
        estimated_value=previous_item.estimated_value,
        procurement_category=previous_item.procurement_category,
        procedure_type=previous_item.procedure_type,
        quarter=previous_item.quarter,
        cpv=previous_item.cpv,
        nuts=previous_item.nuts,
        technique=previous_item.technique,
        conducted_by_other=previous_item.conducted_by_other,
        exemption_basis=previous_item.exemption_basis,
        valuation_method=previous_item.valuation_method,
        note=previous_item.note,
        raw_data=previous_item.raw_data,
    )


def _build_item(version, parsed_item, previous_item):
    if previous_item is None:
        status = PublicProcurementPlanItem.DiffStatus.ADDED
    elif previous_item.content_hash == parsed_item.content_hash:
        status = PublicProcurementPlanItem.DiffStatus.UNCHANGED
    else:
        status = PublicProcurementPlanItem.DiffStatus.CHANGED
    return PublicProcurementPlanItem(
        version=version,
        previous_item=previous_item,
        plan_type=parsed_item.plan_type,
        diff_status=status,
        stable_key=parsed_item.stable_key,
        content_hash=parsed_item.content_hash,
        source_sheet=parsed_item.source_sheet,
        source_row=parsed_item.source_row,
        item_number=parsed_item.item_number,
        subject_type=parsed_item.subject_type,
        title=parsed_item.title,
        estimated_value=parsed_item.estimated_value,
        procurement_category=parsed_item.procurement_category,
        procedure_type=parsed_item.procedure_type,
        quarter=parsed_item.quarter,
        cpv=parsed_item.cpv,
        nuts=parsed_item.nuts,
        technique=parsed_item.technique,
        conducted_by_other=parsed_item.conducted_by_other,
        exemption_basis=parsed_item.exemption_basis,
        valuation_method=parsed_item.valuation_method,
        note=parsed_item.note,
        raw_data=parsed_item.raw_data,
    )


@transaction.atomic
def import_public_procurement_plan(*, excel_file, year, imported_by=None, note=""):
    suffix = Path(excel_file.name).suffix or ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        for chunk in excel_file.chunks():
            tmp.write(chunk)
        temp_path = Path(tmp.name)

    try:
        parsed_items, skipped_sheets, duplicate_keys = parse_public_procurement_excel(temp_path)
    finally:
        temp_path.unlink(missing_ok=True)

    if not parsed_items:
        raise ValueError("Excel ne sadrzi prepoznate stavke plana javnih nabavki.")

    latest_version = (
        PublicProcurementPlanVersion.objects.filter(year=year)
        .order_by("-version_number")
        .first()
    )
    previous_items = {}
    if latest_version:
        previous_items = {
            item.stable_key: item
            for item in latest_version.items.exclude(
                diff_status=PublicProcurementPlanItem.DiffStatus.REMOVED
            )
        }

    next_version_number = (
        PublicProcurementPlanVersion.objects.filter(year=year).aggregate(max_number=Max("version_number"))[
            "max_number"
        ]
        or 0
    ) + 1
    version = PublicProcurementPlanVersion.objects.create(
        year=year,
        version_number=next_version_number,
        source_filename=Path(excel_file.name).name,
        imported_by=imported_by,
        note=note,
    )

    new_keys = {item.stable_key for item in parsed_items}
    objects = [_build_item(version, item, previous_items.get(item.stable_key)) for item in parsed_items]
    for key, previous_item in previous_items.items():
        if key not in new_keys:
            objects.append(_copy_previous_removed_item(version, previous_item))

    # SQL Server/pyodbc can misinfer mixed bulk parameters for JSON and Decimal
    # columns in this import. Plan files are small enough that row-by-row saves
    # are fast and more reliable here.
    for item in objects:
        item.save()

    counts = {
        status: PublicProcurementPlanItem.objects.filter(version=version, diff_status=status).count()
        for status in PublicProcurementPlanItem.DiffStatus.values
    }
    version.total_rows = sum(counts.values())
    version.added_count = counts[PublicProcurementPlanItem.DiffStatus.ADDED]
    version.changed_count = counts[PublicProcurementPlanItem.DiffStatus.CHANGED]
    version.unchanged_count = counts[PublicProcurementPlanItem.DiffStatus.UNCHANGED]
    version.removed_count = counts[PublicProcurementPlanItem.DiffStatus.REMOVED]
    version.save(
        update_fields=[
            "total_rows",
            "added_count",
            "changed_count",
            "unchanged_count",
            "removed_count",
        ]
    )
    return {
        "version": version,
        "skipped_sheets": skipped_sheets,
        "duplicate_keys": duplicate_keys,
    }

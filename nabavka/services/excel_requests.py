from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.utils import timezone

from core.models import OrganizationalUnit
from fleet.models import TrafficCard

from ..models import ProcurementCase, ProcurementItem, ProcurementStatusLog


SOURCE_CODE = "PRACENJE_NABAVKE_ZA_GARAZU"
DEFAULT_REPORT_PATH = "izvestaji/nabavka_zahtevi_za_dopunu.csv"
GARAGE_WAREHOUSE_JOB_CODE = "811002"
REQUEST_DATE_CORRECTIONS = {
    (7, 4): date(2026, 3, 16),
    (7, 5): date(2026, 4, 17),
}
PLATE_CORRECTIONS = {
    "BG1470-VB": "BG1470-XB",
    "BG1549-PJ": "BG1542-PJ",
    "BG2434-GU": "BG2186-GU",
    "BG2504-GU": "BG2504-GL",
}


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _normalize_number(value):
    text = _clean_text(value)
    if text.endswith(".0"):
        return text[:-2]
    return text


def _normalize_plate(value):
    return _clean_text(value).upper()


def _source_marker(page_number, row_number):
    return f"[IMPORT:{SOURCE_CODE}:zahtevi:{page_number}:{row_number}]"


def _as_valid_date(value):
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date) and 2020 <= value.year <= 2030:
        return value
    return None


def _as_decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _title_for_items(items):
    names = []
    for item in items:
        name = item["name"]
        if name not in names:
            names.append(name)
    return "; ".join(names)[:255]


def _imported_created_at(request_date):
    return timezone.make_aware(datetime.combine(request_date, time(hour=12)))


def _report_row(request_row, item_rows, reasons, details):
    plates = sorted({_normalize_plate(item["plate"]) for item in item_rows if item["plate"]})
    return {
        "broj_strane": request_row["page_number"],
        "broj_reda": request_row["row_number"],
        "excel_red_zahteva": request_row["excel_row"],
        "sifra_posla": request_row["job_code"],
        "datum": request_row["request_date"],
        "tip": request_row["case_type"],
        "registracije": ", ".join(plates),
        "broj_stavki": len(item_rows),
        "razlozi": " | ".join(reasons),
        "detalji": " | ".join(details),
    }


def _write_report(report_path, rows):
    output_path = Path(report_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "broj_strane",
        "broj_reda",
        "excel_red_zahteva",
        "sifra_posla",
        "datum",
        "tip",
        "registracije",
        "broj_stavki",
        "razlozi",
        "detalji",
    ]
    with output_path.open("w", encoding="utf-8-sig", newline="") as report_file:
        writer = csv.DictWriter(report_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return output_path


def _load_rows(file_path, sheet_name):
    from openpyxl import load_workbook

    workbook_path = Path(file_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' ne postoji u fajlu {workbook_path}.")

    worksheet = workbook[sheet_name]
    requests = {}
    items = defaultdict(list)
    orphan_items = []

    for excel_row, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
        if any(value is not None for value in row[0:6]):
            key = (row[0], row[1])
            requests[key] = {
                "excel_row": excel_row,
                "page_number": row[0],
                "row_number": row[1],
                "oj": row[2],
                "job_code": _normalize_number(row[3]),
                "request_date": row[4],
                "case_type": _clean_text(row[5]).lower(),
            }

        if any(value is not None for value in row[13:21]):
            key = (row[13], row[14])
            item = {
                "excel_row": excel_row,
                "page_number": row[13],
                "row_number": row[14],
                "ordinal": row[15],
                "name": _clean_text(row[16]),
                "uom": _clean_text(row[17]),
                "quantity": _as_decimal(row[18]),
                "plate": _normalize_plate(row[19]),
                "note": _clean_text(row[20]),
            }
            items[key].append(item)

    for key in sorted(set(items) - set(requests)):
        orphan_items.extend(items[key])

    return requests, items, orphan_items


def _apply_known_corrections(key, request_row, item_rows):
    request_row = dict(request_row)
    item_rows = [dict(item) for item in item_rows]
    corrected_date = False
    corrected_plates = 0

    if key in REQUEST_DATE_CORRECTIONS:
        request_row["request_date"] = REQUEST_DATE_CORRECTIONS[key]
        corrected_date = True

    for item in item_rows:
        corrected_plate = PLATE_CORRECTIONS.get(item["plate"])
        if corrected_plate:
            item["plate"] = corrected_plate
            corrected_plates += 1

    plates = sorted({item["plate"] for item in item_rows if item["plate"]})
    goes_to_garage_warehouse = not plates or len(plates) > 1
    if goes_to_garage_warehouse:
        request_row["job_code"] = GARAGE_WAREHOUSE_JOB_CODE

    return request_row, item_rows, corrected_date, corrected_plates, goes_to_garage_warehouse


def _validate_request(request_row, item_rows, units_by_code, vehicles_by_plate):
    reasons = []
    details = []

    if request_row["case_type"] not in {
        ProcurementCase.CaseType.PROCUREMENT,
        ProcurementCase.CaseType.SERVICE,
    }:
        reasons.append("Nedostaje ili nije prepoznat tip zahteva")

    request_date = _as_valid_date(request_row["request_date"])
    if request_date is None:
        reasons.append("Datum zahteva nije validan")
        details.append(f"datum={request_row['request_date']}")

    if request_row["job_code"] not in units_by_code:
        reasons.append("Sifra posla ne postoji u aplikaciji")

    if not item_rows:
        reasons.append("Zahtev nema stavke")

    incomplete_items = [
        item
        for item in item_rows
        if not item["name"] or not item["uom"] or item["quantity"] is None
    ]
    if incomplete_items:
        reasons.append("Postoje nepotpune stavke")
        details.append(
            "nepotpune stavke: "
            + ", ".join(
                f"Excel red {item['excel_row']} / stavka {item['ordinal']}" for item in incomplete_items
            )
        )

    plates = sorted({item["plate"] for item in item_rows if item["plate"]})
    if not plates:
        details.append("magacin garaze: bez registracije vozila")
    elif len(plates) > 1:
        details.append("magacin garaze: vise vozila=" + ", ".join(plates))
    elif plates[0] not in vehicles_by_plate:
        reasons.append("Registracija vozila ne postoji u aplikaciji")
        details.append("registracija=" + plates[0])

    return reasons, details, request_date, plates


def _create_case(request_row, item_rows, request_date, job_code, vehicle):
    marker = _source_marker(request_row["page_number"], request_row["row_number"])
    created_at = _imported_created_at(request_date)
    procurement_case = ProcurementCase(
        case_type=request_row["case_type"],
        status=ProcurementCase.Status.DRAFT,
        title=_title_for_items(item_rows),
        description="Uvezeno iz Excel evidencije pracenja nabavke za garazu.",
        is_garage=True,
        job_code=job_code,
        vehicle=vehicle,
        needed_by=request_date + timedelta(days=7),
        note=f"{marker}\nDatum zahteva iz Excel evidencije: {request_date:%d.%m.%Y}.",
        created_at=created_at,
    )
    procurement_case.case_number = procurement_case.generate_case_number()
    procurement_case.save()
    ProcurementCase.objects.filter(pk=procurement_case.pk).update(created_at=created_at)

    ProcurementItem.objects.bulk_create(
        [
            ProcurementItem(
                procurement_case=procurement_case,
                name=item["name"],
                uom=item["uom"],
                quantity=item["quantity"],
                note=(
                    f"Excel: strana {request_row['page_number']}, red {request_row['row_number']}, "
                    f"stavka {item['ordinal']}."
                    + (f" {item['note']}" if item["note"] else "")
                )[:255],
            )
            for item in item_rows
        ]
    )
    ProcurementStatusLog.objects.create(
        procurement_case=procurement_case,
        old_status=None,
        new_status=procurement_case.status,
        comment=f"Uvezeno iz Excel evidencije. {marker}",
    )
    return procurement_case


def import_garage_requests_from_excel(
    file_path,
    *,
    sheet_name="zahtevi",
    report_path=DEFAULT_REPORT_PATH,
    commit=False,
):
    requests, items, orphan_items = _load_rows(file_path, sheet_name)
    units_by_code = {
        _normalize_number(unit.code): unit for unit in OrganizationalUnit.objects.all()
    }
    vehicles_by_plate = {
        _normalize_plate(card.registration_number): card.vehicle
        for card in TrafficCard.objects.select_related("vehicle")
    }
    existing_markers = {}
    for procurement_case in ProcurementCase.objects.filter(note__contains=f"[IMPORT:{SOURCE_CODE}:"):
        for line in (procurement_case.note or "").splitlines():
            if line.startswith(f"[IMPORT:{SOURCE_CODE}:"):
                existing_markers[line.strip()] = procurement_case

    result = {
        "requests": len(requests),
        "items": sum(len(group) for group in items.values()),
        "importable": 0,
        "created": 0,
        "created_items": 0,
        "updated_dates": 0,
        "corrected_input_dates": 0,
        "corrected_plates": 0,
        "garage_warehouse_requests": 0,
        "already_imported": 0,
        "skipped": 0,
        "orphan_items": len(orphan_items),
        "skip_reasons": Counter(),
        "commit": commit,
    }
    report_rows = []

    with transaction.atomic():
        for key, request_row in requests.items():
            item_rows = items.get(key, [])
            (
                request_row,
                item_rows,
                corrected_date,
                corrected_plates,
                goes_to_garage_warehouse,
            ) = _apply_known_corrections(key, request_row, item_rows)
            result["corrected_input_dates"] += int(corrected_date)
            result["corrected_plates"] += corrected_plates
            result["garage_warehouse_requests"] += int(goes_to_garage_warehouse)
            reasons, details, request_date, plates = _validate_request(
                request_row,
                item_rows,
                units_by_code,
                vehicles_by_plate,
            )
            if reasons:
                result["skipped"] += 1
                result["skip_reasons"].update(reasons)
                report_rows.append(_report_row(request_row, item_rows, reasons, details))
                continue

            result["importable"] += 1
            marker = _source_marker(request_row["page_number"], request_row["row_number"])
            if marker in existing_markers:
                result["already_imported"] += 1
                created_at = _imported_created_at(request_date)
                existing_case = existing_markers[marker]
                if existing_case.created_at != created_at:
                    if commit:
                        ProcurementCase.objects.filter(pk=existing_case.pk).update(created_at=created_at)
                    result["updated_dates"] += 1
                continue

            if commit:
                _create_case(
                    request_row,
                    item_rows,
                    request_date,
                    units_by_code[request_row["job_code"]],
                    vehicles_by_plate[plates[0]] if len(plates) == 1 else None,
                )
            result["created"] += 1
            result["created_items"] += len(item_rows)

        if not commit:
            transaction.set_rollback(True)

    result["report_path"] = str(_write_report(report_path, report_rows))
    result["skip_reasons"] = dict(result["skip_reasons"])
    return result

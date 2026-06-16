from datetime import datetime
from io import BytesIO
from urllib.parse import parse_qsl, urlencode

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, OuterRef, Q, Subquery, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.html import escape
from django.views import View
from django.views.generic import DetailView, ListView

from core.mixins import RolePermissionRequiredMixin
from fleet.models import JobCode, OrganizationalUnit, Vehicle

from ..forms import (
    EufInvoiceItemLinkForm,
    ProcurementInvoiceContractLinkForm,
    ProcurementInvoiceForm,
    ProcurementInvoiceJobCodeLinkForm,
)
from ..models import (
    ProcurementCase,
    ProcurementInvoice,
    ProcurementInvoiceContractLink,
    ProcurementInvoiceJobCodeLink,
    ProcurementItemInvoiceLink,
)
from ..services.euf import sync_euf_invoice_snapshots
from .cases import NabavkaContextMixin


JOB_CODE_LINK_FORM_PREFIX = "job_code_link"


def _parse_filter_date(value):
    parsed = parse_date(value or "")
    if parsed:
        return parsed
    try:
        return datetime.strptime(value or "", "%d.%m.%Y").date()
    except ValueError:
        return None


def _euf_invoice_base_queryset():
    return (
        ProcurementInvoice.objects.filter(source=ProcurementInvoice.SOURCE_EUF)
        .select_related("job_code", "vehicle")
        .prefetch_related("vehicle__traffic_cards", "job_code_links__job_code")
        .annotate(
            item_links_total=Count("item_links", distinct=True),
            garage_item_links_total=Count(
                "item_links",
                filter=Q(item_links__procurement_item__procurement_case__is_garage=True),
                distinct=True,
            ),
        )
    )


def _filter_euf_invoices(request, invoices=None):
    q = (request.GET.get("invoice_search") or request.GET.get("q") or "").strip()
    supplier = (request.GET.get("supplier") or "").strip()
    center = (request.GET.get("center") or "").strip()
    date_from = _parse_filter_date(request.GET.get("date_from"))
    date_to = _parse_filter_date(request.GET.get("date_to"))
    goes_to_warehouse = request.GET.get("goes_to_warehouse")
    is_garage = request.GET.get("is_garage")
    invoices = invoices if invoices is not None else _euf_invoice_base_queryset()
    if q:
        invoices = invoices.filter(
            Q(invoice_number__icontains=q)
            | Q(supplier_name__icontains=q)
            | Q(center_name__icontains=q)
            | Q(center__icontains=q)
        )
    if supplier:
        invoices = invoices.filter(supplier_name__icontains=supplier)
    if center:
        invoices = invoices.filter(Q(center__icontains=center) | Q(center_name__icontains=center))
    if date_from:
        invoices = invoices.filter(invoice_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(invoice_date__lte=date_to)
    if goes_to_warehouse in {"0", "1"}:
        invoices = invoices.filter(goes_to_warehouse=goes_to_warehouse == "1")
    if is_garage in {"0", "1"}:
        invoices = invoices.filter(is_garage=is_garage == "1")
    return invoices


def _euf_invoice_vehicle_label(invoice):
    return str(invoice.vehicle) if invoice.vehicle else ""


def _invoice_job_code_labels(invoice):
    labels = []
    if invoice.job_code:
        labels.append(getattr(invoice.job_code, "code", "") or "")
    labels.extend(
        link.job_code.code
        for link in invoice.job_code_links.all()
        if link.job_code and link.job_code.code
    )
    return list(dict.fromkeys(label for label in labels if label))


def _hover_text(value, max_length=50):
    text = str(value or "")
    visible = f"{text[:max_length]}..." if len(text) > max_length else text
    return f'<span title="{escape(text)}">{escape(visible)}</span>'


class EufInvoiceListView(NabavkaContextMixin, RolePermissionRequiredMixin, LoginRequiredMixin, ListView):
    template_name = "nabavka/euf_invoice_list.html"
    context_object_name = "invoices"

    def get_queryset(self):
        return ProcurementInvoice.objects.none()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        q = (self.request.GET.get("invoice_search") or self.request.GET.get("q") or "").strip()
        date_from = _parse_filter_date(self.request.GET.get("date_from"))
        date_to = _parse_filter_date(self.request.GET.get("date_to"))
        ctx.update(
            {
                "title": "Preuzete EUF",
                "q": q,
                "supplier": (self.request.GET.get("supplier") or "").strip(),
                "center": (self.request.GET.get("center") or "").strip(),
                "date_from": date_from.isoformat() if date_from else "",
                "date_to": date_to.isoformat() if date_to else "",
                "goes_to_warehouse": self.request.GET.get("goes_to_warehouse") or "",
                "is_garage": self.request.GET.get("is_garage") or "",
            }
        )
        return ctx


class EufInvoiceDataView(NabavkaContextMixin, RolePermissionRequiredMixin, LoginRequiredMixin, View):
    required_permission_code = "nabavka:euf_invoice_list"

    def get(self, request):
        invoices = _filter_euf_invoices(request)
        search_value = request.GET.get("search[value]", "").strip()
        if search_value:
            invoices = invoices.filter(
                Q(invoice_number__icontains=search_value)
                | Q(supplier_name__icontains=search_value)
                | Q(center_name__icontains=search_value)
                | Q(center__icontains=search_value)
                | Q(warehouse__icontains=search_value)
                | Q(registration__icontains=search_value)
                | Q(job_code__code__icontains=search_value)
                | Q(job_code_links__job_code__code__icontains=search_value)
                | Q(job_code_links__job_code__name__icontains=search_value)
                | Q(vehicle__brand__icontains=search_value)
                | Q(vehicle__model__icontains=search_value)
            ).distinct()

        records_total = ProcurementInvoice.objects.filter(source=ProcurementInvoice.SOURCE_EUF).count()
        records_filtered = invoices.count()
        order_map = {
            "0": "invoice_date",
            "1": "supplier_name",
            "2": "invoice_number",
            "3": "amount",
            "4": "job_code__code",
            "5": "goes_to_warehouse",
            "6": "is_garage",
            "7": "is_returned",
            "8": "vehicle__brand",
            "9": "item_links_total",
        }
        order_field = order_map.get(request.GET.get("order[0][column]", "0"), "invoice_date")
        if request.GET.get("order[0][dir]", "desc") == "desc":
            order_field = f"-{order_field}"
        invoices = invoices.order_by(order_field, "-id")

        try:
            start = max(int(request.GET.get("start", 0)), 0)
        except (TypeError, ValueError):
            start = 0
        try:
            length = int(request.GET.get("length", 50))
        except (TypeError, ValueError):
            length = 50
        if length < 0:
            length = 50
        length = min(length, 200)

        rows = []
        for invoice in invoices[start:start + length]:
            detail_url = reverse("nabavka:euf_invoice_detail", kwargs={"pk": invoice.pk})
            returned_url = reverse("nabavka:euf_invoice_returned_toggle", kwargs={"pk": invoice.pk})
            rows.append(
                {
                    "invoice_date": (
                        invoice.invoice_date.strftime("%d.%m.%Y")
                        if invoice.invoice_date
                        else escape(invoice.invoice_date_raw or "")
                    ),
                    "supplier_name": _hover_text(invoice.supplier_name),
                    "invoice_number": escape(invoice.invoice_number or ""),
                    "amount": str(invoice.amount) if invoice.amount is not None else "",
                    "job_code": escape(", ".join(_invoice_job_code_labels(invoice))),
                    "warehouse": (
                        '<span class="invoice-badge ok"><i class="mdi mdi-warehouse"></i> Da</span>'
                        if invoice.goes_to_warehouse
                        else '<span class="invoice-badge muted"><i class="mdi mdi-minus-circle-outline"></i> Ne</span>'
                    ),
                    "is_garage": (
                        '<span class="invoice-badge warn"><i class="mdi mdi-car-wrench"></i> Da</span>'
                        if invoice.is_garage or invoice.garage_item_links_total
                        else '<span class="invoice-badge muted">Ne</span>'
                    ),
                    "is_returned": (
                        '<div class="form-check invoice-returned-check">'
                        f'<input class="form-check-input js-invoice-returned" type="checkbox" '
                        f'data-url="{returned_url}" {"checked" if invoice.is_returned else ""} '
                        f'aria-label="Vraceno za fakturu {escape(invoice.invoice_number or "")}">'
                        "</div>"
                    ),
                    "vehicle": escape(_euf_invoice_vehicle_label(invoice)),
                    "item_links_total": invoice.item_links_total,
                    "actions": (
                        f'<a class="btn btn-outline-primary btn-sm" href="{detail_url}">'
                        '<i class="mdi mdi-eye"></i> Detalj</a>'
                    ),
                }
            )

        try:
            draw = int(request.GET.get("draw", 0))
        except (TypeError, ValueError):
            draw = 0
        return JsonResponse(
            {
                "draw": draw,
                "recordsTotal": records_total,
                "recordsFiltered": records_filtered,
                "data": rows,
            }
        )


class EufInvoiceExportView(NabavkaContextMixin, RolePermissionRequiredMixin, LoginRequiredMixin, View):
    required_permission_code = "nabavka:euf_invoice_list"

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        invoices = _filter_euf_invoices(request).order_by("-invoice_date", "-id")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Preuzete EUF"
        headers = [
            "Datum",
            "Partner",
            "Broj fakture",
            "Iznos",
            "Povezana OJ",
            "Magacin",
            "Garaza",
            "Vraceno",
            "Vozilo",
            "Stavke",
        ]
        worksheet.append(headers)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for invoice in invoices:
            worksheet.append(
                [
                    invoice.invoice_date.strftime("%d.%m.%Y") if invoice.invoice_date else invoice.invoice_date_raw or "",
                    invoice.supplier_name or "",
                    invoice.invoice_number or "",
                    invoice.amount,
                    ", ".join(_invoice_job_code_labels(invoice)),
                    "Da" if invoice.goes_to_warehouse else "Ne",
                    "Da" if invoice.is_garage or invoice.garage_item_links_total else "Ne",
                    "Da" if invoice.is_returned else "Ne",
                    _euf_invoice_vehicle_label(invoice),
                    invoice.item_links_total,
                ]
            )

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 45)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"preuzete-euf-{timezone.localtime().strftime('%Y%m%d-%H%M')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class EufInvoiceReturnedToggleView(RolePermissionRequiredMixin, LoginRequiredMixin, View):
    required_permission_code = "nabavka:euf_invoice_list"

    def post(self, request, pk):
        invoice = get_object_or_404(ProcurementInvoice, pk=pk, source=ProcurementInvoice.SOURCE_EUF)
        invoice.is_returned = (request.POST.get("is_returned") or "").lower() in {"1", "true", "on", "yes"}
        invoice.save(update_fields=["is_returned", "updated_at"])
        return JsonResponse({"ok": True, "is_returned": invoice.is_returned})


class EufInvoiceSyncView(NabavkaContextMixin, RolePermissionRequiredMixin, LoginRequiredMixin, View):
    required_permission_code = "nabavka:euf_invoice_list"

    def post(self, request):
        q = (request.POST.get("invoice_search") or request.POST.get("q") or "").strip()
        try:
            invoices = sync_euf_invoice_snapshots(q=q, limit=2000)
        except Exception as exc:
            messages.error(request, f"Fakture nisu preuzete iz EUF view-a: {exc}")
        else:
            messages.success(request, f"Preuzete EUF su osvezene. Obradjeno zapisa: {len(invoices)}.")
        redirect_url = reverse("nabavka:euf_invoice_list")
        query_string = urlencode(parse_qsl(request.POST.get("next_query") or "", keep_blank_values=False))
        return redirect(f"{redirect_url}?{query_string}" if query_string else redirect_url)


class EufInvoiceUpdateView(NabavkaContextMixin, RolePermissionRequiredMixin, LoginRequiredMixin, View):
    required_permission_code = "nabavka:euf_invoice_detail"

    def post(self, request, pk):
        invoice = get_object_or_404(ProcurementInvoice, pk=pk, source=ProcurementInvoice.SOURCE_EUF)
        form = ProcurementInvoiceForm(request.POST, instance=invoice)
        next_url = request.POST.get("next")
        if form.is_valid():
            form.save()
            messages.success(request, "Detalji fakture su sacuvani.")
        else:
            messages.error(request, "Detalji fakture nisu sacuvani. Proverite unete podatke.")
        return redirect(next_url or reverse("nabavka:euf_invoice_detail", kwargs={"pk": invoice.pk}))


class EufInvoiceDetailView(NabavkaContextMixin, RolePermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = ProcurementInvoice
    template_name = "nabavka/euf_invoice_detail.html"
    context_object_name = "invoice"

    def get_queryset(self):
        return ProcurementInvoice.objects.prefetch_related(
                "item_links__procurement_item__procurement_case",
                "item_links__created_by",
                "contract_links__contract__contract_type",
                "job_code_links__job_code",
        ).select_related("job_code", "vehicle")

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        action = request.POST.get("action")

        if action == "update_details":
            form = ProcurementInvoiceForm(request.POST, instance=self.object)
            if form.is_valid():
                form.save()
                messages.success(request, "Detalji fakture su sacuvani.")
            else:
                messages.error(request, "Detalji fakture nisu sacuvani. Proverite unete podatke.")
                return self.render_to_response(self.get_context_data(invoice_form=form))

        elif action == "link_item":
            form = EufInvoiceItemLinkForm(request.POST)
            if form.is_valid():
                procurement_item = form.cleaned_data["procurement_item"]
                procurement_case = procurement_item.procurement_case
                link, created = ProcurementItemInvoiceLink.objects.get_or_create(
                    procurement_item=procurement_item,
                    defaults={
                        "invoice": self.object,
                        "note": form.cleaned_data.get("note"),
                        "created_by": request.user,
                    },
                )
                if created and procurement_case.status == ProcurementCase.Status.WAITING_INVOICE:
                    procurement_case.status = ProcurementCase.Status.INVOICE_LINKED
                    procurement_case.save(update_fields=["status", "updated_at"])
                messages.success(
                    request,
                    "Stavka je povezana sa fakturom." if created else "Stavka je vec povezana sa fakturom.",
                )
            else:
                messages.error(request, "Stavka nije povezana. Proverite izbor stavke.")
                return self.render_to_response(self.get_context_data(item_link_form=form))

        elif action == "link_contract":
            form = ProcurementInvoiceContractLinkForm(request.POST, invoice=self.object)
            if form.is_valid():
                link = form.save(commit=False)
                link.invoice = self.object
                link.created_by = request.user
                link.save()
                messages.success(request, "Kupovni ugovor je povezan sa fakturom.")
            else:
                messages.error(request, "Ugovor nije povezan. Proverite izbor ugovora.")
                return self.render_to_response(self.get_context_data(contract_link_form=form))

        elif action == "link_job_code":
            form = ProcurementInvoiceJobCodeLinkForm(
                request.POST,
                invoice=self.object,
                prefix=JOB_CODE_LINK_FORM_PREFIX,
            )
            if form.is_valid():
                job_code = form.cleaned_data["job_code"]
                note = form.cleaned_data.get("note")
                if self.object.job_code_id == job_code.pk:
                    messages.warning(
                        request,
                        "Sifra posla je vec dodata kao osnovna sifra fakture.",
                    )
                    return redirect("nabavka:euf_invoice_detail", pk=self.object.pk)
                link, created = ProcurementInvoiceJobCodeLink.objects.get_or_create(
                    invoice=self.object,
                    job_code=job_code,
                    defaults={
                        "note": note,
                        "created_by": request.user,
                    },
                )
                if not created and note and link.note != note:
                    link.note = note
                    link.save(update_fields=["note"])
                messages.success(
                    request,
                    "Sifra posla je povezana sa fakturom."
                    if created
                    else "Sifra posla je vec povezana sa fakturom.",
                )
            else:
                messages.error(request, "Sifra posla nije povezana. Proverite izbor.")
                return self.render_to_response(self.get_context_data(job_code_link_form=form))

        return redirect("nabavka:euf_invoice_detail", pk=self.object.pk)

    @staticmethod
    def _contract_execution_rows(invoice):
        rows = []
        links = invoice.contract_links.select_related("contract", "contract__contract_type")
        for link in links:
            contract = link.contract
            execution_total = (
                ProcurementInvoiceContractLink.objects.filter(contract=contract)
                .aggregate(total=Sum("invoice__amount"))
                .get("total")
                or 0
            )
            has_fixed_value = contract.value_type == contract.VALUE_TYPE_FIXED and contract.value is not None
            remaining = contract.value - execution_total if has_fixed_value else None
            percent = (execution_total / contract.value * 100) if has_fixed_value and contract.value else None
            rows.append(
                {
                    "link": link,
                    "contract": contract,
                    "execution_total": execution_total,
                    "remaining": remaining,
                    "percent": percent,
                    "has_fixed_value": has_fixed_value,
                }
            )
        return rows

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        latest_job_code = (
            JobCode.objects.filter(vehicle_id=OuterRef("pk"))
            .order_by("-assigned_date", "-pk")
            .values("organizational_unit_id")[:1]
        )
        vehicle_job_codes = {
            str(vehicle_id): job_code_id
            for vehicle_id, job_code_id in (
                Vehicle.objects.annotate(current_job_code_id=Subquery(latest_job_code))
                .exclude(current_job_code_id__isnull=True)
                .values_list("pk", "current_job_code_id")
            )
        }
        job_code_labels = {
            job_code.pk: f"{job_code.code} - {job_code.name}"
            for job_code in OrganizationalUnit.objects.filter(pk__in=set(vehicle_job_codes.values()))
        }
        vehicle_job_code_options = {
            vehicle_id: {"id": job_code_id, "text": job_code_labels.get(job_code_id, str(job_code_id))}
            for vehicle_id, job_code_id in vehicle_job_codes.items()
        }
        ctx.update(
            {
                "title": f"Faktura {self.object.invoice_number}",
                "invoice_form": kwargs.get("invoice_form") or ProcurementInvoiceForm(instance=self.object),
                "item_link_form": kwargs.get("item_link_form") or EufInvoiceItemLinkForm(),
                "contract_link_form": kwargs.get("contract_link_form")
                or ProcurementInvoiceContractLinkForm(invoice=self.object),
                "job_code_link_form": kwargs.get("job_code_link_form")
                or ProcurementInvoiceJobCodeLinkForm(
                    invoice=self.object,
                    prefix=JOB_CODE_LINK_FORM_PREFIX,
                ),
                "job_code_links": self.object.job_code_links.select_related("job_code", "created_by"),
                "contract_execution_rows": self._contract_execution_rows(self.object),
                "vehicle_job_codes": vehicle_job_code_options,
                "item_links": self.object.item_links.select_related(
                    "procurement_item",
                    "procurement_item__procurement_case",
                    "created_by",
                ),
            }
        )
        return ctx


class ProcurementInvoiceLinkDeleteView(RolePermissionRequiredMixin, LoginRequiredMixin, View):
    def post(self, request, pk):
        link = get_object_or_404(ProcurementItemInvoiceLink, pk=pk)
        invoice_pk = link.invoice_id
        next_url = request.POST.get("next")
        link.delete()
        messages.success(request, "Veza stavke i fakture je obrisana.")
        return redirect(next_url or reverse("nabavka:euf_invoice_detail", kwargs={"pk": invoice_pk}))


class ProcurementInvoiceContractLinkDeleteView(RolePermissionRequiredMixin, LoginRequiredMixin, View):
    def post(self, request, pk):
        link = get_object_or_404(ProcurementInvoiceContractLink, pk=pk)
        invoice_pk = link.invoice_id
        link.delete()
        messages.success(request, "Veza fakture i ugovora je obrisana.")
        return redirect("nabavka:euf_invoice_detail", pk=invoice_pk)


class ProcurementInvoiceJobCodeLinkDeleteView(RolePermissionRequiredMixin, LoginRequiredMixin, View):
    required_permission_code = "nabavka:euf_invoice_detail"

    def post(self, request, pk):
        link = get_object_or_404(ProcurementInvoiceJobCodeLink, pk=pk)
        invoice_pk = link.invoice_id
        link.delete()
        messages.success(request, "Veza fakture i sifre posla je obrisana.")
        return redirect("nabavka:euf_invoice_detail", pk=invoice_pk)

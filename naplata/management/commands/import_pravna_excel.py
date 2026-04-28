from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from naplata.models import Postupak, PromenaPostupka


LIST_SHEETS = {
    "tuzeni_lista": "tuzeni",
    "tuzili_lista": "tuzili",
    "stecaj_lista": "stecaj",
    "uppr_lista": "uppr",
}

FAZA_SHEETS = {
    "tuzeni_faza": "tuzeni",
    "tuzili_faza": "tuzili",
    "stecaj_faza": "stecaj",
    "uppr_faza": "uppr",
}


def _norm(value):
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace("\xa0", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def _cell(ws, row_idx, col_idx):
    return ws.cell(row=row_idx, column=col_idx).value


def _to_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _to_bool(value):
    if value is None:
        return False
    s = _norm(value)
    return s in {"1", "da", "yes", "true", "x"}


def _to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        dec = value
    else:
        s = str(value).strip().replace(" ", "")
        # If comma is used as decimal separator, strip thousand dots.
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        try:
            dec = Decimal(s)
        except (InvalidOperation, ValueError):
            return None

    # Postupak monetary fields are Decimal(15,2): max 13 integer digits.
    if dec.is_nan() or dec.is_infinite():
        return None
    if dec.adjusted() > 12:
        return None
    try:
        return dec.quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _postupak_lookup(tip, data):
    if tip in {"tuzeni", "stecaj", "uppr"}:
        return {
            "tip": tip,
            "sifra_partnera": data.get("sifra_partnera"),
            "broj_predmeta": data.get("broj_predmeta"),
        }
    return {
        "tip": tip,
        "broj_predmeta": data.get("broj_predmeta"),
        "tuzilac": data.get("tuzilac"),
    }


class Command(BaseCommand):
    help = "Uvoz kompletnog Excel fajla pravne sluzbe (lista + faze) u server_db."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=str(Path("naplata") / "Pravna sluzba - ERP IMS tabele.xlsx"),
            help="Putanja do Excel fajla.",
        )
        parser.add_argument(
            "--database",
            default="server_db",
            help="DB alias iz settings.DATABASES (podrazumevano: server_db).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Samo simulacija bez upisa.",
        )
        parser.add_argument(
            "--truncate",
            action="store_true",
            help="Obrisi postojece podatke iz postupak/promena_postupka pre uvoza.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"])
        db_alias = options["database"]
        dry_run = options["dry_run"]
        truncate = options["truncate"]

        if not file_path.exists():
            raise CommandError(f"Excel fajl ne postoji: {file_path}")

        wb = load_workbook(file_path, data_only=True)

        created_postupci = 0
        updated_postupci = 0
        created_promene = 0
        skipped_rows = 0

        with transaction.atomic(using=db_alias):
            if truncate:
                PromenaPostupka.objects.using(db_alias).all().delete()
                Postupak.objects.using(db_alias).all().delete()

            for sheet_name, tip in LIST_SHEETS.items():
                if sheet_name not in wb.sheetnames:
                    self.stdout.write(self.style.WARNING(f"Preskacem: nedostaje sheet '{sheet_name}'"))
                    continue

                ws = wb[sheet_name]
                for row_idx in range(1, ws.max_row + 1):
                    if tip == "tuzeni":
                        data = {
                            "tip": tip,
                            "sud": _to_str(_cell(ws, row_idx, 2)),
                            "broj_predmeta": _to_str(_cell(ws, row_idx, 3)),
                            "izvrsiteljski_broj": _to_str(_cell(ws, row_idx, 4)),
                            "naziv_partnera": _to_str(_cell(ws, row_idx, 5)),
                            "sifra_partnera": _to_int(_cell(ws, row_idx, 6)),
                            "datum_pokretanja": _to_date(_cell(ws, row_idx, 9)),
                            "predmet_spora": _to_str(_cell(ws, row_idx, 10)),
                            "osnovni_dug": _to_decimal(_cell(ws, row_idx, 11)),
                            "arhivirano": _to_bool(_cell(ws, row_idx, 12)),
                            "valuta": "RSD",
                        }
                    elif tip == "tuzili":
                        data = {
                            "tip": tip,
                            "sud": _to_str(_cell(ws, row_idx, 2)),
                            "broj_predmeta": _to_str(_cell(ws, row_idx, 3)),
                            "tuzilac": _to_str(_cell(ws, row_idx, 4)),
                            "valuta": _to_str(_cell(ws, row_idx, 5)) or "RSD",
                            "vrednost_spora": _to_decimal(_cell(ws, row_idx, 6)),
                            "datum_podnosenja_tuzbe": _to_date(_cell(ws, row_idx, 7)),
                            "arhivirano": _to_bool(_cell(ws, row_idx, 8)),
                        }
                    elif tip == "stecaj":
                        data = {
                            "tip": tip,
                            "sud": _to_str(_cell(ws, row_idx, 2)),
                            "broj_predmeta": _to_str(_cell(ws, row_idx, 3)),
                            "vece": _to_str(_cell(ws, row_idx, 4)),
                            "naziv_partnera": _to_str(_cell(ws, row_idx, 5)),
                            "sifra_partnera": _to_int(_cell(ws, row_idx, 6)),
                            "valuta": _to_str(_cell(ws, row_idx, 9)) or "RSD",
                            "osnovni_dug": _to_decimal(_cell(ws, row_idx, 10)),
                            "kamata": _to_decimal(_cell(ws, row_idx, 11)),
                            "troskovi": _to_decimal(_cell(ws, row_idx, 12)),
                            "ukupan_dug": _to_decimal(_cell(ws, row_idx, 13)),
                            "datum_otvaranja_stecaja": _to_date(_cell(ws, row_idx, 14)),
                            "prijava_potrazivanja": _to_str(_cell(ws, row_idx, 15)),
                            "arhivirano": _to_bool(_cell(ws, row_idx, 16)),
                        }
                    else:  # uppr
                        data = {
                            "tip": tip,
                            "sud": _to_str(_cell(ws, row_idx, 2)) or _to_str(_cell(ws, row_idx, 1)),
                            "broj_predmeta": _to_str(_cell(ws, row_idx, 3)),
                            "novi_broj": _to_str(_cell(ws, row_idx, 4)),
                            "naziv_partnera": _to_str(_cell(ws, row_idx, 5)) or _to_str(_cell(ws, row_idx, 7)),
                            "sifra_partnera": _to_int(_cell(ws, row_idx, 6)),
                            "pib": _to_str(_cell(ws, row_idx, 9)),
                            "valuta": _to_str(_cell(ws, row_idx, 10)) or "RSD",
                            "vrednost_spora": _to_decimal(_cell(ws, row_idx, 11)),
                            "arhivirano": _to_bool(_cell(ws, row_idx, 12)),
                        }

                    # Za 'tuzili' ime tuzioca je u posebnom polju, a naziv_partnera neka ostane None
                    if tip == "tuzili":
                        data["tuzilac"] = data.get("tuzilac") or data.get("naziv_partnera")
                        data["naziv_partnera"] = None

                    lookup = _postupak_lookup(tip, data)
                    if not lookup.get("broj_predmeta"):
                        skipped_rows += 1
                        continue
                    if tip in {"tuzeni", "stecaj", "uppr"} and not lookup.get("sifra_partnera"):
                        skipped_rows += 1
                        continue
                    if tip == "tuzili" and not lookup.get("tuzilac"):
                        skipped_rows += 1
                        continue

                    obj = Postupak.objects.using(db_alias).filter(**lookup).first()
                    if obj is None:
                        obj = Postupak(**lookup)
                        created_postupci += 1
                    else:
                        updated_postupci += 1

                    for key, value in data.items():
                        if key not in lookup:
                            setattr(obj, key, value)

                    if not dry_run:
                        obj.save(using=db_alias)

            for sheet_name, tip in FAZA_SHEETS.items():
                if sheet_name not in wb.sheetnames:
                    self.stdout.write(self.style.WARNING(f"Preskacem: nedostaje sheet '{sheet_name}'"))
                    continue

                ws = wb[sheet_name]
                for row_idx in range(1, ws.max_row + 1):
                    if tip == "tuzili":
                        sifra_partnera = None
                        tuzilac = _to_str(_cell(ws, row_idx, 2))
                        broj_predmeta = _to_str(_cell(ws, row_idx, 3))
                        datum = _to_date(_cell(ws, row_idx, 4))
                        promena_text = _to_str(_cell(ws, row_idx, 5))
                    else:
                        sifra_partnera = _to_int(_cell(ws, row_idx, 2))
                        tuzilac = None
                        broj_predmeta = _to_str(_cell(ws, row_idx, 3))
                        datum = _to_date(_cell(ws, row_idx, 4))
                        promena_text = _to_str(_cell(ws, row_idx, 5))

                    if not broj_predmeta or not datum or not promena_text:
                        skipped_rows += 1
                        continue

                    if tip in {"tuzeni", "stecaj", "uppr"}:
                        if not sifra_partnera:
                            skipped_rows += 1
                            continue
                        postupak = Postupak.objects.using(db_alias).filter(
                            tip=tip,
                            sifra_partnera=sifra_partnera,
                            broj_predmeta=broj_predmeta,
                        ).first()
                    else:
                        if not tuzilac:
                            skipped_rows += 1
                            continue
                        postupak = Postupak.objects.using(db_alias).filter(
                            tip=tip,
                            broj_predmeta=broj_predmeta,
                            tuzilac=tuzilac,
                        ).first()

                    if not postupak:
                        skipped_rows += 1
                        continue

                    exists = PromenaPostupka.objects.using(db_alias).filter(
                        postupak=postupak,
                        datum=datum,
                        promena=promena_text,
                    ).exists()

                    if not exists:
                        created_promene += 1
                        if not dry_run:
                            PromenaPostupka.objects.using(db_alias).create(
                                postupak=postupak,
                                datum=datum,
                                promena=promena_text,
                            )

            if dry_run:
                transaction.set_rollback(True, using=db_alias)

        summary = (
            f"Uvoz zavrsen (dry_run={dry_run}) | "
            f"postupci: +{created_postupci} / ~{updated_postupci}, "
            f"promene: +{created_promene}, "
            f"preskoceno redova: {skipped_rows}."
        )
        self.stdout.write(self.style.SUCCESS(summary))

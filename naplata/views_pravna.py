from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import PermissionDenied
from django.db import connections
from django.db.models import Count
from django.views.decorators.cache import never_cache
from datetime import datetime
from decimal import Decimal
from openpyxl.styles import Alignment

from core.exporting import create_xlsx_workbook, set_column_widths, style_header_row, workbook_response
from fleet.mixins import role_permission_required
from .db_users import resolve_user_pk_for_db
from .models import Postupak, PromenaPostupka
from .forms_pravna import PostupakForm, PromenaPostupkaForm, COLUMNS_BY_TIP


PRAVNA_CASE_TYPES = {
    'tuzeni': 'Tuženi',
    'tuzili': 'Tužili',
    'stecaj': 'Stečaj',
    'uppr': 'UPPR',
}


def _get_pravna_title(case_type):
    return PRAVNA_CASE_TYPES.get(case_type)


def _build_pravna_filtered_context(request, case_type):
    show_archived = request.GET.get('arhivirano') == '1'
    base_qs = Postupak.objects.using('server_db').filter(tip=case_type)
    if not show_archived:
        base_qs = base_qs.filter(arhivirano=False)

    selected_sifra = (request.GET.get('sifra_partnera') or '').strip()
    selected_naziv = (request.GET.get('naziv_partnera') or '').strip()

    sifre = list(
        base_qs.exclude(sifra_partnera__isnull=True)
        .values_list('sifra_partnera', flat=True)
        .distinct()
    )

    partner_meta = {}
    if sifre:
        placeholders = ','.join(['%s'] * len(sifre))
        with connections['server_db'].cursor() as cursor:
            cursor.execute(
                f"SELECT sif_par, naz_par, grupa FROM partneri WHERE sif_par IN ({placeholders})",
                sifre,
            )
            for sif_par, naz_par, grupa in cursor.fetchall():
                partner_meta[int(sif_par)] = {
                    'naziv': naz_par,
                    'nivo': grupa,
                }

    qs = base_qs
    if selected_sifra:
        try:
            qs = qs.filter(sifra_partnera=int(selected_sifra))
        except ValueError:
            qs = qs.none()

    if selected_naziv:
        naziv_sifre = [
            sifra for sifra, meta in partner_meta.items()
            if (meta.get('naziv') or '') == selected_naziv
        ]
        qs = qs.filter(
            naziv_partnera=selected_naziv,
        ) | qs.filter(
            sifra_partnera__in=naziv_sifre,
        )

    partner_options = sorted(
        [
            {
                'sifra': sifra,
                'naziv': (partner_meta.get(int(sifra), {}).get('naziv') or ''),
            }
            for sifra in sifre
        ],
        key=lambda item: item['sifra'],
    )
    naziv_options = sorted(
        {
            (meta.get('naziv') or '').strip()
            for meta in partner_meta.values()
            if (meta.get('naziv') or '').strip()
        }
        | {
            (naziv or '').strip()
            for naziv in base_qs.exclude(naziv_partnera__isnull=True).values_list('naziv_partnera', flat=True)
            if (naziv or '').strip()
        }
    )

    return {
        'qs': qs,
        'show_archived': show_archived,
        'selected_sifra': selected_sifra,
        'selected_naziv': selected_naziv,
        'partner_options': partner_options,
        'naziv_options': naziv_options,
    }


# ─── Lista postupaka po tipu ──────────────────────────────────────

@never_cache
@role_permission_required()
def pravna_cases_list(request, case_type):
    title = _get_pravna_title(case_type)
    if not title:
        raise PermissionDenied('Nepoznat tip pravnog slučaja.')

    ctx = _build_pravna_filtered_context(request, case_type)
    qs = ctx['qs'].annotate(promene_count=Count('promene'))

    columns = COLUMNS_BY_TIP.get(case_type, [])

    return render(request, 'naplata/pravna_lista.html', {
        'title': f'Pravna služba - {title}',
        'case_type': case_type,
        'postupci': qs,
        'columns': columns,
        'show_archived': ctx['show_archived'],
        'selected_sifra': ctx['selected_sifra'],
        'selected_naziv': ctx['selected_naziv'],
        'partner_options': ctx['partner_options'],
        'naziv_options': ctx['naziv_options'],
    })


@never_cache
@role_permission_required()
def pravna_izvestaj(request, case_type):
    title = _get_pravna_title(case_type)
    if not title:
        raise PermissionDenied('Nepoznat tip pravnog slučaja.')

    ctx = _build_pravna_filtered_context(request, case_type)
    postupci = list(ctx['qs'].order_by('naziv_partnera', 'sifra_partnera', 'broj_predmeta', 'id'))
    postupak_ids = [p.id for p in postupci]

    promene_map = {pid: [] for pid in postupak_ids}
    if postupak_ids:
        promene = PromenaPostupka.objects.using('server_db').filter(
            postupak_id__in=postupak_ids
        ).order_by('datum', 'created_at')
        for pr in promene:
            promene_map.setdefault(pr.postupak_id, []).append(pr)

    groups = {}
    for p in postupci:
        sifra = p.sifra_partnera
        naziv = (p.naziv_partnera or '').strip()
        if not naziv and getattr(p, 'tuzilac', None):
            naziv = p.tuzilac
        if not naziv:
            naziv = 'Nepoznat partner'

        key = (sifra, naziv)
        if key not in groups:
            groups[key] = {
                'sifra_partnera': sifra,
                'naziv_partnera': naziv,
                'postupci': [],
            }

        groups[key]['postupci'].append({
            'postupak': p,
            'promene': promene_map.get(p.id, []),
        })

    filter_badges = []
    if ctx['selected_sifra']:
        filter_badges.append(f"Sifra partnera: {ctx['selected_sifra']}")
    if ctx['selected_naziv']:
        filter_badges.append(f"Naziv partnera: {ctx['selected_naziv']}")
    filter_badges.append('Arhivirano: ukljuceno' if ctx['show_archived'] else 'Arhivirano: iskljuceno')

    return render(request, 'naplata/pravna_izvestaj.html', {
        'title': f'Izvestaj - Pravna sluzba - {title}',
        'case_type': case_type,
        'report_date': datetime.now(),
        'groups': list(groups.values()),
        'columns': COLUMNS_BY_TIP.get(case_type, []),
        'filter_badges': filter_badges,
        'total_postupci': len(postupci),
    })


@never_cache
@role_permission_required()
def pravna_izvestaj_excel(request, case_type):
    title = _get_pravna_title(case_type)
    if not title:
        raise PermissionDenied('Nepoznat tip pravnog slucaja.')

    ctx = _build_pravna_filtered_context(request, case_type)
    postupci = list(ctx['qs'].order_by('sud', 'broj_predmeta', 'naziv_partnera', 'id'))
    postupak_ids = [p.id for p in postupci]

    promene_map = {pid: [] for pid in postupak_ids}
    if postupak_ids:
        promene = PromenaPostupka.objects.using('server_db').filter(
            postupak_id__in=postupak_ids
        ).order_by('datum', 'created_at')
        for pr in promene:
            promene_map.setdefault(pr.postupak_id, []).append(pr)

    def _format_date(value):
        return value.strftime('%d.%m.%Y') if value else '-'

    def _duznik_tuzeni(postupak):
        return (
            (postupak.naziv_partnera or '').strip()
            or (postupak.tuzilac or '').strip()
            or (str(postupak.sifra_partnera) if postupak.sifra_partnera else '-')
        )

    def _datum_pokretanja(postupak):
        return (
            postupak.datum_pokretanja
            or postupak.datum_podnosenja_tuzbe
            or postupak.datum_otvaranja_stecaja
        )

    def _predmet_spora(postupak):
        return (
            (postupak.predmet_spora or '').strip()
            or (postupak.prijava_potrazivanja or '').strip()
            or '-'
        )

    def _vrednost_spora(postupak):
        if postupak.vrednost_spora is not None:
            return postupak.vrednost_spora
        if postupak.osnovni_dug is not None:
            return postupak.osnovni_dug
        if postupak.ukupan_dug is not None:
            return postupak.ukupan_dug
        return '-'

    def _faze_postupka(postupak_id):
        promene = promene_map.get(postupak_id, [])
        if not promene:
            return 'Nema unetih faza.'
        lines = []
        for idx, pr in enumerate(promene, start=1):
            datum = _format_date(pr.datum)
            promena = (pr.promena or '').strip()
            if promena:
                lines.append(f'{idx}. {datum} - {promena}')
            else:
                lines.append(f'{idx}. {datum}')
        return '\n'.join(lines)

    wb, ws = create_xlsx_workbook('Pregled sudskih postupaka')

    headers = [
        'Sud i broj postupka',
        'Duznik - tuzeni',
        'Datum pokretanja postupka',
        'Predmet spora',
        'Vrednost spora',
        'Faze postupka',
    ]
    ws.append(headers)
    style_header_row(ws)

    widths = [38, 34, 24, 48, 18, 70]
    set_column_widths(ws, widths)

    for row_num, postupak in enumerate(postupci, 2):
        sud = (postupak.sud or '').strip()
        broj = (postupak.broj_predmeta or '').strip()
        sud_i_broj = f'{sud} / {broj}' if sud and broj else (sud or broj or '-')

        ws.cell(row=row_num, column=1, value=sud_i_broj)
        ws.cell(row=row_num, column=2, value=_duznik_tuzeni(postupak))
        ws.cell(row=row_num, column=3, value=_format_date(_datum_pokretanja(postupak)))
        ws.cell(row=row_num, column=4, value=_predmet_spora(postupak))
        ws.cell(row=row_num, column=5, value=_vrednost_spora(postupak))
        ws.cell(row=row_num, column=6, value=_faze_postupka(postupak.id))

        ws.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row_num, column=6).alignment = Alignment(wrap_text=True, vertical='top')

    filename = f'pravna_{case_type}_pregled_postupaka_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response = workbook_response(wb, filename)
    response.set_cookie('excel_download', '1', max_age=120, path='/', samesite='Lax')
    return response


# ─── Detalj + faze ───────────────────────────────────────────────

@never_cache
@role_permission_required()
def pravna_detalj(request, pk):
    postupak = get_object_or_404(Postupak.objects.using('server_db'), pk=pk)
    promene = PromenaPostupka.objects.using('server_db').filter(postupak=postupak)
    columns = COLUMNS_BY_TIP.get(postupak.tip, [])

    # Partner iz baze
    partner = None
    if postupak.sifra_partnera:
        with connections['server_db'].cursor() as cursor:
            cursor.execute(
                "SELECT naz_par, mesto_par, pib FROM partneri WHERE sif_par = %s",
                [postupak.sifra_partnera],
            )
            row = cursor.fetchone()
            if row:
                partner = {
                    'naziv_u_bazi': row[0],
                    'mesto': row[1],
                    'pib': row[2],
                }

    forma_promena = PromenaPostupkaForm()

    return render(request, 'naplata/pravna_detalj.html', {
        'title': f'{postupak.get_tip_display()} - {postupak.naziv_partnera or postupak.sifra_partnera}',
        'postupak': postupak,
        'promene': promene,
        'partner': partner,
        'columns': columns,
        'forma_promena': forma_promena,
    })


# ─── Dodaj postupak ──────────────────────────────────────────────

@role_permission_required()
def pravna_dodaj(request, case_type):
    title = _get_pravna_title(case_type)
    if not title:
        raise PermissionDenied('Nepoznat tip pravnog slučaja.')

    if request.method == 'POST':
        form = PostupakForm(request.POST, tip=case_type)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tip = case_type
            obj.created_by_id = resolve_user_pk_for_db(request.user, 'server_db')
            obj.save(using='server_db')
            return redirect('naplata:pravna_cases_list', case_type=case_type)
    else:
        form = PostupakForm(tip=case_type)

    return render(request, 'naplata/pravna_forma.html', {
        'title': f'Novi postupak - {title}',
        'form': form,
        'case_type': case_type,
    })


# ─── Izmeni postupak ─────────────────────────────────────────────

@role_permission_required()
def pravna_izmeni(request, pk):
    postupak = get_object_or_404(Postupak.objects.using('server_db'), pk=pk)

    if request.method == 'POST':
        form = PostupakForm(request.POST, instance=postupak, tip=postupak.tip)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.save(using='server_db')
            return redirect('naplata:pravna_detalj', pk=pk)
    else:
        form = PostupakForm(instance=postupak, tip=postupak.tip)

    return render(request, 'naplata/pravna_forma.html', {
        'title': f'Izmeni - {postupak}',
        'form': form,
        'case_type': postupak.tip,
        'postupak': postupak,
    })


# ─── Obriši postupak ─────────────────────────────────────────────

@role_permission_required()
def pravna_obrisi(request, pk):
    postupak = get_object_or_404(Postupak.objects.using('server_db'), pk=pk)
    case_type = postupak.tip
    if request.method == 'POST':
        postupak.delete(using='server_db')
        return redirect('naplata:pravna_cases_list', case_type=case_type)
    return redirect('naplata:pravna_detalj', pk=pk)


@role_permission_required()
def pravna_arhiviraj(request, pk):
    postupak = get_object_or_404(Postupak.objects.using('server_db'), pk=pk)
    if request.method == 'POST':
        postupak.arhivirano = not postupak.arhivirano
        postupak.save(using='server_db', update_fields=['arhivirano'])
    return redirect('naplata:pravna_detalj', pk=pk)


# ─── Dodaj promenu (fazu) ────────────────────────────────────────

@role_permission_required()
def pravna_dodaj_promenu(request, pk):
    postupak = get_object_or_404(Postupak.objects.using('server_db'), pk=pk)

    if request.method == 'POST':
        form = PromenaPostupkaForm(request.POST)
        if form.is_valid():
            promena = form.save(commit=False)
            promena.postupak = postupak
            promena.created_by_id = resolve_user_pk_for_db(request.user, 'server_db')
            promena.save(using='server_db')
    return redirect('naplata:pravna_detalj', pk=pk)


# ─── Obriši promenu ──────────────────────────────────────────────

@role_permission_required()
def pravna_obrisi_promenu(request, pk):
    promena = get_object_or_404(PromenaPostupka.objects.using('server_db'), pk=pk)
    postupak_pk = promena.postupak_id
    if request.method == 'POST':
        promena.delete(using='server_db')
    return redirect('naplata:pravna_detalj', pk=postupak_pk)


# ─── Stari view za listu iz SQL view-a (tuzeni) ──────────────────

@never_cache
@role_permission_required()
def lista_tuzenih(request):
    with connections['server_db'].cursor() as cursor:
        cursor.execute("""
            SELECT
                god,
                sif_par,
                naz_par,
                datum,
                knt,
                naz_knt,
                duguju,
                platili
            FROM dbo.v_tuzeni
            ORDER BY god DESC, duguju DESC
        """)
        tuzeni = cursor.fetchall()

    total_duguju = sum((Decimal(row[6] or 0) for row in tuzeni), Decimal("0"))
    total_platili = sum((Decimal(row[7] or 0) for row in tuzeni), Decimal("0"))

    return render(request, 'naplata/tuzeni_list.html', {
        'tuzeni': tuzeni,
        'total_duguju': total_duguju,
        'total_platili': total_platili,
        'title': 'Utuženi Klijenti',
    })

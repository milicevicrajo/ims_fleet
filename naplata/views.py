import logging
from datetime import date, datetime, time
from decimal import Decimal
from urllib.parse import urlencode

from openpyxl import load_workbook

from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.db import connections
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST

from core.exporting import rows_to_xlsx_response
from core.mixins import role_permission_required
from core.models import OrganizationalUnit

from .db_users import resolve_user_pk_for_db
from .forms import (
    KontaktiForm,
    NapomeneForm,
    OpomeneForm,
    PozivPismoForm,
    PoziviTelForm,
    TuzbeForm,
)
from .models import (
    AvansKlijent,
    Kontakti,
    Napomene,
    Opomene,
    PozivPismo,
    PoziviTel,
    Tuzbe,
)
from .queries import (
    dugovanja_po_bucketima_rows,
    izvestaj_po_siframa_posla_data,
    neodobrene_if_filters_from_request,
    neodobrene_if_options,
    neodobrene_if_rows,
)

logger = logging.getLogger(__name__)


def _allowed_centers_from_user(user):
    codes = []
    raw = (user.allowed_center_codes or "").strip()
    if raw:
        parts = [p.strip() for p in raw.replace(";", ",").split(",")]
        codes.extend([p for p in parts if p])

    unit_centers = user.allowed_centers.values_list("center", flat=True)
    codes.extend([str(center).strip() for center in unit_centers if str(center).strip()])
    return sorted({str(code).strip() for code in codes if str(code).strip()})


def _allowed_sif_pos_from_user(user):
    if user.is_superuser:
        return []

    allowed_centers = _allowed_centers_from_user(user)
    explicit_codes = user.allowed_centers.values_list("code", flat=True)
    codes = {str(code).strip() for code in explicit_codes if str(code).strip()}

    if allowed_centers:
        center_codes = OrganizationalUnit.objects.filter(
            center__in=allowed_centers
        ).values_list("code", flat=True)
        codes.update({str(code).strip() for code in center_codes if str(code).strip()})

    return sorted(codes)

@role_permission_required()
def lista_dugovanja(request):
    with connections['server_db'].cursor() as cursor:
        cursor.execute("SELECT sif_par, naz_par, dug, pot FROM dbo.baza ORDER BY dug DESC")
        dugovanja = cursor.fetchall()
    total_dug = sum((Decimal(row[2] or 0) for row in dugovanja), Decimal("0"))
    total_pot = sum((Decimal(row[3] or 0) for row in dugovanja), Decimal("0"))

    return render(request, 'naplata/dugovanja.html', {
        'dugovanja': dugovanja,
        'total_dug': total_dug,
        'total_pot': total_pot,
    })


@never_cache
@role_permission_required()
def lista_dugovanja_po_bucketima(request):
    marked_ids = set(AvansKlijent.objects.values_list('sif_par', flat=True))
    dugovanja = dugovanja_po_bucketima_rows()

    return render(request, 'naplata/dugovanja_bucketi.html', {
        'dugovanja': dugovanja,
        'marked_ids': marked_ids,
        'title': 'Lista dugovanja po bucketima',
    })


@never_cache
def lista_avans_klijenti(request):
    marked_ids = list(AvansKlijent.objects.values_list('sif_par', flat=True))
    dugovanja = dugovanja_po_bucketima_rows(marked_ids)

    return render(request, 'naplata/dugovanja_bucketi_avans.html', {
        'dugovanja': dugovanja,
        'marked_ids': set(marked_ids),
        'title': 'Spisak za proveru',
    })


@never_cache
@role_permission_required()
def izvestaj_po_siframa_posla(request):
    allowed_sif_pos = _allowed_sif_pos_from_user(request.user)
    selected_sif_pos = (request.GET.get("sif_pos") or "").strip()
    marked_ids = set(AvansKlijent.objects.values_list('sif_par', flat=True))

    if selected_sif_pos and not request.user.is_superuser and selected_sif_pos not in allowed_sif_pos:
        raise PermissionDenied('Nemate pristup izabranoj šifri posla.')

    dugovanja, sif_pos_options = izvestaj_po_siframa_posla_data(
        request.user.is_superuser,
        allowed_sif_pos,
        selected_sif_pos,
    )
    return render(request, 'naplata/izvestaj_sifra_posla.html', {
        'dugovanja': dugovanja,
        'marked_ids': marked_ids,
        'title': 'Izveštaj po šiframa posla',
        'report_mode': True,
        'report_allowed_sif_pos': allowed_sif_pos,
        'sif_pos_options': sif_pos_options,
        'selected_sif_pos': selected_sif_pos,
    })


@never_cache
@role_permission_required()
def neodobrene_if_izvestaj(request):
    filters = neodobrene_if_filters_from_request(request)
    rows = neodobrene_if_rows(filters)
    god_options, status_options, partner_options, naziv_options = neodobrene_if_options()
    query_string = urlencode({k: v for k, v in filters.items() if v})

    return render(request, 'naplata/neodobrene_if.html', {
        'title': 'Neodobrene IF',
        'rows': rows,
        'filters': filters,
        'god_options': god_options,
        'status_options': status_options,
        'partner_options': partner_options,
        'naziv_options': naziv_options,
        'rows_count': len(rows),
        'query_string': query_string,
    })


@role_permission_required()
def export_neodobrene_if_excel(request):
    filters = neodobrene_if_filters_from_request(request)
    rows = neodobrene_if_rows(filters)

    headers = [
        "OJ",
        "God",
        "Datum",
        "Sifra partnera",
        "Naziv partnera",
        "Faktura",
        "Status na sefu",
        "Vreme statusa",
        "Komentar",
    ]

    return rows_to_xlsx_response(
        "neodobrene_if.xlsx",
        "Neodobrene IF",
        headers,
        rows,
        bold_header=True,
        auto_width=True,
    )


@require_POST
@role_permission_required()
def toggle_avans_klijent(request):
    sif_par = request.POST.get('sif_par')
    if not sif_par:
        return redirect(request.META.get('HTTP_REFERER', 'naplata:lista_dugovanja_po_bucketima'))

    try:
        sif_par_int = int(sif_par)
    except (TypeError, ValueError):
        return redirect(request.META.get('HTTP_REFERER', 'naplata:lista_dugovanja_po_bucketima'))

    obj, created = AvansKlijent.objects.get_or_create(
        sif_par=sif_par_int,
        defaults={'created_by_id': resolve_user_pk_for_db(request.user, AvansKlijent.objects.db)},
    )
    if not created:
        obj.delete()

    return redirect(request.META.get('HTTP_REFERER', 'naplata:lista_dugovanja_po_bucketima'))


@role_permission_required()
def export_dugovanja_bucketi_excel(request):
    # Header
    headers = [
        "Šifra partnera", "Naziv partnera", "Nedospelo", "Dospelo - baket 1", "Dospelo - baket 2",
        "Dospelo - baket 3", "Dospelo - baket 4", "Dospelo - baket 5", "Dospelo - baket 6", "Ukupno - Dospelo",
        "Ukupno", "Veliki", "INO"
    ]
    rows = dugovanja_po_bucketima_rows()

    return rows_to_xlsx_response("dugovanja_po_baketima.xlsx", "Dugovanja po baketima", headers, rows)


@never_cache
@role_permission_required()
def detalji_partner(request, sif_par):
    report_mode = request.GET.get('report') == '1'
    report_allowed_sif_pos = _allowed_sif_pos_from_user(request.user) if report_mode else []

    if report_mode and not request.user.is_superuser and not report_allowed_sif_pos:
        raise PermissionDenied('Nije definisana šifra posla za korisnika.')

    if report_mode and not request.user.is_superuser:
        with connections['server_db'].cursor() as cursor:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            cursor.execute(f"""
                SELECT TOP 1 1
                FROM dodela_baketa
                WHERE sif_par = %s AND sif_pos IN ({placeholders})
            """, [sif_par, *report_allowed_sif_pos])
            if cursor.fetchone() is None:
                raise PermissionDenied('Nemate pristup ovom partneru za vašu šifru posla.')

    # 1. Osnovne informacije o partneru (view partneri)
    with connections['server_db'].cursor() as cursor:
        cursor.execute("""
            SELECT sif_pred, grupa, sif_par, naz_par, ulica_par, p_b_par, mesto_par, 
                   zr, mb, telefon, fax, email, lice, web, proc_rabata, br_dana, vlasnik, pib, zemlja, 
                   proc_rabata1, proc_rabata2, pdv_obveznik, sif_ter, JBKJS, CRF
            FROM partneri
            WHERE sif_par = %s
        """, [sif_par])
        partner = cursor.fetchone()


    # 2.1 Dugovanja (view baza)
    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT 
                god AS Godina,
                oj AS OJ,
                sif_pos AS [šifra posla],
                sif_vrs AS vrsta,
                datum,
                vez_dok AS veza,
                dpo,
                skr_naz AS valuta,
                dug AS duguje,
                pot AS potražuje
            FROM baza
            WHERE sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += """
            ORDER BY datum
        """
        cursor.execute(sql, params)
        dugovanja = cursor.fetchall()


    # 2.2 Dugovanja baketi (view baza)
    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT 
                sb.opis,
                b.baket,
                b.sif_pos,
                b.vez_dok,
                b.dpo,
                b.duguje,
                b.potrazuje,
                b.saldo,
                sb.akcija
            FROM dodela_baketa b
            LEFT JOIN sif_baket sb ON b.baket = sb.baket
            WHERE b.sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND b.sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += """
            ORDER BY b.baket
        """
        cursor.execute(sql, params)
        dugovanja_baket = cursor.fetchall()

        
    # 2.3 Dugovanja po fakturama
    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT 
                br_naloga, 
                sif_par, 
                naz_par, 
                sif_vrs, 
                MAX(dat_naloga) AS poslednji_datum,  
                SUM(dug) AS ukupno_duguje, 
                SUM(pot) AS ukupno_potrazuje, 
                SUM(dug) - SUM(pot) AS saldo
            FROM baza
            WHERE sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += """
            GROUP BY br_naloga, sif_par, naz_par, sif_vrs
            ORDER BY poslednji_datum DESC
        """
        cursor.execute(sql, params)
        dugovanja_sumarno = cursor.fetchall()


    # 3. Dospela potraživanja po bucketima (view dodela_bucketa)
    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT sif_par, naz_par, vez_dok, sif_pos, duguje, potražuje, saldo,
                   dpo, danasnji_datum, broj_dana, baket, kategorija
            FROM dodela_baketa
            WHERE sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += """
            ORDER BY baket
        """
        cursor.execute(sql, params)
        baketi = cursor.fetchall()

    # 4. Otpisana potraživanja (view ispravke)
    with connections['server_db'].cursor() as cursor:
        cursor.execute("""
            SELECT 
                god,
                oj,
                sif_pos AS [šifra posla],
                sif_vrs,
                datum,
                dpo,
                knt,
                naz_knt,
                dug,
                pot
            FROM ispravke
            WHERE sif_par = %s
            ORDER BY datum DESC
        """, [sif_par])
        ispravke = cursor.fetchall()




    # # 5. Ostale tabele (kontakti, pozivi, opomene, tužbe, napomene)
    kontakti = Kontakti.objects.using('server_db').filter(sif_par=sif_par)
    napomene = Napomene.objects.using('server_db').filter(sif_par=sif_par)
    opomene = Opomene.objects.using('server_db').filter(sif_par=sif_par)
    poziv_pismo = PozivPismo.objects.using('server_db').filter(sif_par=sif_par)
    pozivi_tel = PoziviTel.objects.using('server_db').filter(sif_par=sif_par)
    tuzbe = Tuzbe.objects.using('server_db').filter(sif_par=sif_par)
    

    # 6. Spisak faktura iz baketa 6
    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT sif_par, naz_par as Naziv, TRIM(vez_dok) AS veza, sif_pos as [šifra posla], 
                    SUM(duguje) AS duguje, SUM(potrazuje) AS potrazuje, SUM(saldo) AS saldo
            FROM dodela_baketa
            WHERE baket = 181 AND sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += """
            GROUP BY sif_par, naz_par, vez_dok, sif_pos
            ORDER BY sif_par
        """
        cursor.execute(sql, params)
        spisak_utuzenje = cursor.fetchall()

    # 7. Spisak faktura iz baketa 5
    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT sif_par, naz_par, TRIM(vez_dok) AS veza, sif_pos, 
                SUM(duguje) AS duguje, SUM(potrazuje) AS potrazuje, 
                SUM(saldo) AS saldo
            FROM dodela_baketa
            WHERE baket = 180 AND sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += """
            GROUP BY sif_par, naz_par, vez_dok, sif_pos
            ORDER BY sif_par
        """
        cursor.execute(sql, params)
        opomene_fakture = cursor.fetchall()

    # 8. Spisak faktura iz baketa 4
    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT sif_par, naz_par, LTRIM(RTRIM(vez_dok)) AS veza, sif_pos,
                SUM(duguje) AS duguje, SUM(potrazuje) AS potrazuje, SUM(saldo) AS saldo
            FROM dodela_baketa
            WHERE baket = 90 AND sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += """
            GROUP BY sif_par, naz_par, vez_dok, sif_pos
            ORDER BY sif_par
        """
        cursor.execute(sql, params)
        fakture_baket_90 = cursor.fetchall()

    # 9. Spisak faktura iz baketa 3
    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT sif_par, naz_par as Naziv, 
                   LTRIM(RTRIM(vez_dok)) AS veza, 
                   sif_pos AS [šifra posla], 
                   SUM(duguje) AS duguje, 
                   SUM(potrazuje) AS potrazuje, 
                   SUM(saldo) AS saldo
            FROM dodela_baketa
            WHERE baket = 60 AND sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += """
            GROUP BY sif_par, naz_par, vez_dok, sif_pos
            ORDER BY sif_par
        """
        cursor.execute(sql, params)
        fakture_baket_60 = cursor.fetchall()

    # Slanje svih podataka u template
    return render(request, 'naplata/detalji_partner.html', {
        'partner': partner,
        'dugovanja': dugovanja,
        'baketi': baketi,
        'ispravke': ispravke,
        'kontakti': kontakti,
        'pozivi_tel': pozivi_tel,
        'opomene': opomene,
        'poziv_pismo': poziv_pismo,
        'tuzbe': tuzbe,
        'napomene': napomene,
        'dugovanja_baket':dugovanja_baket,
        'dugovanja_sumarno':dugovanja_sumarno,
        'spisak_utuzenje': spisak_utuzenje,
        'opomene_fakture': opomene_fakture,
        'fakture_baket_90': fakture_baket_90,
        'fakture_baket_60':fakture_baket_60,
        'report_mode': report_mode,
        'report_allowed_sif_pos': report_allowed_sif_pos,
    })


@role_permission_required()
def export_partner_baketi_excel(request, sif_par):
    report_mode = request.GET.get('report') == '1'
    report_allowed_sif_pos = _allowed_sif_pos_from_user(request.user) if report_mode else []

    if report_mode and not request.user.is_superuser and not report_allowed_sif_pos:
        raise PermissionDenied('Nije definisana šifra posla za korisnika.')

    with connections['server_db'].cursor() as cursor:
        sql = """
            SELECT
                sb.opis,
                b.baket,
                b.sif_pos,
                b.vez_dok,
                b.dpo,
                b.duguje,
                b.potrazuje,
                b.saldo,
                sb.akcija
            FROM dodela_baketa b
            LEFT JOIN sif_baket sb ON b.baket = sb.baket
            WHERE b.sif_par = %s
        """
        params = [sif_par]
        if report_mode and not request.user.is_superuser:
            placeholders = ",".join(["%s"] * len(report_allowed_sif_pos))
            sql += f" AND b.sif_pos IN ({placeholders})"
            params.extend(report_allowed_sif_pos)
        sql += " ORDER BY b.baket"

        cursor.execute(sql, params)
        rows = cursor.fetchall()

    headers = [
        "Opis", "Baketi", "Šifra posla", "Veza", "DPO", "Duguje", "Potražuje", "Saldo", "Akcija"
    ]
    return rows_to_xlsx_response(
        f"dugovanja_baketi_partner_{sif_par}.xlsx",
        "Dugovanja baketi",
        headers,
        rows,
    )

@role_permission_required()
def export_utuzene_fakture_excel(request, sif_par):
    # 1. Izvrši upit
    with connections['server_db'].cursor() as cursor:
        cursor.execute("""
            SELECT sif_par, naz_par as Naziv, TRIM(vez_dok) AS veza, sif_pos as [šifra posla], 
                   SUM(duguje) AS duguje, SUM(potrazuje) AS potrazuje, SUM(saldo) AS saldo
            FROM dodela_baketa
            WHERE baket = 181 AND sif_par = %s
            GROUP BY sif_par, naz_par, vez_dok, sif_pos
            ORDER BY sif_par
        """, [sif_par])
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    return rows_to_xlsx_response(
        "utuzene_fakture.xlsx",
        "Za utuÅ¾enje",
        columns,
        rows,
        bold_header=True,
        fixed_column_width=18,
    )

@role_permission_required()
def export_opomene_excel(request,sif_par):
    with connections['server_db'].cursor() as cursor:
        cursor.execute("""
            SELECT sif_par, naz_par as Naziv, TRIM(vez_dok) AS veza, sif_pos as [šifra posla], 
                   SUM(duguje) AS duguje, SUM(potrazuje) AS potrazuje, SUM(saldo) AS saldo
            FROM dodela_baketa
            WHERE baket = 180 AND sif_par = %s
            GROUP BY sif_par, naz_par, vez_dok, sif_pos
            ORDER BY sif_par
        """, [sif_par])
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]

    return rows_to_xlsx_response(
        "opomene_fakture.xlsx",
        "Za opomene",
        columns,
        rows,
        bold_header=True,
        fixed_column_width=18,
    )


@role_permission_required()
def export_baket_90_excel(request, sif_par):
    with connections['server_db'].cursor() as cursor:
        cursor.execute("""
            SELECT sif_par, naz_par, LTRIM(RTRIM(vez_dok)), sif_pos,
                   SUM(duguje), SUM(potrazuje), SUM(saldo)
            FROM dodela_baketa
            WHERE baket = 90 AND sif_par = %s
            GROUP BY sif_par, naz_par, vez_dok, sif_pos
            ORDER BY sif_par
        """, [sif_par])
        data = cursor.fetchall()

    headers = ['Šifra partnera', 'Naziv', 'Veza', 'Šifra posla', 'Duguje', 'Potražuje', 'Saldo']
    return rows_to_xlsx_response("baket_90.xlsx", "Baket 90", headers, data)

@role_permission_required()
def export_baket_60_excel(request, sif_par):
    with connections['server_db'].cursor() as cursor:
        cursor.execute("""
            SELECT sif_par, naz_par, LTRIM(RTRIM(vez_dok)), sif_pos,
                   SUM(duguje), SUM(potrazuje), SUM(saldo)
            FROM dodela_baketa
            WHERE baket = 60 AND sif_par = %s
            GROUP BY sif_par, naz_par, vez_dok, sif_pos
            ORDER BY sif_par
        """, [sif_par])
        data = cursor.fetchall()

    headers = ['Šifra partnera', 'Naziv', 'Veza', 'Šifra posla', 'Duguje', 'Potražuje', 'Saldo']
    return rows_to_xlsx_response(f"baket_60_sifpar_{sif_par}.xlsx", "Baket 60", headers, data)

# <!-- ======================================================================= -->
#                           <!-- KONTAKTI -->
# <!-- ======================================================================= -->

@role_permission_required()
def lista_kontakata(request):
    kontakti = Kontakti.objects.using('server_db').all()
    return render(request, 'naplata/kontakti_lista.html', {'kontakti': kontakti})


@role_permission_required()
def dodaj_kontakt(request, sif_par, naz_par):
    cleaned_naz_par = (naz_par or "").strip()

    if request.method == "POST":
        post_data = request.POST.copy()
        post_data["sif_par"] = str(sif_par)
        post_data["naz_par"] = cleaned_naz_par
        form = KontaktiForm(post_data)
        if form.is_valid():
            kontakt = form.save(commit=False)
            kontakt.sif_par = sif_par  # Automatski dodeljujemo sifru partnera
            kontakt.naz_par = cleaned_naz_par
            kontakt.save(using='server_db')
            return redirect('naplata:detalji_partner', sif_par=sif_par)
    else:
        form = KontaktiForm(initial={'sif_par': sif_par, 'naz_par': cleaned_naz_par})

    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def izmeni_kontakt(request, id):
    kontakt = get_object_or_404(Kontakti.objects.using('server_db'), id=id)
    if request.method == "POST":
        post_data = request.POST.copy()
        post_data["sif_par"] = str(kontakt.sif_par)
        post_data["naz_par"] = (kontakt.naz_par or "").strip()
        form = KontaktiForm(post_data, instance=kontakt)
        if form.is_valid():
            kontakt = form.save(commit=False)
            kontakt.save(using='server_db')
            return redirect('naplata:detalji_partner', sif_par=kontakt.sif_par)
    else:
        form = KontaktiForm(instance=kontakt)
    return render(request, 'naplata/form_naplata.html', {'form': form})
@role_permission_required()
def obrisi_kontakt(request, id):
    kontakt = get_object_or_404(Kontakti.objects.using('server_db'), id=id)
    if request.method == "POST":
        kontakt.delete(using='server_db')
    return redirect(request.META.get('HTTP_REFERER', 'lista_kontakata'))  # Ostaje na istoj stranici


# <!-- ======================================================================= -->
#                           <!-- NAPOMENE -->
# <!-- ======================================================================= -->
@role_permission_required()
def lista_napomena(request):
    napomene = Napomene.objects.using('server_db').all()
    return render(request, 'naplata/napomena_lista.html', {'napomene': napomene})

@role_permission_required()
def dodaj_napomenu(request, sif_par, naz_par):
    if request.method == "POST":
        form = NapomeneForm(request.POST)
        if form.is_valid():
            napomena = form.save(commit=False)
            napomena.sif_par = sif_par  # Postavljamo partnera automatski
            napomena.naz_par = naz_par
            napomena.save(using='server_db')
            return redirect('naplata:detalji_partner', sif_par = sif_par)
    else:
        form = NapomeneForm(initial={'sif_par': sif_par, 'naz_par': naz_par})  # Automatsko popunjavanje

    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def izmeni_napomenu(request, id):
    napomena = get_object_or_404(Napomene.objects.using('server_db'), id=id)
    if request.method == "POST":
        form = NapomeneForm(request.POST, instance=napomena)
        if form.is_valid():
            napomena = form.save(commit=False)
            napomena.save(using='server_db')
            return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici
    else:
        form = NapomeneForm(instance=napomena)
    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def obrisi_napomenu(request, id):
    napomena = get_object_or_404(Napomene.objects.using('server_db'), id=id)
    if request.method == "POST":
        napomena.delete(using='server_db')
    return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici


# <!-- ======================================================================= -->
#                           <!-- OPOMENE -->
# <!-- ======================================================================= -->
@role_permission_required()
def lista_opomena(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                messages.error(request, 'Excel je prazan.')
                return redirect('naplata:lista_opomena')

            def _normalize_header(value):
                return str(value or '').strip().lower().replace(' ', '_')

            aliases = {
                'sif_par': {'sif_par', 'sifra', 'sifra_partnera', 'sifrapartnera'},
                'naz_par': {'naz_par', 'partner_naziv', 'partner', 'naziv_partnera'},
                'god': {'god', 'godina'},
                'br_opomene': {'br_opomene', 'broj_opomene', 'br_opomena', 'broj_opomena'},
                'datum': {'datum', 'na_dan', 'na_datum'},
                'iznos': {'iznos', 'ukupno'},
                'fakture': {'fakture', 'faktura'},
                'napomene': {'napomene', 'napomena'},
            }

            header_row = rows[0]
            header_map = {}
            for index, raw_name in enumerate(header_row):
                normalized = _normalize_header(raw_name)
                for target, names in aliases.items():
                    if normalized in names and target not in header_map:
                        header_map[target] = index

            required_fields = {'sif_par', 'naz_par', 'datum', 'iznos'}
            if not required_fields.issubset(set(header_map.keys())):
                missing = ', '.join(sorted(required_fields - set(header_map.keys())))
                messages.error(request, f'Nedostaju obavezne kolone u Excel fajlu: {missing}')
                return redirect('naplata:lista_opomena')

            def _cell_value(row, field_name):
                column_index = header_map.get(field_name)
                if column_index is None or column_index >= len(row):
                    return None
                return row[column_index]

            def _to_int(value):
                if value in (None, ''):
                    return None
                if isinstance(value, (int, float)):
                    return int(value)
                text = str(value).strip()
                if not text:
                    return None
                return int(float(text))

            def _to_decimal(value):
                if value in (None, ''):
                    return None
                if isinstance(value, Decimal):
                    return value
                if isinstance(value, (int, float)):
                    return Decimal(str(value))
                text = str(value).strip()
                if not text:
                    return None
                if ',' in text and '.' in text:
                    if text.rfind(',') > text.rfind('.'):
                        text = text.replace('.', '').replace(',', '.')
                    else:
                        text = text.replace(',', '')
                elif ',' in text:
                    text = text.replace(',', '.')
                return Decimal(text)

            def _to_datetime(value):
                if value in (None, ''):
                    return None
                if isinstance(value, datetime):
                    return value
                if isinstance(value, date):
                    return datetime.combine(value, time.min)
                text = str(value).strip()
                if not text:
                    return None
                formats = (
                    '%Y-%m-%d %H:%M:%S.%f',
                    '%Y-%m-%d %H:%M:%S',
                    '%Y-%m-%d',
                    '%d.%m.%Y',
                    '%d.%m.%Y %H:%M:%S',
                    '%d/%m/%Y',
                    '%d/%m/%Y %H:%M:%S',
                )
                for fmt in formats:
                    try:
                        parsed = datetime.strptime(text, fmt)
                        if fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                            return datetime.combine(parsed.date(), time.min)
                        return parsed
                    except ValueError:
                        continue
                raise ValueError(f'Neispravan datum: {text}')

            insert_rows = []
            skipped_rows = 0

            for row in rows[1:]:
                if not row or all(value in (None, '') for value in row):
                    continue

                try:
                    sif_par = _to_int(_cell_value(row, 'sif_par'))
                    naz_par = str(_cell_value(row, 'naz_par') or '').strip() or None
                    datum = _to_datetime(_cell_value(row, 'datum'))
                    iznos = _to_decimal(_cell_value(row, 'iznos'))

                    if sif_par is None or not naz_par or datum is None or iznos is None:
                        skipped_rows += 1
                        continue

                    god_value = int(_to_int(_cell_value(row, 'god')) or datum.year)
                    br_opomene_value = _to_int(_cell_value(row, 'br_opomene'))
                    fakture_value = _cell_value(row, 'fakture')
                    napomene_value = _cell_value(row, 'napomene')

                    insert_rows.append((
                        sif_par,
                        naz_par,
                        god_value,
                        br_opomene_value,
                        datum,
                        float(iznos),
                        str(fakture_value).strip() if fakture_value not in (None, '') else None,
                        str(napomene_value).strip() if napomene_value not in (None, '') else None,
                    ))
                except Exception:
                    skipped_rows += 1

            if insert_rows:
                with connections['server_db'].cursor() as cursor:
                    cursor.executemany(
                        """
                        INSERT INTO opomene (sif_par, naz_par, god, br_opomene, datum, iznos, fakture, napomene)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        insert_rows,
                    )

                messages.success(
                    request,
                    f'Učitano redova: {len(insert_rows)}. Preskočeno redova: {skipped_rows}.',
                )
            else:
                messages.warning(request, 'Nijedan red nije učitan. Proverite format Excel fajla.')

            return redirect('naplata:lista_opomena')
        except Exception as exc:
            messages.error(request, f'Greška pri učitavanju Excel fajla: {exc}')
            return redirect('naplata:lista_opomena')

    selected_year = request.GET.get('god')
    with connections['server_db'].cursor() as cursor:
        if selected_year:
            cursor.execute("""
                SELECT
                    sif_par,
                    naz_par,
                    CAST(god AS INT) AS god,
                    br_opomene,
                    datum,
                    iznos,
                    fakture,
                    napomene,
                    id
                FROM opomene
                WHERE god = %s
                ORDER BY god DESC, datum DESC
            """, [selected_year])
        else:
            cursor.execute("""
                SELECT
                    sif_par,
                    naz_par,
                    CAST(god AS INT) AS god,
                    br_opomene,
                    datum,
                    iznos,
                    fakture,
                    napomene,
                    id
                FROM opomene
                ORDER BY god DESC, datum DESC
            """)
        opomene = cursor.fetchall()
    return render(request, 'naplata/opomene_list.html', {
        'opomene': opomene,
        'title': 'Lista opomena',
        'selected_year': selected_year,
    })

@role_permission_required()
def dodaj_opomenu(request, sif_par, naz_par):
    if request.method == "POST":
        form = OpomeneForm(request.POST)
        if form.is_valid():
            opomena = form.save(commit=False)
            opomena.sif_par = sif_par  
            opomena.naz_par = naz_par
            opomena.save(using='server_db')
            return redirect('naplata:detalji_partner', sif_par = sif_par)
        else:
            logger.warning("Forma nije validna! %s", form.errors)
    else:
        form = OpomeneForm(initial={'sif_par': sif_par, 'naz_par': naz_par})  

    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def izmeni_opomenu(request, id):
    opomena = get_object_or_404(Opomene.objects.using('server_db'), id=id)
    if request.method == "POST":
        form = OpomeneForm(request.POST, instance=opomena)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.save(using='server_db')
            return redirect('lista_opomena')
    else:
        form = OpomeneForm(instance=opomena)
    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def obrisi_opomenu(request, id):
    opomena = get_object_or_404(Opomene.objects.using('server_db'), id=id)
    if request.method == "POST":
        opomena.delete(using='server_db')
        return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici

    return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici


# <!-- ======================================================================= -->
#                           <!-- POZIVI -->
# <!-- ======================================================================= -->
@role_permission_required()
def lista_poziva(request):
    pozivi = PoziviTel.objects.using('server_db').all()
    return render(request, 'naplata/pozivi_tel/lista.html', {'pozivi': pozivi})

@role_permission_required()
def dodaj_poziv(request, sif_par, naz_par):
    if request.method == "POST":
        form = PoziviTelForm(request.POST)
        if form.is_valid():
            poziv = form.save(commit=False)
            poziv.sif_par = sif_par  # Automatski postavljamo vrednost
            poziv.naz_par = naz_par
            poziv.save(using='server_db')
            return redirect('naplata:detalji_partner', sif_par = sif_par)

    else:
        form = PoziviTelForm(initial={'sif_par': sif_par, 'naz_par': naz_par})  

    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def izmeni_poziv(request, id):
    poziv = get_object_or_404(PoziviTel.objects.using('server_db'), id=id)
    if request.method == "POST":
        form = PoziviTelForm(request.POST, instance=poziv)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.save(using='server_db')
            return redirect('lista_poziva')
    else:
        form = PoziviTelForm(instance=poziv)
    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def obrisi_poziv(request, id):
    poziv = get_object_or_404(PoziviTel.objects.using('server_db'), id=id)
    if request.method == "POST":
        poziv.delete(using='server_db')
        return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici
    return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici


# <!-- ======================================================================= -->
#                           <!-- POZIVI/PISMA -->
# <!-- ======================================================================= -->
@role_permission_required()
def lista_pozivnih_pisma(request):
    pozivi = PozivPismo.objects.using('server_db').all()
    return render(request, 'naplata/poziv_pismo/lista.html', {'pozivi': pozivi})

@role_permission_required()
def dodaj_poziv_pismo(request, sif_par, naz_par):
    if request.method == "POST":
        form = PozivPismoForm(request.POST)
        if form.is_valid():
            poziv = form.save(commit=False)
            poziv.sif_par = sif_par  # Automatski postavljamo vrednost
            poziv.naz_par = naz_par
            poziv.save(using='server_db')
            return redirect('naplata:detalji_partner', sif_par = sif_par)
    else:
        form = PozivPismoForm(initial={'sif_par': sif_par, 'naz_par': naz_par})  

    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def izmeni_poziv_pismo(request, id):
    poziv = get_object_or_404(PozivPismo.objects.using('server_db'), id=id)
    if request.method == "POST":
        form = PozivPismoForm(request.POST, instance=poziv)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.save(using='server_db')
            return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici
    else:
        form = PozivPismoForm(instance=poziv)
    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def obrisi_poziv_pismo(request, id):
    poziv = get_object_or_404(PozivPismo.objects.using('server_db'), id=id)
    if request.method == "POST":
        poziv.delete(using='server_db')
        return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici
    return render(request, 'naplata/poziv_pismo/obrisi.html', {'poziv': poziv})


# <!-- ======================================================================= -->
#                           <!-- TUZBE -->
# <!-- ======================================================================= -->
@role_permission_required()
def lista_tuzbi(request):
    tuzbe = Tuzbe.objects.using('server_db').all()
    return render(request, 'naplata/tuzbe/lista.html', {'tuzbe': tuzbe})

@role_permission_required()
def dodaj_tuzbu(request, sif_par, naz_par):
    if request.method == "POST":
        form = TuzbeForm(request.POST)
        if form.is_valid():
            tuzba = form.save(commit=False)
            tuzba.sif_par = sif_par  # Automatski postavljamo vrednost
            tuzba.naz_par = naz_par
            tuzba.save(using='server_db')
            return redirect('naplata:detalji_partner', sif_par = sif_par)
    else:
        form = TuzbeForm(initial={'sif_par': sif_par, 'naz_par': naz_par})  

    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def izmeni_tuzbu(request, id):
    tuzba = get_object_or_404(Tuzbe.objects.using('server_db'), id=id)
    if request.method == "POST":
        form = TuzbeForm(request.POST, instance=tuzba)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.save(using='server_db')
        return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici
    else:
        form = TuzbeForm(instance=tuzba)
    return render(request, 'naplata/form_naplata.html', {'form': form})

@role_permission_required()
def obrisi_tuzbu(request, id):
    tuzba = get_object_or_404(Tuzbe.objects.using('server_db'), id=id)
    if request.method == "POST":
        tuzba.delete(using='server_db')
        return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici
    return redirect(request.META.get('HTTP_REFERER', 'lista_napomena'))  # Ostaje na istoj stranici


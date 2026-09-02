import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.models import PermissionCode, Role
from fleet.models import Employee
from mobilni.forms.mobile import MobileParkingExemptionForm, MobileUserForm
from mobilni.models import MobileAssignment, MobilePackage, MobileParkingExemption, MobileUsage, MobileUser
from mobilni.support.mobile import import_assignments, import_usages, sync_employee_links
from mobilni.withholdings import (
    REPORT_ALL,
    REPORT_EMPLOYEES,
    REPORT_FORMER_EMPLOYEES,
    REPORT_NON_EMPLOYEES,
    calculate_withholding,
    get_withholding_rows,
)
from ugovori.models import Contract, ContractType


class MobilePermissionTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("mobile-test", password="test")
        self.client.force_login(self.user)

    def test_mobile_dashboard_requires_mobile_permission(self):
        response = self.client.get(reverse("mobilni:mobile_dashboard"))

        self.assertEqual(response.status_code, 403)

    def test_mobile_role_can_open_dashboard(self):
        permission = PermissionCode.objects.create(code="mobilni:mobile_dashboard")
        role = Role.objects.create(name="Mobilni", slug="mobilni")
        role.permissions.add(permission)
        self.user.roles.add(role)

        response = self.client.get(reverse("mobilni:mobile_dashboard"))

        self.assertEqual(response.status_code, 200)

    def test_mobile_export_requires_its_permission(self):
        response = self.client.get(reverse("mobilni:mobile_assignment_export"))

        self.assertEqual(response.status_code, 403)

    def test_contract_delete_unlinks_mobile_packages_before_delete(self):
        contract_type = ContractType.objects.create(name="Mobilna telefonija")
        contract = Contract.objects.create(
            contract_type=contract_type,
            contract_number="20-test",
            title="Test ugovor",
            contract_date=datetime.date(2026, 1, 1),
        )
        package = MobilePackage.objects.create(name="Test paket", contract=contract)
        permission = PermissionCode.objects.create(code="ugovori:contract_delete")
        role = Role.objects.create(name="Brisanje ugovora", slug="brisanje-ugovora")
        role.permissions.add(permission)
        self.user.roles.add(role)

        response = self.client.post(reverse("ugovori:contract_delete", args=[contract.pk]))

        self.assertRedirects(
            response,
            reverse("ugovori:contract_list"),
            fetch_redirect_response=False,
        )
        self.assertFalse(Contract.objects.filter(pk=contract.pk).exists())
        package.refresh_from_db()
        self.assertIsNone(package.contract_id)


class MobileWithholdingTests(TestCase):
    def setUp(self):
        self.package = MobilePackage.objects.create(
            partner_code="1",
            partner_name="MTS",
            name="Paket 1",
            net_amount=Decimal("100.00"),
        )

    def create_usage(
        self,
        *,
        employee_code=500,
        phone_number="381631234567",
        employee_active=True,
        departure_date=None,
        year=2026,
        month=6,
        vat_base=Decimal("150.00"),
        parking=Decimal("20.00"),
        nzrd=Decimal("5.00"),
        total=Decimal("200.00"),
        package=None,
    ):
        package = package or self.package
        employee = self.create_employee(employee_code, is_active=employee_active)
        mobile_user, _ = MobileUser.objects.update_or_create(
            employee_code=employee_code,
            defaults={
                "full_name": f"Radnik {employee_code}",
                "is_active": employee_active,
                "departure_date": departure_date,
                "employee": employee,
                "link_status": MobileUser.LinkStatus.AUTO,
            },
        )
        assignment = MobileAssignment.objects.create(
            year=year,
            month=month,
            phone_number=phone_number,
            package=package,
            mobile_user=mobile_user,
            source_employee_code=employee_code,
            source_full_name=mobile_user.full_name,
            employee=employee,
        )
        return MobileUsage.objects.create(
            year=year,
            month=month,
            phone_number=phone_number,
            assignment=assignment,
            employee=employee,
            vat_base=vat_base,
            parking=parking,
            nzrd=nzrd,
            total=total,
        )

    def create_employee(
        self,
        employee_code,
        *,
        is_active=True,
        first_name=None,
        last_name="",
        original_full_name=None,
    ):
        employee, _ = Employee.objects.update_or_create(
            employee_code=employee_code,
            defaults={
                "first_name": first_name or f"Radnik {employee_code}",
                "last_name": last_name,
                "position": "Referent",
                "department_code": 1,
                "org_unit_code": "1",
                "gender": "M",
                "date_of_birth": datetime.date(1990, 1, 1),
                "date_of_joining": datetime.date(2026, 1, 1),
                "personal_number": f"0101990{employee_code:06d}"[:13],
                "is_active": is_active,
                "original_full_name": original_full_name or "",
            },
        )
        return employee

    def grant_permission(self, user, code):
        permission = PermissionCode.objects.create(code=code)
        role = Role.objects.create(name=f"Role {code}", slug=code.replace(":", "-"))
        role.permissions.add(permission)
        user.roles.add(role)

    def test_calculate_withholding_matches_source_view_formula(self):
        usage = self.create_usage()

        self.assertEqual(calculate_withholding(usage), Decimal("75.00"))

    def test_calculate_withholding_applies_parking_exemption_by_phone_number(self):
        MobileParkingExemption.objects.create(phone_number="381631111111")
        parking_exempt = self.create_usage(employee_code=999, phone_number="381631111111")

        self.assertEqual(calculate_withholding(parking_exempt), Decimal("55.00"))

    def test_calculate_withholding_applies_parking_exemption_while_number_exists(self):
        MobileParkingExemption.objects.create(phone_number="381 63 111 1111")
        usage = self.create_usage(employee_code=999, phone_number="381631111111")

        self.assertEqual(calculate_withholding(usage), Decimal("55.00"))

    def test_calculate_withholding_applies_special_phone_exception(self):
        special_phone = self.create_usage(employee_code=501, phone_number="381637781481")

        self.assertEqual(calculate_withholding(special_phone), Decimal("175.00"))

    def test_reports_separate_current_and_former_employees(self):
        active = self.create_usage(employee_code=500, phone_number="381631111111")
        former = self.create_usage(
            employee_code=501,
            phone_number="381632222222",
            employee_active=False,
            departure_date=datetime.date(2026, 5, 31),
        )
        self.create_usage(
            employee_code=502,
            phone_number="381633333333",
            employee_active=False,
            departure_date=datetime.date(2026, 6, 30),
        )

        employee_rows = get_withholding_rows(REPORT_EMPLOYEES, year=2026, month=6)
        former_rows = get_withholding_rows(REPORT_FORMER_EMPLOYEES, year=2026, month=6)

        self.assertEqual([row.usage for row in employee_rows], [active])
        self.assertEqual([row.usage for row in former_rows], [former])

    def test_employee_report_excludes_unassigned_active_numbers(self):
        usage = self.create_usage()
        usage.assignment.mobile_user = None
        usage.assignment.employee = None
        usage.assignment.save(update_fields=["mobile_user", "employee"])

        rows = get_withholding_rows(REPORT_EMPLOYEES, year=2026, month=6)

        self.assertEqual(rows, [])

    def test_employee_csv_contains_only_required_columns(self):
        self.create_usage()
        self.create_usage(
            employee_code=501,
            phone_number="381632222222",
            vat_base=Decimal("100.00"),
            parking=Decimal("0.00"),
            nzrd=Decimal("0.00"),
            total=Decimal("100.00"),
        )
        user = get_user_model().objects.create_user("withholding-export", password="test")
        self.grant_permission(user, "mobilni:mobile_withholding_employees_export")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_withholding_employees_export"),
            {"year": 2026, "month": 6},
        )

        self.assertEqual(response.status_code, 200)
        lines = response.content.decode("utf-8-sig").splitlines()
        self.assertEqual(lines[0], "Godina;Mesec;Šifra radnika;Iznos obustave")
        self.assertEqual(lines[1], "2026;6;500;75.00")
        self.assertEqual(len(lines), 2)

    def test_usage_accounting_csv_uses_period_and_phone_filter(self):
        self.create_usage(employee_code=500, phone_number="381631111111")
        self.create_usage(employee_code=501, phone_number="381632222222")
        user = get_user_model().objects.create_user("usage-accounting-export", password="test")
        self.grant_permission(user, "mobilni:mobile_usage_accounting_csv")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_usage_accounting_csv"),
            {"year": 2026, "month": 6, "phone_number": "1111"},
        )

        self.assertEqual(response.status_code, 200)
        lines = response.content.decode("utf-8-sig").splitlines()
        self.assertEqual(lines[0], "Godina;Mesec;Šifra radnika;Iznos obustave")
        self.assertEqual(lines[1], "2026;6;500;75.00")
        self.assertEqual(len(lines), 2)

    def test_employee_report_page_renders_calculated_row(self):
        self.create_usage()
        user = get_user_model().objects.create_user("withholding-report", password="test")
        self.grant_permission(user, "mobilni:mobile_withholding_employees")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_withholding_employees"),
            {"year": 2026, "month": 6},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Radnik 500")
        self.assertContains(response, "75.00")
        self.assertContains(response, "Obustava = osnovica za PDV - neto iznos paketa + parking + NZRD")
        self.assertContains(response, "381637781481")
        self.assertContains(response, "Šifra radnika")

    def test_withholding_report_page_uses_phone_filter(self):
        self.create_usage(employee_code=500, phone_number="381631111111")
        self.create_usage(employee_code=501, phone_number="381632222222")
        MobileParkingExemption.objects.create(phone_number="381631111111")
        user = get_user_model().objects.create_user("withholding-phone-filter", password="test")
        self.grant_permission(user, "mobilni:mobile_withholding_all")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_withholding_all"),
            {"year": 2026, "month": 6, "phone_number": "1111"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="withholding-phone"')
        self.assertContains(response, "Broj telefona")
        self.assertContains(response, "381631111111")
        self.assertContains(response, "Parking izuzet")
        self.assertContains(response, "Ne ide u obustavu")
        self.assertNotContains(response, "381632222222")
        self.assertEqual(response.context["phone_number"], "1111")

    def test_non_employee_report_lists_assignments_without_employee(self):
        usage = self.create_usage(phone_number="381631111111")
        usage.assignment.mobile_user.employee = None
        usage.assignment.mobile_user.link_status = MobileUser.LinkStatus.NON_EMPLOYEE
        usage.assignment.mobile_user.save(update_fields=["employee", "link_status"])
        usage.assignment.employee = None
        usage.assignment.save(update_fields=["employee"])
        usage.employee = None
        usage.save(update_fields=["employee"])
        user = get_user_model().objects.create_user("withholding-non-employee", password="test")
        self.grant_permission(user, "mobilni:mobile_withholding_non_employees")
        self.client.force_login(user)

        rows = get_withholding_rows(REPORT_NON_EMPLOYEES, year=2026, month=6)
        response = self.client.get(
            reverse("mobilni:mobile_withholding_non_employees"),
            {"year": 2026, "month": 6},
        )

        self.assertEqual([row.usage for row in rows], [usage])
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nezaposleni")
        self.assertContains(response, "381631111111")
        self.assertEqual(response.context["row_count"], 1)

    def test_all_report_can_be_filtered_by_phone_employee_and_package(self):
        matching = self.create_usage(employee_code=500, phone_number="381631111111")
        other_package = MobilePackage.objects.create(
            partner_code="1",
            partner_name="MTS",
            name="Paket 2",
            net_amount=Decimal("50.00"),
        )
        self.create_usage(
            employee_code=501,
            phone_number="381632222222",
            package=other_package,
        )

        rows = get_withholding_rows(
            REPORT_ALL,
            year=2026,
            month=6,
            phone_number="1111",
            employee="Radnik 500",
            package_id=self.package.pk,
        )

        self.assertEqual([row.usage for row in rows], [matching])

    def test_usage_page_uses_phone_filter(self):
        self.create_usage(employee_code=500, phone_number="381631111111")
        self.create_usage(employee_code=501, phone_number="381632222222")
        MobileParkingExemption.objects.create(phone_number="381631111111")
        user = get_user_model().objects.create_user("usage-list", password="test")
        self.grant_permission(user, "mobilni:mobile_usage_list")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_usage_list"),
            {"year": 2026, "month": 6, "phone_number": "1111"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="usage-phone"')
        self.assertContains(response, "Broj telefona")
        self.assertContains(response, "Potrošnja mobilnih")
        self.assertContains(response, "Šifra radnika")
        self.assertEqual(len(response.context["usages"]), 1)
        self.assertContains(response, "381631111111")
        self.assertContains(response, "Parking izuzet")
        self.assertContains(response, "Ne ide u obustavu")
        self.assertNotContains(response, "381632222222")

    def test_usage_page_allows_empty_period_filter(self):
        self.create_usage(employee_code=500, phone_number="381631111111", year=2026, month=6)
        self.create_usage(employee_code=501, phone_number="381632222222", year=2025, month=5)
        user = get_user_model().objects.create_user("usage-list-all-periods", password="test")
        self.grant_permission(user, "mobilni:mobile_usage_list")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_usage_list"),
            {"period": "", "year": "", "month": ""},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Svi periodi")
        self.assertIsNone(response.context["selected_year"])
        self.assertIsNone(response.context["selected_month"])
        self.assertEqual(len(response.context["usages"]), 2)

    def test_period_pages_default_to_latest_existing_period(self):
        self.create_usage(employee_code=510, phone_number="381631111111", year=2025, month=12)
        self.create_usage(employee_code=511, phone_number="381632222222", year=2026, month=7)
        user = get_user_model().objects.create_user("mobile-period-default", password="test")
        self.grant_permission(user, "mobilni:mobile_assignment_list")
        self.grant_permission(user, "mobilni:mobile_usage_list")
        self.grant_permission(user, "mobilni:mobile_dashboard")
        self.client.force_login(user)

        for route_name in [
            "mobilni:mobile_assignment_list",
            "mobilni:mobile_usage_list",
            "mobilni:mobile_dashboard",
        ]:
            response = self.client.get(reverse(route_name))

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.context["selected_year"], 2026)
            self.assertEqual(response.context["selected_month"], 7)

    def test_period_pages_fallback_to_current_period_without_data(self):
        today = timezone.localdate()
        user = get_user_model().objects.create_user("mobile-period-empty-default", password="test")
        self.grant_permission(user, "mobilni:mobile_assignment_list")
        self.client.force_login(user)

        response = self.client.get(reverse("mobilni:mobile_assignment_list"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_year"], today.year)
        self.assertEqual(response.context["selected_month"], today.month)

    def test_import_period_forms_default_to_current_period(self):
        today = timezone.localdate()
        user = get_user_model().objects.create_user("mobile-import-default", password="test")
        self.grant_permission(user, "mobilni:mobile_import")
        self.client.force_login(user)

        response = self.client.get(reverse("mobilni:mobile_import"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["forms"]["assignments"].initial["year"], today.year)
        self.assertEqual(response.context["forms"]["assignments"].initial["month"], today.month)
        self.assertEqual(response.context["forms"]["usages"].initial["year"], today.year)
        self.assertEqual(response.context["forms"]["usages"].initial["month"], today.month)

    def test_assignment_page_marks_parking_exempt_numbers(self):
        self.create_usage(employee_code=500, phone_number="381631111111")
        MobileParkingExemption.objects.create(phone_number="381631111111")
        user = get_user_model().objects.create_user("assignment-list", password="test")
        self.grant_permission(user, "mobilni:mobile_assignment_list")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_assignment_list"),
            {"year": 2026, "month": 6},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "381631111111")
        self.assertContains(response, "Parking izuzet")

    def test_assignment_page_links_phone_detail_and_employee(self):
        usage = self.create_usage(employee_code=500, phone_number="381631111111")
        user = get_user_model().objects.create_user("assignment-links", password="test")
        self.grant_permission(user, "mobilni:mobile_assignment_list")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_assignment_list"),
            {"year": 2026, "month": 6},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            reverse("mobilni:mobile_phone_detail", args=[usage.phone_number]),
        )
        self.assertContains(response, reverse("employee_detail", args=[usage.assignment.employee_id]))
        self.assertContains(response, "Sifra: 500")

    def test_phone_detail_page_renders_assignment_and_usage_history(self):
        self.create_usage(employee_code=500, phone_number="381631111111")
        user = get_user_model().objects.create_user("phone-detail", password="test")
        self.grant_permission(user, "mobilni:mobile_assignment_list")
        self.client.force_login(user)

        response = self.client.get(reverse("mobilni:mobile_phone_detail", args=["381631111111"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "381631111111")
        self.assertContains(response, "Paket 1")
        self.assertContains(response, "Radnik 500")
        self.assertContains(response, "Sifra: 500")
        self.assertContains(response, "75.00")

    def test_parking_exemption_form_uses_only_phone_number(self):
        self.create_usage(employee_code=500, phone_number="381631111111")
        form = MobileParkingExemptionForm(data={"phone_number": "381 63 111 1111"})
        choices = dict(MobileParkingExemptionForm().fields["phone_number"].widget.choices)

        self.assertTrue(form.is_valid())
        self.assertEqual(list(form.fields), ["phone_number"])
        self.assertEqual(form.cleaned_data["phone_number"], "381631111111")
        self.assertIn("select2-method", form.fields["phone_number"].widget.attrs["class"])
        self.assertIn("381631111111", choices)
        self.assertIn("Radnik 500", choices["381631111111"])

    def test_mobile_user_form_renders_employee_choices(self):
        employee = self.create_employee(503)

        form = MobileUserForm()
        rendered = str(form["employee"])

        self.assertIn(f'value="{employee.pk}"', rendered)
        self.assertIn("503 -", rendered)
        self.assertNotIn("django-select2", rendered)

    def test_parking_exemption_page_renders_configured_numbers(self):
        MobileParkingExemption.objects.create(phone_number="381631111111")
        user = get_user_model().objects.create_user("parking-exemption-list", password="test")
        self.grant_permission(user, "mobilni:mobile_parking_exemption_list")
        self.client.force_login(user)

        response = self.client.get(reverse("mobilni:mobile_parking_exemption_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Izuzeci parkinga")
        self.assertContains(response, "381631111111")
        self.assertContains(response, "Parking izuzet")
        self.assertContains(response, "Parking se ne obracunava za obustavu")

    def test_usage_excel_contains_obustava_and_respects_phone_filter(self):
        self.create_usage(employee_code=500, phone_number="381631111111")
        self.create_usage(employee_code=501, phone_number="381632222222")
        user = get_user_model().objects.create_user("usage-export", password="test")
        self.grant_permission(user, "mobilni:mobile_usage_export")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_usage_export"),
            {"year": 2026, "month": 6, "phone_number": "1111"},
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(workbook.active.title, "Potrošnja")
        self.assertIn("Šifra radnika", rows[0])
        self.assertEqual(rows[0][-1], "Obustava")
        self.assertIn("Parking izuzet", rows[0])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][6], "381631111111")
        self.assertEqual(rows[1][7], "Ne")
        self.assertEqual(rows[1][-1], 75)

    def test_dashboard_calculates_package_and_institute_totals(self):
        self.create_usage(
            employee_code=500,
            phone_number="381631111111",
            total=Decimal("200.00"),
        )
        MobileParkingExemption.objects.create(phone_number="381631111111")
        self.create_usage(
            employee_code=501,
            phone_number="381632222222",
            vat_base=Decimal("130.00"),
            parking=Decimal("0.00"),
            nzrd=Decimal("0.00"),
            total=Decimal("180.00"),
        )
        user = get_user_model().objects.create_user("mobile-dashboard", password="test")
        self.grant_permission(user, "mobilni:mobile_dashboard")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_dashboard"),
            {"year": 2026, "month": 6},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="mobile-phone"')
        self.assertContains(response, "Broj telefona")
        self.assertContains(response, "Parking izuzet")
        self.assertEqual(response.context["usage_total"], Decimal("380.00"))
        self.assertEqual(response.context["withholding_total"], Decimal("85.00"))
        self.assertEqual(response.context["institute_total"], Decimal("295.00"))
        self.assertEqual(response.context["active_number_count"], 2)
        self.assertEqual(response.context["contracted_number_count"], 2)
        self.assertEqual(response.context["package_summary"][0]["number_count"], 2)

    def test_dashboard_renders_monthly_usage_and_withholding_charts(self):
        other_package = MobilePackage.objects.create(
            partner_code="2",
            partner_name="MTS",
            name="Paket 2",
            net_amount=Decimal("50.00"),
        )
        self.create_usage(
            employee_code=600,
            phone_number="381631111111",
            year=2026,
            month=1,
            vat_base=Decimal("130.00"),
            parking=Decimal("0.00"),
            nzrd=Decimal("0.00"),
            total=Decimal("100.00"),
        )
        self.create_usage(
            employee_code=601,
            phone_number="381632222222",
            year=2026,
            month=1,
            package=other_package,
            vat_base=Decimal("90.00"),
            parking=Decimal("10.00"),
            nzrd=Decimal("0.00"),
            total=Decimal("300.00"),
        )
        self.create_usage(
            employee_code=602,
            phone_number="381633333333",
            year=2026,
            month=2,
            vat_base=Decimal("150.00"),
            parking=Decimal("0.00"),
            nzrd=Decimal("5.00"),
            total=Decimal("50.00"),
        )
        user = get_user_model().objects.create_user("mobile-dashboard-chart", password="test")
        self.grant_permission(user, "mobilni:mobile_dashboard")
        self.client.force_login(user)

        response = self.client.get(
            reverse("mobilni:mobile_dashboard"),
            {"year": 2026, "month": 1},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="mobile-usage-monthly-chart"')
        self.assertContains(response, 'id="mobile-withholding-monthly-chart"')
        usage_chart = response.context["usage_monthly_chart"]
        january_segments = {item["label"]: item["amount"] for item in usage_chart["months"][0]["segments"]}
        self.assertEqual(usage_chart["months"][0]["total"], Decimal("400.00"))
        self.assertEqual(january_segments["Paket 1"], Decimal("100.00"))
        self.assertEqual(january_segments["Paket 2"], Decimal("300.00"))
        withholding_chart = response.context["withholding_monthly_chart"]
        self.assertEqual(withholding_chart["months"][0]["total"], Decimal("80.00"))
        self.assertEqual(withholding_chart["months"][1]["total"], Decimal("55.00"))

    def test_employee_sync_does_not_guess_active_name_for_old_code(self):
        old_employee = self.create_employee(
            519,
            is_active=False,
            first_name="Dr.zagorka",
            last_name="Radojevic",
            original_full_name="RADOJEVIC DR ZAGORKA",
        )
        active_employee = self.create_employee(
            1037,
            is_active=True,
            first_name="Zagorka",
            last_name="Radojevic",
            original_full_name="RADOJEVIC DR ZAGORKA",
        )
        mobile_user = MobileUser.objects.create(
            employee_code=519,
            full_name="Radojevic Dr. Zagorka",
            is_active=True,
            employee=old_employee,
            link_status=MobileUser.LinkStatus.AUTO,
        )
        assignment = MobileAssignment.objects.create(
            year=2026,
            month=6,
            phone_number="38163384942",
            package=self.package,
            mobile_user=mobile_user,
            source_employee_code=519,
            source_full_name="Radojevic Dr. Zagorka",
            employee=old_employee,
        )
        usage = MobileUsage.objects.create(
            year=2026,
            month=6,
            phone_number="38163384942",
            assignment=assignment,
            employee=old_employee,
            vat_base=Decimal("150.00"),
            total=Decimal("200.00"),
        )

        result = sync_employee_links()
        mobile_user.refresh_from_db()
        assignment.refresh_from_db()
        usage.refresh_from_db()

        self.assertEqual(result["mobile_users_linked"], 1)
        self.assertEqual(result["assignments_employee_linked"], 1)
        self.assertEqual(result["usages_employee_linked"], 1)
        self.assertIsNone(mobile_user.employee)
        self.assertEqual(mobile_user.link_status, MobileUser.LinkStatus.AMBIGUOUS)
        self.assertEqual(assignment.mobile_user, mobile_user)
        self.assertIsNone(assignment.employee)
        self.assertIsNone(usage.employee)
        self.assertNotEqual(active_employee.employee_code, mobile_user.employee_code)
        self.assertEqual(get_withholding_rows(REPORT_FORMER_EMPLOYEES, year=2026, month=6), [])
        self.assertEqual(get_withholding_rows(REPORT_EMPLOYEES, year=2026, month=6), [])
        self.assertEqual([row.usage for row in get_withholding_rows(REPORT_NON_EMPLOYEES, year=2026, month=6)], [usage])

    def test_employee_sync_preserves_manual_link_to_new_code(self):
        old_employee = self.create_employee(
            519,
            is_active=False,
            first_name="Dr.zagorka",
            last_name="Radojevic",
            original_full_name="RADOJEVIC DR ZAGORKA",
        )
        active_employee = self.create_employee(
            1037,
            is_active=True,
            first_name="Zagorka",
            last_name="Radojevic",
            original_full_name="RADOJEVIC DR ZAGORKA",
        )
        mobile_user = MobileUser.objects.create(
            employee_code=519,
            full_name="Radojevic Dr. Zagorka",
            is_active=True,
            employee=active_employee,
            link_status=MobileUser.LinkStatus.MANUAL,
        )
        assignment = MobileAssignment.objects.create(
            year=2026,
            month=6,
            phone_number="38163384942",
            package=self.package,
            mobile_user=mobile_user,
            source_employee_code=519,
            source_full_name="Radojevic Dr. Zagorka",
            employee=old_employee,
        )
        usage = MobileUsage.objects.create(
            year=2026,
            month=6,
            phone_number="38163384942",
            assignment=assignment,
            employee=old_employee,
            vat_base=Decimal("150.00"),
            total=Decimal("200.00"),
        )

        result = sync_employee_links()
        mobile_user.refresh_from_db()
        assignment.refresh_from_db()
        usage.refresh_from_db()

        self.assertEqual(result["mobile_users_linked"], 0)
        self.assertEqual(result["assignments_employee_linked"], 1)
        self.assertEqual(result["usages_employee_linked"], 1)
        self.assertEqual(mobile_user.employee, active_employee)
        self.assertEqual(mobile_user.link_status, MobileUser.LinkStatus.MANUAL)
        self.assertEqual(assignment.employee, active_employee)
        self.assertEqual(usage.employee, active_employee)
        self.assertEqual(get_withholding_rows(REPORT_FORMER_EMPLOYEES, year=2026, month=6), [])
        self.assertEqual([row.usage for row in get_withholding_rows(REPORT_EMPLOYEES, year=2026, month=6)], [usage])

    def test_import_assignments_links_existing_employee_and_package(self):
        employee = self.create_employee(700)
        uploaded_file = SimpleUploadedFile(
            "dodele.csv",
            b"broj;paket;rasif\n381 63 999 9999;Paket 1;700\n",
            content_type="text/csv",
        )

        result = import_assignments(uploaded_file, year=2026, month=6)

        assignment = MobileAssignment.objects.get(phone_number="381639999999")
        self.assertEqual(result.imported, 1)
        self.assertEqual(result.updated, 0)
        self.assertEqual(assignment.employee, employee)
        self.assertEqual(assignment.mobile_user.employee, employee)
        self.assertEqual(assignment.mobile_user.link_status, MobileUser.LinkStatus.AUTO)
        self.assertEqual(assignment.source_employee_code, 700)
        self.assertEqual(assignment.package, self.package)

    def test_import_assignments_keeps_assignment_without_existing_employee(self):
        uploaded_file = SimpleUploadedFile(
            "dodele.csv",
            b"broj;paket;rasif\n381639999998;Nepostojeci;9999\n381639999997;Paket 1;9999\n",
            content_type="text/csv",
        )

        result = import_assignments(uploaded_file, year=2026, month=6)

        assignment = MobileAssignment.objects.get(phone_number="381639999997")
        self.assertEqual(result.imported, 1)
        self.assertEqual(result.skipped, 1)
        self.assertEqual(assignment.package, self.package)
        self.assertIsNone(assignment.employee)
        self.assertIsNotNone(assignment.mobile_user)
        self.assertEqual(assignment.mobile_user.employee_code, 9999)
        self.assertIsNone(assignment.mobile_user.employee)
        self.assertEqual(assignment.mobile_user.link_status, MobileUser.LinkStatus.UNMATCHED)
        self.assertEqual(assignment.source_employee_code, 9999)
        self.assertIn("paket 'Nepostojeci' nije pronadjen", result.errors[0])

    def test_import_assignments_does_not_guess_active_employee_by_name(self):
        self.create_employee(
            519,
            is_active=False,
            first_name="Dr.zagorka",
            last_name="Radojevic",
            original_full_name="RADOJEVIC DR ZAGORKA",
        )
        self.create_employee(
            1037,
            is_active=True,
            first_name="Zagorka",
            last_name="Radojevic",
            original_full_name="RADOJEVIC DR ZAGORKA",
        )
        uploaded_file = SimpleUploadedFile(
            "dodele.csv",
            b"broj;paket;rasif;ranaz;aktivan_radnik\n38163384942;Paket 1;519;Radojevic Dr. Zagorka;1\n",
            content_type="text/csv",
        )

        result = import_assignments(uploaded_file, year=2026, month=6)

        assignment = MobileAssignment.objects.select_related("mobile_user", "mobile_user__employee").get(
            phone_number="38163384942",
        )
        self.assertEqual(result.imported, 1)
        self.assertEqual(assignment.source_employee_code, 519)
        self.assertEqual(assignment.source_full_name, "Radojevic Dr. Zagorka")
        self.assertEqual(assignment.mobile_user.link_status, MobileUser.LinkStatus.AMBIGUOUS)
        self.assertIsNone(assignment.mobile_user.employee)
        self.assertIsNone(assignment.employee)

    def test_import_usages_does_not_link_previous_period_assignment(self):
        employee = self.create_employee(701)
        mobile_user = MobileUser.objects.create(
            employee_code=701,
            full_name="Radnik 701",
            employee=employee,
            link_status=MobileUser.LinkStatus.AUTO,
        )
        MobileAssignment.objects.create(
            year=2026,
            month=5,
            phone_number="381639999996",
            package=self.package,
            mobile_user=mobile_user,
            source_employee_code=701,
            source_full_name="Radnik 701",
            employee=employee,
        )
        uploaded_file = SimpleUploadedFile(
            "potrosnja.csv",
            b"pretpl.broj;ukupno za naplatu;osnovica za pdv\n381639999996;200;150\n",
            content_type="text/csv",
        )

        result = import_usages(uploaded_file, year=2026, month=6)

        usage = MobileUsage.objects.get(phone_number="381639999996", year=2026, month=6)
        self.assertEqual(result.imported, 1)
        self.assertIsNone(usage.assignment)
        self.assertIsNone(usage.employee)

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import connections, transaction

from .apr_openapi import APR_OPENAPI_SOURCE
from .models import Contract, ContractParty, ContractType, Partner


@dataclass
class PartnerSyncResult:
    loaded: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    skipped: int = 0


@dataclass
class ContractImportResult:
    rows: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    skipped_duplicate: int = 0
    skipped_invalid: int = 0
    parties_rows: int = 0
    parties_created: int = 0
    parties_unchanged: int = 0
    parties_skipped: int = 0
    commit: bool = False


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if text.endswith(".") and text.count(".") >= 3:
        text = text.rstrip(".")
    for date_format in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def parse_excel_decimal(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def residency_from_country(country):
    country = (country or "").strip().upper()
    if not country or country in {"RS", "SRB", "SRBIJA", "SERBIA"}:
        return Partner.DOMESTIC
    return Partner.FOREIGN


BANK_GROUPS = {11, 13, 14}
APR_PROTECTED_PARTNER_FIELDS = {"name", "is_active"}


def parse_finance_group(value):
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def partner_type_from_finance_group(group):
    group = parse_finance_group(group)
    if group == 10:
        return Partner.PERSON
    if group in BANK_GROUPS:
        return Partner.BANK
    return Partner.LEGAL_ENTITY


def get_finance_partner(sif_par, source_db="server_db"):
    rows = fetch_finance_partners(source_db=source_db, sif_par=sif_par)
    return rows[0] if rows else None


def count_finance_partners(source_db="server_db"):
    with connections[source_db].cursor() as cursor:
        cursor.execute("SELECT COUNT(DISTINCT sif_par) FROM dbo.partneri WHERE sif_par IS NOT NULL")
        return int(cursor.fetchone()[0] or 0)


def fetch_finance_partners(source_db="server_db", limit=None, offset=None, sif_par=None):
    use_offset = offset is not None
    top_sql = f"TOP ({int(limit)}) " if limit and not use_offset else ""
    where = ["rn = 1"]
    params = []
    if sif_par is not None:
        where.append("sif_par = %s")
        params.append(sif_par)

    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    paging_sql = ""
    if use_offset:
        offset_value = max(int(offset or 0), 0)
        limit_value = max(int(limit or 250), 1)
        paging_sql = "OFFSET %s ROWS FETCH NEXT %s ROWS ONLY"
        params.extend([offset_value, limit_value])

    sql = f"""
        WITH ranked_partners AS (
            SELECT
                naz_grup,
                grupa,
                sif_par,
                naz_par,
                ulica_par,
                mesto_par,
                mb,
                telefon,
                email,
                lice,
                pib,
                zemlja,
                ROW_NUMBER() OVER (
                    PARTITION BY sif_par
                    ORDER BY
                        CASE
                            WHEN grupa = 1 THEN 1
                            WHEN grupa = 10 THEN 2
                            WHEN grupa IN (11, 13, 14) THEN 3
                            ELSE 4
                        END,
                        grupa
                ) AS rn
            FROM dbo.partneri
            WHERE sif_par IS NOT NULL
        )
        SELECT {top_sql}
            naz_grup,
            grupa,
            sif_par,
            naz_par,
            ulica_par,
            mesto_par,
            mb,
            telefon,
            email,
            lice,
            pib,
            zemlja
        FROM ranked_partners
        {where_sql}
        ORDER BY sif_par
        {paging_sql}
    """

    with connections[source_db].cursor() as cursor:
        cursor.execute(sql, params)
        columns = [column[0] for column in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def partner_defaults_from_finance(row):
    return {
        "name": clean_text(row["naz_par"]) or f"Partner {row['sif_par']}",
        "partner_type": partner_type_from_finance_group(row.get("grupa")),
        "residency": residency_from_country(row["zemlja"]),
        "pib": clean_text(row["pib"]),
        "maticni_broj": clean_text(row["mb"]),
        "country": clean_text(row["zemlja"]),
        "city": clean_text(row["mesto_par"]),
        "address": clean_text(row["ulica_par"]),
        "email": clean_text(row["email"]),
        "phone": clean_text(row["telefon"]),
        "contact_person": clean_text(row["lice"]),
        "is_active": True,
    }


def sync_partner_from_finance(sif_par, source_db="server_db", target_db="default", commit=True):
    return sync_finance_partners(
        source_db=source_db,
        target_db=target_db,
        sif_par=sif_par,
        commit=commit,
    )


def sync_finance_partners(source_db="server_db", target_db="default", limit=None, sif_par=None, commit=True):
    rows = fetch_finance_partners(source_db=source_db, limit=limit, sif_par=sif_par)
    return sync_finance_partner_rows(rows, target_db=target_db, commit=commit)


def sync_finance_partner_batch(source_db="server_db", target_db="default", offset=0, limit=250, commit=True):
    rows = fetch_finance_partners(source_db=source_db, limit=limit, offset=offset)
    return sync_finance_partner_rows(rows, target_db=target_db, commit=commit)


def sync_finance_partner_rows(rows, target_db="default", commit=True):
    result = PartnerSyncResult(loaded=len(rows))
    sif_values = []
    seen_sif_values = set()
    for row in rows:
        sif_par = row["sif_par"]
        if sif_par is not None and sif_par not in seen_sif_values:
            sif_values.append(sif_par)
            seen_sif_values.add(sif_par)

    with transaction.atomic(using=target_db):
        existing_partners = {}
        if sif_values:
            partners = (
                Partner.objects.using(target_db)
                .filter(external_sif_par__in=sif_values)
                .order_by("id")
            )
            for partner in partners:
                existing_partners.setdefault(partner.external_sif_par, partner)

        for row in rows:
            if row["sif_par"] is None:
                result.skipped += 1
                continue

            defaults = partner_defaults_from_finance(row)
            partner = existing_partners.get(row["sif_par"])
            if partner is not None and partner.data_source == APR_OPENAPI_SOURCE and partner.data_validated:
                defaults = {
                    field: value
                    for field, value in defaults.items()
                    if field not in APR_PROTECTED_PARTNER_FIELDS
                }

            if partner is None:
                result.created += 1
                if commit:
                    partner = Partner.objects.using(target_db).create(
                        external_sif_par=row["sif_par"],
                        **defaults,
                    )
                    existing_partners[row["sif_par"]] = partner
                continue

            changed_fields = []
            for field, value in defaults.items():
                if getattr(partner, field) != value:
                    setattr(partner, field, value)
                    changed_fields.append(field)

            if changed_fields:
                result.updated += 1
                if commit:
                    partner.save(using=target_db, update_fields=[*changed_fields, "updated_at"])
            else:
                result.unchanged += 1

        if not commit:
            transaction.set_rollback(True, using=target_db)

    return result


CONTRACT_IMPORT_FIELDS = {
    "kind",
    "contract_type",
    "parent_contract",
    "title",
    "subject",
    "contract_date",
    "valid_from",
    "valid_to",
    "value",
    "currency",
    "status",
    "note",
}


def excel_headers(worksheet):
    return [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]


def excel_rows(worksheet):
    headers = excel_headers(worksheet)
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = dict(zip(headers, row))
        yield row_index, values


def normalize_contract_number(value):
    value = clean_text(value)
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return value


def get_contract_type_from_excel(row):
    type_id = row.get("contract_type_id")
    type_name = clean_text(row.get("contract_type_name"))
    if type_id not in (None, ""):
        try:
            contract_type = ContractType.objects.filter(pk=int(float(str(type_id).strip()))).first()
            if contract_type:
                return contract_type
        except (TypeError, ValueError):
            pass
    if type_name:
        return ContractType.objects.filter(name__iexact=type_name).first()
    return None


def contract_defaults_from_excel(row, contract_type, parent_contract=None):
    value = parse_excel_decimal(row.get("value"))
    value_raw = clean_text(row.get("value_raw"))
    note = clean_text(row.get("note"))
    if value is None and value_raw:
        note = f"{note}\nVrednost: {value_raw}" if note else f"Vrednost: {value_raw}"

    return {
        "kind": clean_text(row.get("kind")) or Contract.MAIN,
        "contract_type": contract_type,
        "parent_contract": parent_contract,
        "title": clean_text(row.get("title")) or normalize_contract_number(row.get("contract_number")),
        "subject": clean_text(row.get("subject")),
        "contract_date": parse_excel_date(row.get("contract_date")),
        "valid_from": parse_excel_date(row.get("valid_from")),
        "valid_to": parse_excel_date(row.get("valid_to")),
        "value": value,
        "currency": clean_text(row.get("currency")) or "RSD",
        "status": clean_text(row.get("status")) or Contract.STATUS_ACTIVE,
        "note": note,
    }


def update_contract_from_defaults(contract, defaults):
    changed_fields = []
    for field_name in CONTRACT_IMPORT_FIELDS:
        value = defaults[field_name]
        if getattr(contract, field_name) != value:
            setattr(contract, field_name, value)
            changed_fields.append(field_name)
    return changed_fields


def import_contracts_from_excel(
    file_path,
    *,
    contracts_sheet="contracts_import",
    parties_sheet="contract_parties_import",
    commit=False,
):
    from openpyxl import load_workbook

    workbook_path = Path(file_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if contracts_sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet '{contracts_sheet}' ne postoji u fajlu {workbook_path}.")
    if parties_sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet '{parties_sheet}' ne postoji u fajlu {workbook_path}.")

    result = ContractImportResult(commit=commit)
    contracts_ws = workbook[contracts_sheet]
    parties_ws = workbook[parties_sheet]

    seen_numbers = set()
    contract_rows = []
    for row_index, row in excel_rows(contracts_ws):
        contract_number = normalize_contract_number(row.get("contract_number"))
        if not contract_number:
            continue
        result.rows += 1
        if contract_number in seen_numbers:
            result.skipped_duplicate += 1
            continue
        seen_numbers.add(contract_number)
        contract_rows.append((row_index, contract_number, row))

    imported_contracts = {
        contract.contract_number: contract
        for contract in Contract.objects.filter(contract_number__in=seen_numbers)
    }

    with transaction.atomic():
        # Main contracts first, then annexes whose parent can now be resolved.
        for wanted_kind in (Contract.MAIN, Contract.ANNEX):
            for row_index, contract_number, row in contract_rows:
                kind = clean_text(row.get("kind")) or Contract.MAIN
                if kind != wanted_kind:
                    continue

                parent_contract = None
                contract_type = get_contract_type_from_excel(row)
                if kind == Contract.ANNEX:
                    parent_number = normalize_contract_number(row.get("parent_contract_number"))
                    parent_contract = imported_contracts.get(parent_number)
                    if not parent_contract:
                        parent_contract = Contract.objects.filter(contract_number=parent_number).first()
                    if not parent_contract:
                        result.skipped_invalid += 1
                        continue
                    if contract_type is None:
                        contract_type = parent_contract.contract_type

                if contract_type is None:
                    result.skipped_invalid += 1
                    continue

                defaults = contract_defaults_from_excel(row, contract_type, parent_contract)
                if not defaults["contract_date"]:
                    result.skipped_invalid += 1
                    continue

                contract = imported_contracts.get(contract_number)
                if contract is None:
                    result.created += 1
                    contract = Contract.objects.create(
                        contract_number=contract_number,
                        **defaults,
                    )
                    imported_contracts[contract_number] = contract
                    continue

                changed_fields = update_contract_from_defaults(contract, defaults)
                if changed_fields:
                    result.updated += 1
                    contract.save(update_fields=[*changed_fields, "updated_at"])
                else:
                    result.unchanged += 1

        for row_index, row in excel_rows(parties_ws):
            contract_number = normalize_contract_number(row.get("contract_number"))
            if not contract_number:
                continue
            result.parties_rows += 1
            contract = imported_contracts.get(contract_number)
            if not contract:
                result.parties_skipped += 1
                continue
            try:
                sif_par = int(float(str(row.get("sif_par")).strip()))
            except (TypeError, ValueError):
                result.parties_skipped += 1
                continue
            partner = Partner.objects.filter(external_sif_par=sif_par).order_by("id").first()
            if not partner:
                result.parties_skipped += 1
                continue
            role = clean_text(row.get("role")) or "ostalo"
            note = clean_text(row.get("note"))
            existing_party = ContractParty.objects.filter(
                contract=contract,
                partner=partner,
                role=role,
            ).first()
            if existing_party:
                result.parties_unchanged += 1
                continue
            result.parties_created += 1
            ContractParty.objects.create(
                contract=contract,
                partner=partner,
                role=role,
                note=note,
            )

        if not commit:
            transaction.set_rollback(True)

    return result

from argparse import Namespace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from .models import Menica, UlaznaMenica


FIELD_MAP = {
    "Naziv duznika": "naziv_duznika",
    "Maticni broj duznika": "maticni_broj_duznika",
    "Poreski broj duznika": "poreski_broj_duznika",
    "Strana rezultata": "strana_rezultata",
    "Serijski broj menice": "serijski_broj_menice",
    "Datum izdavanja": "datum_izdavanja",
    "Iznos menice": "iznos_menice",
    "Valuta menice": "valuta_menice",
    "Datum dospeca": "datum_dospeca",
    "Izdavalac menice": "izdavalac_menice",
    "Vrsta menice": "vrsta_menice",
    "Redni broj": "redni_broj",
    "Osnov izdavanja": "osnov_izdavanja",
    "Iznos iz osnova": "iznos_iz_osnova",
    "Valuta osnova": "valuta_osnova",
    "Datum registracije": "datum_registracije",
    "Naziv banke": "naziv_banke",
    "Status": "status",
    "Avalisti detalji": "avalisti_detalji",
    "Avalisti broj zapisa": "avalisti_broj_zapisa",
}

DATE_FIELDS = {"datum_izdavanja", "datum_dospeca", "datum_registracije"}
DECIMAL_FIELDS = {"iznos_menice", "iznos_iz_osnova"}
INTEGER_FIELDS = {"strana_rezultata", "redni_broj", "avalisti_broj_zapisa"}
NBS_FIELDS = set(FIELD_MAP.values())
MANUAL_EXCEL_FIELD_MAP = {
    "Broj ugovora": "broj_ugovora",
    "Datum ugovora": "datum_ugovora",
    "OJ": "oj",
    "Napomena": "napomena",
    "Status (0-aktivna, 1-obrisana,2-nepoznat)": "interni_status",
}
EXCEL_KEY_FIELD_MAP = {
    "Serijski broj menice": "serijski_broj_menice",
    "Серијски број менице": "serijski_broj_menice",
    "Datum registracije": "datum_registracije",
    "Датум регистрације": "datum_registracije",
    "Redni broj": "redni_broj",
    "Редни број": "redni_broj",
}
ULAZNE_EXCEL_FIELD_MAP = {
    "Serijski broj menice": "serijski_broj_menice",
    "Serijski broj menice+A1:H39": "serijski_broj_menice",
    "Osnov izdavanja": "osnov_izdavanja",
    "Datum prijema menice": "datum_prijema_menice",
    "Procenat - iznos": "procenat_iznos",
    "Sifra poslovnog partnera": "sifra_poslovnog_partnera",
    "Naziv pravnog lica": "naziv_pravnog_lica",
    "Broj naseg ugovora": "broj_naseg_ugovora",
    "Datum ugovora": "datum_ugovora",
    "Ugovor vazi do": "ugovor_vazi_do",
    "Lokacija menice, sifra centra": "sifra_centra",
}
ULAZNE_DATE_FIELDS = {"datum_prijema_menice", "datum_ugovora", "ugovor_vazi_do"}
ULAZNE_DECIMAL_FIELDS = {"procenat_iznos"}
ULAZNE_INTEGER_FIELDS = {"sifra_poslovnog_partnera"}


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if text.endswith(".") and text.count(".") >= 3:
        text = text.rstrip(".")
    for date_format in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def parse_decimal(value):
    if value in (None, ""):
        return None
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_integer(value):
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def normalize_text(value):
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def normalize_record(record):
    values = {"tip": Menica.TIP_IZLAZNA}
    for source_name, field_name in FIELD_MAP.items():
        raw_value = record.get(source_name, "")
        if field_name in DATE_FIELDS:
            values[field_name] = parse_date(raw_value)
        elif field_name in DECIMAL_FIELDS:
            values[field_name] = parse_decimal(raw_value)
        elif field_name in INTEGER_FIELDS:
            values[field_name] = parse_integer(raw_value)
        else:
            values[field_name] = str(raw_value).strip() or None
    return values


def _update_instance_fields(instance, values, allowed_fields, *, skip_empty=False):
    changed_fields = []
    for field_name in allowed_fields:
        if field_name not in values:
            continue
        value = values[field_name]
        if skip_empty and value in (None, ""):
            continue
        if getattr(instance, field_name) != value:
            setattr(instance, field_name, value)
            changed_fields.append(field_name)
    return changed_fields


def _find_izlazna_menica(values):
    qs = Menica.objects.filter(
        tip=Menica.TIP_IZLAZNA,
        serijski_broj_menice=values.get("serijski_broj_menice"),
        datum_registracije=values.get("datum_registracije"),
    )
    redni_broj = values.get("redni_broj")
    if redni_broj is not None:
        qs = qs.filter(redni_broj=redni_broj)
    return qs.order_by("id").first()


def sync_izlazne_menice(
    *,
    tax_code="100223617",
    national_code="",
    serial_number="",
    registration_date="",
    page_size=100,
    include_avalists=True,
    max_pages=None,
    timeout=30,
):
    from .scraper import scrape

    args = Namespace(
        tax_code=tax_code,
        national_code=national_code,
        serial_number=serial_number,
        registration_date=registration_date,
        page_size=page_size,
        include_avalists=include_avalists,
        max_pages=max_pages,
        timeout=timeout,
    )
    records = scrape(args)
    created_count = 0
    updated_count = 0
    unchanged_count = 0
    skipped_count = 0

    for record in records:
        values = normalize_record(record)
        serial = values.get("serijski_broj_menice")
        registration = values.get("datum_registracije")
        if not serial or not registration:
            skipped_count += 1
            continue

        menica = _find_izlazna_menica(values)
        if menica:
            changed_fields = _update_instance_fields(menica, values, NBS_FIELDS)
            if changed_fields:
                menica.save(update_fields=[*changed_fields, "updated_at"])
                updated_count += 1
            else:
                unchanged_count += 1
            continue

        Menica.objects.create(**values)
        created_count += 1

    return {
        "fetched": len(records),
        "created": created_count,
        "updated": updated_count,
        "unchanged": unchanged_count,
        "skipped": skipped_count,
    }


def _excel_header_map(worksheet, header_row):
    headers = {}
    for column_index in range(1, worksheet.max_column + 1):
        value = normalize_text(worksheet.cell(header_row, column_index).value)
        if value:
            headers[value] = column_index
    return headers


def _find_header(headers, expected_header):
    if expected_header in headers:
        return headers[expected_header]
    for header, column_index in headers.items():
        if header.startswith(expected_header):
            return column_index
    return None


def _excel_row_values(worksheet, row_index, headers):
    values = {}
    for header, field_name in EXCEL_KEY_FIELD_MAP.items():
        column_index = headers.get(header)
        if not column_index:
            continue
        raw_value = worksheet.cell(row_index, column_index).value
        if field_name == "datum_registracije":
            values[field_name] = parse_date(raw_value)
        elif field_name == "redni_broj":
            values[field_name] = parse_integer(raw_value)
        else:
            values[field_name] = normalize_text(raw_value)

    for header, field_name in MANUAL_EXCEL_FIELD_MAP.items():
        column_index = headers.get(header)
        if not column_index:
            continue
        raw_value = worksheet.cell(row_index, column_index).value
        if field_name == "datum_ugovora":
            values[field_name] = parse_date(raw_value)
        elif field_name == "interni_status":
            status = parse_integer(raw_value)
            if status in {
                Menica.STATUS_AKTIVNA,
                Menica.STATUS_OBRISANA,
                Menica.STATUS_NEPOZNAT,
            }:
                values[field_name] = status
        else:
            values[field_name] = normalize_text(raw_value)
    return values


def import_izlazne_menice_manual_fields_from_excel(
    file_path,
    *,
    sheet_name="Izlazne menice",
    header_row=2,
    first_data_row=3,
    commit=False,
):
    from openpyxl import load_workbook

    workbook_path = Path(file_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' ne postoji u fajlu {workbook_path}.")

    worksheet = workbook[sheet_name]
    headers = _excel_header_map(worksheet, header_row)
    required_headers = [
        "Серијски број менице",
        "Датум регистрације",
        *MANUAL_EXCEL_FIELD_MAP.keys(),
    ]
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        raise ValueError("Nedostaju kolone u Excel-u: " + ", ".join(missing_headers))

    result = {
        "rows": 0,
        "matched": 0,
        "updated": 0,
        "unchanged": 0,
        "skipped_missing_key": 0,
        "skipped_not_found": 0,
        "commit": commit,
    }

    manual_fields = set(MANUAL_EXCEL_FIELD_MAP.values())
    for row_index in range(first_data_row, worksheet.max_row + 1):
        result["rows"] += 1
        values = _excel_row_values(worksheet, row_index, headers)
        serial = values.get("serijski_broj_menice")
        registration = values.get("datum_registracije")
        if not serial or not registration:
            result["skipped_missing_key"] += 1
            continue

        menica = _find_izlazna_menica(values)
        if not menica:
            result["skipped_not_found"] += 1
            continue

        result["matched"] += 1
        changed_fields = _update_instance_fields(
            menica,
            values,
            manual_fields,
            skip_empty=True,
        )
        if changed_fields:
            result["updated"] += 1
            if commit:
                menica.save(update_fields=[*changed_fields, "updated_at"])
        else:
            result["unchanged"] += 1

    return result


def _ulazna_excel_row_values(worksheet, row_index, headers):
    values = {}
    for header, field_name in ULAZNE_EXCEL_FIELD_MAP.items():
        column_index = _find_header(headers, header)
        if not column_index:
            continue
        raw_value = worksheet.cell(row_index, column_index).value
        if field_name in ULAZNE_DATE_FIELDS:
            values[field_name] = parse_date(raw_value)
        elif field_name in ULAZNE_DECIMAL_FIELDS:
            values[field_name] = parse_decimal(raw_value)
        elif field_name in ULAZNE_INTEGER_FIELDS:
            values[field_name] = parse_integer(raw_value)
        else:
            values[field_name] = normalize_text(raw_value)
    return values


def import_ulazne_menice_from_excel(
    file_path,
    *,
    sheet_name="ulazne menice",
    header_row=1,
    first_data_row=2,
    commit=False,
):
    from openpyxl import load_workbook

    workbook_path = Path(file_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' ne postoji u fajlu {workbook_path}.")

    worksheet = workbook[sheet_name]
    headers = _excel_header_map(worksheet, header_row)
    required_headers = ["Serijski broj menice", "Datum prijema menice"]
    missing_headers = [
        header for header in required_headers if _find_header(headers, header) is None
    ]
    if missing_headers:
        raise ValueError("Nedostaju kolone u Excel-u: " + ", ".join(missing_headers))

    result = {
        "rows": 0,
        "created": 0,
        "updated": 0,
        "unchanged": 0,
        "skipped_missing_key": 0,
        "commit": commit,
    }
    allowed_fields = set(ULAZNE_EXCEL_FIELD_MAP.values())

    for row_index in range(first_data_row, worksheet.max_row + 1):
        result["rows"] += 1
        values = _ulazna_excel_row_values(worksheet, row_index, headers)
        serial = values.get("serijski_broj_menice")
        if not serial:
            result["skipped_missing_key"] += 1
            continue

        menica = (
            UlaznaMenica.objects.filter(serijski_broj_menice=serial)
            .order_by("id")
            .first()
        )
        if menica is None:
            result["created"] += 1
            if commit:
                UlaznaMenica.objects.create(**values)
            continue

        changed_fields = _update_instance_fields(menica, values, allowed_fields)
        if changed_fields:
            result["updated"] += 1
            if commit:
                menica.save(update_fields=[*changed_fields, "updated_at"])
        else:
            result["unchanged"] += 1

    return result
